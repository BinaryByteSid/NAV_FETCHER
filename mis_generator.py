"""
MIS Generator Module for Mutual Fund Portfolios.

Produces the three standard MIS views:

  MIS 1 — portfolio vs Nifty 50, for the day and for the period, plus Flows,
          MTD and YTD (financial-year-to-date) columns.
  MIS 2 — portfolio vs each scheme's own benchmark, for the day and the period.
  MIS 3 — performance since 1 April (financial year start) vs both Nifty and
          each scheme's own benchmark.

Every figure traces to real data. Scheme NAVs come from AMFI; benchmark index
levels come from benchmark_proxy.py, which maps each benchmark to a passive
index fund and reads its NAV history from the same AMFI feed. Where a benchmark
cannot be sourced the report prints "N/A" — it never substitutes a value.

Returns are aligned to actual observation dates rather than calendar dates, so
a market holiday cannot silently collapse a one-day return to 0.00%.
"""

from __future__ import annotations

import io
import re
from datetime import datetime, date, timedelta
from typing import Dict, List, Tuple, Any, Optional

import pandas as pd
import streamlit as st

# openpyxl styling imports
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Re-use existing NAV fetch services from parent codebase
from nav_fetcher import (
    fetch_amfi_data_chunked,
    fetch_latest_navs,
    parse_amfi_date_series,
    populate_actual_aum,
    load_portfolio_aum_data,
    calculate_flows_for_dataframe,
    _parse_amfi_date_str,
)
from benchmark_proxy import (
    required_proxy_isins,
    build_benchmark_series,
    parse_supplied_levels,
    apply_supplied_levels,
    apply_tri_levels,
    nav_frame_to_isin_series,
    overlay_live_navs,
    describe_gaps,
    describe_resolution,
    fetch_nifty_price_index,
    BenchmarkSeries,
    STATUS_EXACT,
    STATUS_APPROX,
    STATUS_NONE,
    DEFAULT_BENCHMARK,
)
import mis_history
from ui_theme import render_section_header, render_info_card, finance_panel

# ─── Default Sample Portfolio ──────────────────────────────────────────────────

DEFAULT_SAMPLE_PORTFOLIO = [
    {"Scheme Name": "Invesco India Smallcap Reg Gr",   "ISIN": "INF205K011T7", "Allocation (%)": 7.0,  "Benchmark": "S&P BSE 250 SmallCap TR INR"},
    {"Scheme Name": "Kotak Midcap Reg Gr",             "ISIN": "INF174K01DS9", "Allocation (%)": 6.0,  "Benchmark": "Nifty Midcap 150 TR INR"},
    {"Scheme Name": "Quant Large Cap Reg Gr",          "ISIN": "INF966L01AW4", "Allocation (%)": 10.0, "Benchmark": "Nifty 100 TR INR"},
    {"Scheme Name": "SBI Infrastructure Reg Gr",       "ISIN": "INF200K01CT2", "Allocation (%)": 6.0,  "Benchmark": "Nifty Infrastructure TR INR"},
    {"Scheme Name": "Invesco India Focused Reg Gr",    "ISIN": "INF205KA1189", "Allocation (%)": 7.0,  "Benchmark": "BSE 500 India TR INR"},
    {"Scheme Name": "SBI Large & Midcap Reg Gr",       "ISIN": "INF200K01305", "Allocation (%)": 6.0,  "Benchmark": "Nifty LargeMidcap 250 TR INR"},
    {"Scheme Name": "Bandhan Large & Mid Cap Gr",      "ISIN": "INF194K01524", "Allocation (%)": 7.0,  "Benchmark": "Nifty LargeMidcap 250 TR INR"},
    {"Scheme Name": "ICICI Pru Focused Equity Gr",     "ISIN": "INF109K01BZ4", "Allocation (%)": 7.0,  "Benchmark": "BSE 500 India TR INR"},
    {"Scheme Name": "Kotak Multicap Reg Gr",           "ISIN": "INF174KA1HS9", "Allocation (%)": 7.0,  "Benchmark": "Nifty 500 Multicap 50:25:25 TR INR"},
    {"Scheme Name": "Canara Robeco Multi Cap Reg Gr",  "ISIN": "INF760K01KR2", "Allocation (%)": 7.0,  "Benchmark": "Nifty 500 Multicap 50:25:25 TR INR"},
    {"Scheme Name": "DSP Large & Mid Cap Fund Reg Gr", "ISIN": "INF740K01094", "Allocation (%)": 9.0,  "Benchmark": "Nifty LargeMidcap 250 TR INR"},
    {"Scheme Name": "HDFC Flexi Cap Gr",               "ISIN": "INF179K01608", "Allocation (%)": 9.0,  "Benchmark": "Nifty 500 TR INR"},
    {"Scheme Name": "ICICI Pru Dividend Yield Eq Gr",  "ISIN": "INF109KA1TX4", "Allocation (%)": 7.0,  "Benchmark": "Nifty 500 TR INR"},
    {"Scheme Name": "HDFC Small Cap Gr",               "ISIN": "INF179KA1RZ8", "Allocation (%)": 5.0,  "Benchmark": "S&P BSE 250 SmallCap TR INR"},
]

# Reports carry full working precision in the cell and show two decimals.
# Rounding the stored value to two threw the precision away in the file, so a
# figure could not be re-derived or reconciled from the workbook afterwards;
# the number format governs display only.
STORE_DP = 5
DISPLAY_FMT = "0.00"

NIFTY_KEY = "NIFTY 50"


# ─── Helper Date Functions ────────────────────────────────────────────────────

def get_fy_start_date(target_date: date) -> date:
    """April 1st of the financial year containing target_date."""
    if target_date.month >= 4:
        return date(target_date.year, 4, 1)
    return date(target_date.year - 1, 4, 1)


def get_mtd_start_date(target_date: date) -> date:
    """Last day of the previous calendar month — the MTD baseline."""
    return date(target_date.year, target_date.month, 1) - timedelta(days=1)


# ─── Observation series helpers ───────────────────────────────────────────────
# A "series" here is {observation_date: value}. Both scheme NAVs and benchmark
# index levels use this shape so one set of helpers serves both.

Series = Dict[date, float]


SKIP_WEEKEND_MARKS = True  # module-level switch, set from the UI each run


def _is_skipped_day(d: date) -> bool:
    """True for days whose observations must not anchor a return.

    Indian equity markets are shut on Saturdays, but some AMCs still stamp a
    NAV on them. Anchoring a return to a Saturday mark measures the AMC's
    bookkeeping rather than a market move, so those observations are passed
    over. Sunday is excluded for the same reason.
    """
    if not SKIP_WEEKEND_MARKS:
        return False
    return d.weekday() >= 5  # 5 = Saturday, 6 = Sunday


def value_asof(series: Optional[Series], target: date) -> Tuple[Optional[float], Optional[date]]:
    """Latest weekday value on or before ``target``, with the date it came from."""
    if not series:
        return None, None
    candidates = [d for d in series if d <= target and not _is_skipped_day(d)]
    if not candidates:
        # Fall back to any observation rather than reporting N/A outright: a
        # weekend-only series is still better than nothing, and the caller
        # flags a stale anchor separately.
        candidates = [d for d in series if d <= target]
    if not candidates:
        return None, None
    d = max(candidates)
    return series[d], d


def value_prev(series: Optional[Series], before: date) -> Tuple[Optional[float], Optional[date]]:
    """Latest value strictly before ``before``.

    Used for one-day returns. Resolving both ends with "on or before" would let
    a holiday pick the same observation twice and report a 0.00% move that never
    happened.
    """
    if not series:
        return None, None
    candidates = [d for d in series if d < before and not _is_skipped_day(d)]
    if not candidates:
        candidates = [d for d in series if d < before]
    if not candidates:
        return None, None
    d = max(candidates)
    return series[d], d


def _num(v: Any) -> Optional[float]:
    """Coerce to float, mapping None and NaN alike to None.

    Pandas hands back NaN rather than None for missing cells, and `NaN is None`
    is False — so a plain None check lets NaN through and it then contaminates
    every sum and average it touches.
    """
    if v is None:
        return None
    try:
        f = float(v)
    except (TypeError, ValueError):
        return None
    return None if pd.isna(f) else f


def calc_return(end_val: Optional[float], start_val: Optional[float]) -> Optional[float]:
    """Non-annualised return %: ((end - start) / start) * 100."""
    e, s = _num(end_val), _num(start_val)
    if e is None or s is None or s <= 0:
        return None
    return ((e - s) / s) * 100.0


def _diff(a: Optional[float], b: Optional[float]) -> Optional[float]:
    """a - b, propagating missing values so a gap never becomes a fake 0."""
    x, y = _num(a), _num(b)
    if x is None or y is None:
        return None
    return x - y


# ─── Data Ingestion & Validation ──────────────────────────────────────────────

def _is_percent_format(fmt: Any) -> bool:
    """True if an Excel number format displays its value as a percentage.

    A '%' inside a quoted literal or escaped with a backslash is just text --
    only a bare '%' makes Excel scale the stored value by 100.
    """
    if not fmt:
        return False
    stripped = re.sub(r'"[^"]*"', "", str(fmt))
    stripped = re.sub(r"\\.", "", stripped)
    return "%" in stripped


def read_portfolio_excel(file_obj) -> pd.DataFrame:
    """Read an uploaded portfolio workbook, honouring Excel percent formatting.

    Excel stores a cell displaying 7.00% as 0.07, so the stored number alone
    is ambiguous: 0.07 could mean 7% or 0.07%. The cell's number format says
    which, and reading it means a cell holding 7 is always taken as 7 -- no
    inference from column totals.
    """
    df = pd.read_excel(file_obj)
    if df.empty:
        return df

    try:
        file_obj.seek(0)
        ws = openpyxl.load_workbook(file_obj, data_only=True).worksheets[0]
    except Exception:
        # .xls, or a workbook openpyxl cannot parse. Values stand as read.
        return df

    # Map each header to its column, then scale the columns Excel is
    # displaying as percentages.
    header_at = {}
    for cell in ws[1]:
        if cell.value is not None:
            header_at.setdefault(str(cell.value).strip(), cell.column)

    for col_name in df.columns:
        col_idx = header_at.get(str(col_name).strip())
        if col_idx is None or not pd.api.types.is_numeric_dtype(df[col_name]):
            continue
        formats = [
            ws.cell(row=r, column=col_idx).number_format
            for r in range(2, min(ws.max_row, len(df) + 1) + 1)
            if isinstance(ws.cell(row=r, column=col_idx).value, (int, float))
        ]
        if formats and all(_is_percent_format(f) for f in formats):
            # Round away the binary-float artifact: 0.07 * 100 is
            # 7.000000000000001.
            df[col_name] = (df[col_name] * 100.0).round(6)

    return df


_PORTFOLIO_KEY_COLS = ("Scheme Name", "ISIN", "Allocation (%)", "Benchmark")


def _report_input_signature(current, previous, d_start, d_end, m3_start, m3_end,
                            include_flows, skip_saturdays) -> tuple:
    """Everything that changes a report's contents, in one comparable value.

    Used to tell the user their downloads are stale, since the buttons render
    from the last generated result rather than from what is currently on screen.
    """
    return (
        _portfolio_signature(current),
        _portfolio_signature(previous),
        d_start, d_end, m3_start, m3_end, bool(include_flows), bool(skip_saturdays),
    )


def _portfolio_signature(df: Optional[pd.DataFrame]) -> tuple:
    """Order-insensitive fingerprint of a portfolio, for change detection.

    Compares only the four fields that define a holding, as strings, so a
    Streamlit rerun that reorders rows or re-types a number does not read as an
    edit. Reordering alone is not a portfolio change, so rows are sorted.
    """
    if df is None or df.empty:
        return ()
    cols = [c for c in _PORTFOLIO_KEY_COLS if c in df.columns]
    if not cols:
        return ()
    rows = []
    for _, r in df[cols].iterrows():
        cells = []
        for c in cols:
            v = r[c]
            if isinstance(v, float):
                cells.append(f"{v:.6f}")          # 7.0 and 7.000001 are the same holding
            else:
                cells.append(str(v).strip().upper())
        if any(cell not in ("", "NAN", "NONE") for cell in cells):
            rows.append(tuple(cells))
    return tuple(sorted(rows))


def validate_and_normalize_portfolio(df_input: pd.DataFrame) -> Tuple[pd.DataFrame, List[str], List[str]]:
    """Validate portfolio input and prepare normalized allocations.

    Returns (cleaned_df, warnings, errors).
    """
    warnings: List[str] = []
    errors: List[str] = []

    if df_input is None or df_input.empty:
        errors.append("Portfolio input is empty. Please enter schemes manually or upload an Excel file.")
        return pd.DataFrame(), warnings, errors

    df = df_input.copy()

    isin_col = name_col = alloc_col = bm_col = None
    for c in df.columns:
        c_low = str(c).lower().strip()
        if "isin" in c_low:
            isin_col = c
        elif "bench" in c_low or c_low == "bm":
            bm_col = c
        elif "alloc" in c_low or "weight" in c_low or c_low == "wt" or "%" in c_low:
            alloc_col = c
        elif "scheme" in c_low or "fund" in c_low or "name" in c_low:
            name_col = c

    if not isin_col:
        errors.append("Missing required column 'ISIN' in input.")
        return pd.DataFrame(), warnings, errors

    clean_rows = []
    seen_isins = set()

    for _, row in df.iterrows():
        isin = str(row.get(isin_col, "")).strip().upper()
        if not isin or isin in ("NAN", "NONE", "-"):
            continue

        if isin in seen_isins:
            warnings.append(f"Duplicate ISIN '{isin}' detected. Only the first occurrence was retained.")
            continue
        seen_isins.add(isin)

        name = str(row.get(name_col, isin)).strip() if name_col else isin
        if not name or name.upper() in ("NAN", "NONE"):
            name = isin

        alloc_raw = row.get(alloc_col, 0.0) if alloc_col else 0.0
        try:
            alloc_val = float(str(alloc_raw).replace("%", "").strip())
        except (TypeError, ValueError):
            alloc_val = 0.0

        bm_val = str(row.get(bm_col, DEFAULT_BENCHMARK)).strip() if bm_col else DEFAULT_BENCHMARK
        if not bm_val or bm_val.upper() in ("NAN", "NONE", "-"):
            bm_val = DEFAULT_BENCHMARK

        clean_rows.append({
            "Scheme Name": name,
            "ISIN": isin,
            "Allocation (%)": alloc_val,
            "Benchmark": bm_val,
        })

    if not clean_rows:
        errors.append("No valid ISIN rows found in the portfolio input.")
        return pd.DataFrame(), warnings, errors

    res_df = pd.DataFrame(clean_rows)

    tot_alloc = res_df["Allocation (%)"].sum()
    if tot_alloc <= 0:
        warnings.append("Total allocation was 0%. Weights have been distributed equally.")
        res_df["Allocation (%)"] = 100.0 / len(res_df)
        tot_alloc = 100.0

    if abs(tot_alloc - 100.0) > 0.01:
        warnings.append(
            f"Total allocation sum is {tot_alloc:.2f}% (not 100%). Weights have been normalized to 100%."
        )

    res_df["Normalized_Weight"] = res_df["Allocation (%)"] / tot_alloc
    return res_df, warnings, errors


# ─── NAV Fetching ─────────────────────────────────────────────────────────────

def fetch_mis_nav_history(isin_list: List[str], earliest_date: date,
                          end_date: date, progress=None) -> Tuple[pd.DataFrame, List[Tuple[date, date]]]:
    """Fetch historical NAVs for all ISINs, overlaid with the live AMFI feed.

    ``isin_list`` carries both the portfolio schemes and the benchmark proxy
    funds so the whole report costs one pass over AMFI. include_direct is on
    for the proxies' sake; an ISIN belongs to exactly one plan, so scheme rows
    are unaffected.

    Returns (frame, gaps) — gaps are date windows AMFI refused to serve, which
    the caller must report rather than silently treat as "no NAVs existed".
    """
    fetch_start = earliest_date - timedelta(days=20)
    df_raw, gaps = fetch_amfi_data_chunked(
        fetch_start, end_date, isin_list, include_direct=True, return_gaps=True,
        progress=progress,
        # Smaller windows than the 90-day default: each AMFI call blocks the
        # script, and a long silence lets the Spaces proxy drop the websocket
        # mid-report. Shorter calls keep the progress bar talking.
        chunk_days=45,
    )
    if df_raw.empty:
        return pd.DataFrame(), gaps

    df_raw = df_raw.copy()
    df_raw["NAV Date"] = parse_amfi_date_series(df_raw["NAV Date"])
    df_raw = df_raw.dropna(subset=["NAV Date", "NAV"])
    df_raw["ISIN_G"] = df_raw["ISIN Div Payout / ISIN Growth"].astype(str).str.strip().str.upper()
    df_raw["ISIN_R"] = df_raw["ISIN Div Reinvestment"].astype(str).str.strip().str.upper()
    df_raw["NAV Date_Date"] = df_raw["NAV Date"].dt.date

    latest_live = fetch_latest_navs(isin_list) or {}
    live_rows = []
    for isin_u, info in latest_live.items():
        parsed_iso = _parse_amfi_date_str(info.get("date", ""))
        if not parsed_iso:
            continue
        try:
            l_dt = datetime.strptime(parsed_iso, "%Y-%m-%d").date()
        except ValueError:
            continue
        # A stale-dated live record must not leak past the report's end date.
        if l_dt > end_date:
            continue
        live_rows.append({
            "ISIN_G": isin_u,
            "ISIN_R": isin_u,
            "Scheme Name": info.get("scheme_name", isin_u),
            "NAV": info["nav"],
            "NAV Date_Date": l_dt,
            "NAV Date": pd.Timestamp(l_dt),
        })

    if live_rows:
        df_raw = pd.concat([df_raw, pd.DataFrame(live_rows)], ignore_index=True)

    return df_raw, gaps


def build_nav_series(df_nav_raw: pd.DataFrame, isin_list: List[str]) -> Dict[str, Series]:
    """Collapse the raw NAV frame into {ISIN: {date: nav}}."""
    out: Dict[str, Series] = {i.strip().upper(): {} for i in isin_list}
    if df_nav_raw.empty:
        return out

    wanted = set(out.keys())
    for _, r in df_nav_raw.iterrows():
        d = r.get("NAV Date_Date")
        nav = r.get("NAV")
        if d is None or pd.isna(d) or pd.isna(nav):
            continue
        try:
            nav_f = float(nav)
        except (TypeError, ValueError):
            continue
        if nav_f <= 0:
            continue
        for col in ("ISIN_G", "ISIN_R"):
            isin = r.get(col)
            if isin in wanted:
                out[isin][d] = nav_f
    return out


# ─── Flows ────────────────────────────────────────────────────────────────────

# Quant publishes its AUM a day ahead of the rest of the industry, so the usual
# pairing measures the wrong interval for their schemes. AMFI names every one of
# their schemes 'quant <fund>', so a prefix test is enough and will not catch an
# unrelated fund with 'quant' elsewhere in its name.


def _align_quant_nav(df_aum: pd.DataFrame) -> pd.DataFrame:
    """Pair Quant AMC's AUM with the return that actually spans it.

    Quant publishes AUM a day ahead of the rest of the industry: the figure
    stamped for day D already reflects day D+1's book. The flow formula
    AUM_t - AUM_(t-1) * (1 + return) then charges the wrong day's
    mark-to-market against the pair, which is why Quant Large Cap reported a
    51.15cr flow where the reference shows 2.04.

    Taking the reported AUM as the previous day's and the next day's as the
    current one -- with the return otherwise normal -- is the same as pairing
    the unchanged AUM column with the *preceding* day's return. Shifting the
    NAV column forward one observation does exactly that, and leaves the AUM
    figures themselves untouched. Other AMCs are unaffected.
    """
    if df_aum.empty or "NAV" not in df_aum.columns or "Scheme Name" not in df_aum.columns:
        return df_aum

    is_quant = df_aum["Scheme Name"].astype(str).str.lower().str.startswith("quant")
    if not is_quant.any():
        return df_aum

    date_col = "NAV Date" if "NAV Date" in df_aum.columns else None
    if date_col is None:
        return df_aum

    out = df_aum.copy()
    for _isin, grp in out[is_quant].groupby("ISIN Div Payout / ISIN Growth"):
        grp = grp.sort_values(date_col)
        shifted = grp["NAV"].shift(1)
        # The first observation has no predecessor; keep its own NAV so the
        # opening day is simply not counted rather than becoming NaN.
        shifted.iloc[0] = grp["NAV"].iloc[0]
        out.loc[grp.index, "NAV"] = shifted.values

    return out


def compute_scheme_flows(df_nav_raw: pd.DataFrame, isin_list: List[str],
                         report_date: date, mtd_start: date, fy_start: date,
                         notes: Optional[List[str]] = None) -> Dict[str, Dict[str, Optional[float]]]:
    """Net flows per ISIN, via the project's existing AUM engine.

    Reuses populate_actual_aum + calculate_flows_for_dataframe from nav_fetcher
    so the figures match what the rest of the app reports. Returns
    {ISIN: {"day": ..., "mtd": ..., "ytd": ...}} in crores, where "day" is the
    net flow on the flow date and the other two are cumulative net flows over
    the month-to-date and financial-year-to-date windows.

    AUM is published a day behind NAV, so every flow figure stops at the last
    trading day *before* report_date -- a report dated 23 July carries flows to
    22 July, which is how the reference MIS labels the column. Returns
    (flows, flow_date); flow_date is None when no such day exists.
    """
    empty = {"day": None, "mtd": None, "ytd": None}
    result: Dict[str, Dict[str, Optional[float]]] = {
        i.strip().upper(): dict(empty) for i in isin_list
    }
    if df_nav_raw.empty or "Scheme Code" not in df_nav_raw.columns:
        return result, None

    # Flows accumulate day by day, so the whole FY-to-date window is needed —
    # plus one earlier observation to seed the first day's prior AUM.
    dates_present = sorted({d for d in df_nav_raw["NAV Date_Date"]
                            if d is not None and d <= report_date})
    if len(dates_present) < 2:
        return result, None

    # AUM lags NAV by a day, so the report date's own flow is not yet
    # published. Stop at the last trading day strictly before it.
    earlier = [d for d in dates_present if d < report_date]
    if not earlier:
        return result, None
    flow_date = max(earlier)
    dates_present = [d for d in dates_present if d <= flow_date]
    if len(dates_present) < 2:
        return result, None

    # One day earlier than the first day being counted, so the seed observation
    # supplies the prior-day AUM that 1 April's own flow is measured against.
    window_start = min(fy_start - timedelta(days=1), mtd_start)
    seed = [d for d in dates_present if d <= window_start]
    keep = {d for d in dates_present if d > window_start}
    if seed:
        keep.add(max(seed))
    if len(keep) < 2:
        return result, None

    df_slice = df_nav_raw[df_nav_raw["NAV Date_Date"].isin(keep)].copy()
    df_slice = df_slice.dropna(subset=["Scheme Code"])
    if df_slice.empty:
        return result, None

    partial: List[str] = []

    def _note_partial(done: int, total: int) -> None:
        partial.append(
            f"Live AUM was retrieved for {done} of {total} day/category combinations; the rest fall "
            f"back to a derived AUM, so Flows, MTD and YTD are approximate."
        )

    try:
        df_port = load_portfolio_aum_data()
        if df_port.empty and notes is not None:
            notes.append(
                "The fallback AUM workbook ('portfolio last 6 months.xlsx') was not found, so every "
                "AUM figure depends on the live AMFI API; any day it declines leaves that scheme's "
                "Flows, MTD and YTD blank rather than approximate."
            )
        # One API call per day/category: a financial year is hundreds of them.
        # Bound the wall time and parallelise modestly rather than letting the
        # report hang -- partial live AUM plus a warning beats no report.
        df_aum = populate_actual_aum(
            df_slice, df_port, want_aum=True, fetch_live_aum=True,
            budget_seconds=AUM_BUDGET_SECONDS, max_workers=AUM_FETCH_WORKERS,
            on_incomplete=_note_partial,
            # Flows measure AUM movement between consecutive days. A carried
            # figure is indistinguishable from AMFI repeating a stale one, so
            # every carried day would be scored as a zero flow.
            gap_fill="fallback",
        )
        df_aum = _align_quant_nav(df_aum)
        df_flows = calculate_flows_for_dataframe(
            df_aum,
            min(keep),
            ["Asset Class", "Scheme Code", "ISIN Div Payout / ISIN Growth",
             "ISIN Div Reinvestment", "Scheme Name", "Plan Type", "Option Type"],
        )
    except Exception:
        # Flows are supplementary; a failure here must not sink the whole report.
        return result, flow_date

    if notes is not None:
        notes.extend(partial)

    if df_flows.empty or "Net flows on current day" not in df_flows.columns:
        return result, flow_date

    # AMFI sometimes repeats yesterday's AUM verbatim. The flow formula is
    # AUM_t - AUM_(t-1) * (1 + return), so an unchanged AUM books the whole
    # mark-to-market move as a subscription: Kotak Midcap showed a 913cr
    # "flow" purely because a 69,849cr AUM was served twice while the NAV fell
    # 1.3%. An AUM that does not move at all while the NAV does is
    # arithmetically impossible, so treat those days as missing, not as flows.
    stale_days = 0
    stale_names: set = set()

    missing_aum_days = 0

    for _, r in df_flows.iterrows():
        flow = r.get("Net flows on current day")
        # No AUM for the day means no flow can be measured. Book zero rather
        # than dropping the day, so a scheme with a patchy feed still reports a
        # number instead of N/A -- at the cost of understating a real flow.
        flow_missing = flow is None or pd.isna(flow)
        if flow_missing:
            missing_aum_days += 1

        prev_aum = pd.to_numeric(r.get("Closing AUM as on previous day"), errors="coerce")
        curr_aum = pd.to_numeric(r.get("Actual AUM as on current date"), errors="coerce")
        day_ret = pd.to_numeric(r.get("Daily return"), errors="coerce")
        # An AUM that does not move while the NAV does is a feed that has not
        # been updated, not a subscription. The flow formula would otherwise
        # book the entire mark-to-market move as one: Kotak Midcap showed a
        # 913cr "flow" purely because a 69,849cr AUM was served twice while the
        # NAV fell 1.3%. Report those days as a flow of zero.
        aum_stale = (
            not flow_missing
            and pd.notna(prev_aum) and pd.notna(curr_aum) and pd.notna(day_ret)
            and prev_aum == curr_aum
            and abs(float(day_ret)) > STALE_AUM_MIN_MOVE_PCT
        )
        if aum_stale:
            stale_days += 1
            stale_names.add(str(r.get("Scheme Name", "")).strip())
        try:
            d = datetime.strptime(str(r.get("NAV Date", "")).strip(), "%d-%m-%Y").date()
        except ValueError:
            continue
        if d > flow_date:
            continue

        for col in ("ISIN Div Payout / ISIN Growth", "ISIN Div Reinvestment"):
            isin = str(r.get(col, "")).strip().upper()
            if isin not in result:
                continue
            bucket = result[isin]
            val = 0.0 if (aum_stale or flow_missing) else float(flow)
            if d == flow_date:
                bucket["day"] = val
            # mtd_start is a baseline date (last day of the previous month), so
            # it is excluded. fy_start is the first day OF the year, so a flow
            # dated 1 April belongs inside it — hence >= rather than >.
            if d > mtd_start:
                bucket["mtd"] = (bucket["mtd"] or 0.0) + val
            if d >= fy_start:
                bucket["ytd"] = (bucket["ytd"] or 0.0) + val
            break

    if missing_aum_days and notes is not None:
        notes.append(
            f"AMFI served no AUM on {missing_aum_days} scheme-day(s); those days are "
            f"counted as zero flow, so Flows, MTD and YTD understate any real movement "
            f"on them."
        )

    if stale_days and notes is not None:
        shown = ", ".join(sorted(n for n in stale_names if n)[:4])
        more = f" and {len(stale_names) - 4} more" if len(stale_names) > 4 else ""
        notes.append(
            f"AMFI repeated the previous day's AUM on {stale_days} scheme-day(s) "
            f"({shown}{more}). Those days are counted as zero flow, since an unchanged "
            f"AUM against a moved NAV is a feed that has not updated rather than a "
            f"subscription. MTD and YTD therefore understate any real flow on those days."
        )

    return result, flow_date


# ─── Per-scheme metric computation ────────────────────────────────────────────

# A one-day return may legitimately span a long weekend or a cluster of
# holidays, but not much more. Beyond this the two observations are too far
# apart to be called a daily move.
MAX_DAILY_GAP_DAYS = 7

# Wall-clock cap on the live-AUM fetch behind the Flows column. A financial
# year spans ~78 trading days across ~8 equity categories, so a full book is
# ~600 calls; 90s only got through a third of them, and whichever categories
# lost the race came back N/A while their neighbours reported numbers.
AUM_BUDGET_SECONDS = 300.0
AUM_FETCH_WORKERS = 6

# A daily move smaller than this is too small to tell a stale AUM apart from a
# genuinely flat one, so staleness is only called above it.
STALE_AUM_MIN_MOVE_PCT = 0.05


def _window_return(series: Optional[Series], d_base: date, d_to: date) -> Optional[float]:
    """Return over (d_base, d_to], both ends resolved to real observations.

    d_base is the day *before* the window opens, so the opening day's own move
    is inside the window -- the same prior-close convention the period and FY
    figures use.
    """
    v_to, _ = value_asof(series, d_to)
    v_base, _ = value_asof(series, d_base)
    return calc_return(v_to, v_base)


def _scheme_metrics(series: Series, d_end: date, d_period_base: date,
                    d_fy_base: date, d_mtd_base: date) -> Dict[str, Any]:
    """All return figures for a single observation series."""
    v_end, dt_end = value_asof(series, d_end)
    v_prev, dt_prev = value_prev(series, dt_end) if dt_end else (None, None)
    v_period, dt_period = value_asof(series, d_period_base)
    v_fy, dt_fy = value_asof(series, d_fy_base)
    v_mtd, dt_mtd = value_asof(series, d_mtd_base)

    # Don't dress a multi-week move up as a daily one when the feed has a hole.
    day_gap = (dt_end - dt_prev).days if (dt_end and dt_prev) else None
    day_stale = day_gap is not None and day_gap > MAX_DAILY_GAP_DAYS
    day_ret = None if day_stale else calc_return(v_end, v_prev)

    return {
        "nav_end": v_end, "date_end": dt_end,
        "nav_prev": v_prev, "date_prev": dt_prev,
        "day": day_ret,
        "day_gap_days": day_gap,
        "day_stale": day_stale,
        "period": calc_return(v_end, v_period),
        "fy": calc_return(v_end, v_fy),
        "mtd": calc_return(v_end, v_mtd),
        "date_period_base": dt_period,
        "date_fy_base": dt_fy,
        "date_mtd_base": dt_mtd,
    }


def _weighted(values: List[Optional[float]], weights: List[float]) -> Tuple[Optional[float], float]:
    """Weighted average over entries that have a value.

    Returns (weighted_average, covered_weight). Weights are renormalized across
    the covered rows; covered_weight lets the caller flag partial coverage
    rather than let a gap quietly inflate the remaining holdings.
    """
    pairs = []
    for v, w in zip(values, weights):
        vn, wn = _num(v), _num(w)
        if vn is not None and wn is not None:
            pairs.append((vn, wn))
    covered = sum(w for _, w in pairs)
    if not pairs or covered <= 0:
        return None, 0.0
    return sum(v * w for v, w in pairs) / covered, covered


# ─── MIS report construction ──────────────────────────────────────────────────

def _ordinal(n: int) -> str:
    if 11 <= (n % 100) <= 13:
        return f"{n}th"
    return f"{n}{ {1: 'st', 2: 'nd', 3: 'rd'}.get(n % 10, 'th') }"


def _long_date(d: date) -> str:
    return f"{_ordinal(d.day)} {d:%B %Y}"


def _build_reports(portfolio_df: pd.DataFrame, nav_series: Dict[str, Series],
                   bm_series: Dict[str, Any], flows: Dict[str, Optional[float]],
                   d_end: date, d_start: date, d_fy: date, d_mtd: date,
                   label: str, d_flow: Optional[date] = None,
                   d_mis3_start: Optional[date] = None,
                   d_mis3_end: Optional[date] = None) -> Dict[str, Any]:
    """Assemble MIS 1, 2 and 3 for one portfolio.

    ``d_flow`` is the day the Flows/MTD/YTD figures run to -- a day behind
    d_end, since AUM publishes a day late. It only labels the columns; the
    values themselves are already cut to that day upstream.
    """
    d_flow = d_flow or d_end
    # A period's return is measured from the close of the day BEFORE it opens:
    # "since 1st April" compounds 1 April's own move, so it bases off 31 March.
    # Basing off 1 April instead silently drops the first day, which understates
    # every return by roughly one day's move. d_mtd is already a baseline date
    # (the last day of the previous month), so it needs no shift.
    d_period_base = d_start - timedelta(days=1)
    d_fy_base = d_fy - timedelta(days=1)

    # MIS 3 defaults to the financial year but can be pointed at any window.
    d_m3_start = d_mis3_start or d_fy
    d_m3_end = d_mis3_end or d_end
    d_m3_base = d_m3_start - timedelta(days=1)

    nifty = bm_series.get(NIFTY_KEY)
    nifty_levels = nifty.levels if nifty is not None else {}
    n_metrics = _scheme_metrics(nifty_levels, d_end, d_period_base, d_fy_base, d_mtd)

    rows: List[Dict[str, Any]] = []
    for _, r in portfolio_df.iterrows():
        isin = str(r["ISIN"]).strip().upper()
        sm = _scheme_metrics(nav_series.get(isin, {}), d_end, d_period_base, d_fy_base, d_mtd)

        bm_name = r["Benchmark"]
        bser = bm_series.get(bm_name)
        bm_levels = bser.levels if bser is not None else {}
        bm = _scheme_metrics(bm_levels, d_end, d_period_base, d_fy_base, d_mtd)

        rows.append({
            "Scheme Name": r["Scheme Name"],
            "ISIN": isin,
            "Benchmark": bm_name,
            "Allocation": r["Allocation (%)"],
            "Weight": r["Normalized_Weight"],
            # day
            "Day Scheme Return": sm["day"],
            "Day Nifty Return": n_metrics["day"],
            "Day Excess vs Nifty": _diff(sm["day"], n_metrics["day"]),
            "Day Benchmark Return": bm["day"],
            "Day Excess vs Benchmark": _diff(sm["day"], bm["day"]),
            # period
            "Period Scheme Return": sm["period"],
            "Period Nifty Return": n_metrics["period"],
            "Period Excess vs Nifty": _diff(sm["period"], n_metrics["period"]),
            "Period Benchmark Return": bm["period"],
            "Period Excess vs Benchmark": _diff(sm["period"], bm["period"]),
            # MIS 3 window — the financial year unless a separate range is set.
            # The "since 1st April" footers on MIS 1/2 keep using sm["fy"].
            "FY Scheme Return": _window_return(nav_series.get(isin, {}), d_m3_base, d_m3_end),
            "FY Nifty Return": _window_return(nifty_levels, d_m3_base, d_m3_end),
            "FY Excess vs Nifty": _diff(_window_return(nav_series.get(isin, {}), d_m3_base, d_m3_end),
                                        _window_return(nifty_levels, d_m3_base, d_m3_end)),
            "FY Benchmark Return": _window_return(bm_levels, d_m3_base, d_m3_end),
            "FY Excess vs Benchmark": _diff(_window_return(nav_series.get(isin, {}), d_m3_base, d_m3_end),
                                            _window_return(bm_levels, d_m3_base, d_m3_end)),
            # Supplementary flow columns, in crores — these mirror the
            # reference report, where MTD and YTD are cumulative net flows
            # rather than returns.
            "Flows": (flows.get(isin) or {}).get("day"),
            "MTD": (flows.get(isin) or {}).get("mtd"),
            "YTD": (flows.get(isin) or {}).get("ytd"),
            # Return-based month/FY figures, kept for the detail frame.
            "MTD Return": sm["mtd"],
            # provenance
            "_nav_date": sm["date_end"],
            "_nav_prev_date": sm["date_prev"],
            "_day_gap_days": sm["day_gap_days"],
            "_day_stale": sm["day_stale"],
            "_bm_status": bser.status if bser is not None else STATUS_NONE,
        })

    df = pd.DataFrame(rows)
    weights = df["Weight"].tolist()

    def wavg(col: str) -> Tuple[Optional[float], float]:
        return _weighted(df[col].tolist(), weights)

    p_day, cov_day = wavg("Day Scheme Return")
    p_period, cov_period = wavg("Period Scheme Return")
    p_fy, _ = wavg("FY Scheme Return")
    b_day, _ = wavg("Day Benchmark Return")
    b_period, _ = wavg("Period Benchmark Return")
    b_fy, _ = wavg("FY Benchmark Return")

    n_day, n_period, n_fy = n_metrics["day"], n_metrics["period"], n_metrics["fy"]

    day_label = _long_date(d_end)
    period_label = f"{_long_date(d_start)} - {_long_date(d_end)}"
    fy_label = f"{d_fy:%b'%y}"

    # ── MIS 1 ────────────────────────────────────────────────────────────────
    mis1_cols = [
        "S.No", "Scheme Name", "ISIN", "Allocation",
        "Day Scheme Return", "Day Nifty Return", "Day Excess vs Nifty",
        "Period Scheme Return", "Period Nifty Return", "Period Excess vs Nifty",
        "Flows", "MTD", "YTD",
    ]
    mis1 = df.sort_values("Period Excess vs Nifty", ascending=False, na_position="last").reset_index(drop=True)
    mis1.insert(0, "S.No", range(1, len(mis1) + 1))

    mis1_spec = {
        "title": f"{label} — Performance w.r.t Nifty 50",
        "columns": mis1_cols,
        "headers": {
            "S.No": "S.No", "Scheme Name": "Scheme Name", "ISIN": "ISIN",
            "Allocation": "Allocation",
            "Day Scheme Return": "Scheme Return in %",
            "Day Nifty Return": "Return-Nifty 50",
            "Day Excess vs Nifty": "Excess Return Over Nifty 50",
            "Period Scheme Return": "Scheme Return in %",
            "Period Nifty Return": "Return-Nifty 50",
            "Period Excess vs Nifty": "Excess Return Over Nifty 50",
            "Flows": f"Flows- {_long_date(d_flow)}",
            "MTD": f"MTD- {d_flow:%B'%y}",
            "YTD": f"YTD ({_ordinal(d_fy.day)} {d_fy:%B'%y} to {_ordinal(d_flow.day)} {d_flow:%B'%y})",
        },
        # Flows/MTD/YTD are rupee amounts in crores, not percentages.
        "bands": [
            ("", ["S.No", "Scheme Name", "ISIN", "Allocation"]),
            (f"For the day {day_label}", ["Day Scheme Return", "Day Nifty Return", "Day Excess vs Nifty"]),
            (f"For the Period {period_label}", ["Period Scheme Return", "Period Nifty Return", "Period Excess vs Nifty"]),
            ("", ["Flows", "MTD", "YTD"]),
        ],
        "rows": mis1[mis1_cols],
        "pct_cols": ["Day Scheme Return", "Day Nifty Return", "Day Excess vs Nifty",
                     "Period Scheme Return", "Period Nifty Return", "Period Excess vs Nifty"],
        "alloc_cols": ["Allocation"],
        "num_cols": ["Flows", "MTD", "YTD"],
        "colour_cols": ["Day Excess vs Nifty", "Period Excess vs Nifty", "Flows", "MTD", "YTD"],
        "footers": [
            ("day", "Weighted Average Daily Portfolio Return", p_day, False),
            ("day", "Nifty's Daily Return", n_day, False),
            ("day", "Excess Return", _diff(p_day, n_day), True),
            ("period", "Weighted Average Portfolio Return", p_period, False),
            ("period", "Nifty's Return", n_period, False),
            ("period", "Excess Return", _diff(p_period, n_period), True),
            ("period", f"Excess Return ( since 1st April'{d_fy:%y} )", _diff(p_fy, n_fy), True),
        ],
    }

    # ── MIS 2 ────────────────────────────────────────────────────────────────
    mis2_cols = [
        "S.No", "Scheme Name", "ISIN", "Benchmark", "Allocation",
        "Day Scheme Return", "Day Benchmark Return", "Day Excess vs Benchmark",
        "Period Scheme Return", "Period Benchmark Return", "Period Excess vs Benchmark",
    ]
    mis2 = df.sort_values("Period Excess vs Benchmark", ascending=False, na_position="last").reset_index(drop=True)
    mis2.insert(0, "S.No", range(1, len(mis2) + 1))

    mis2_spec = {
        "title": f"{label} — Performance w.r.t Own Benchmark",
        "columns": mis2_cols,
        "headers": {
            "S.No": "S.No", "Scheme Name": "Scheme Name", "ISIN": "ISIN",
            "Benchmark": "Benchmark", "Allocation": "Allocation",
            "Day Scheme Return": "Scheme Return in %",
            "Day Benchmark Return": "Return-Own Benchmark",
            "Day Excess vs Benchmark": "Excess Return Over Own Benchmark",
            "Period Scheme Return": "Scheme Return in %",
            "Period Benchmark Return": "Return-Own Benchmark",
            "Period Excess vs Benchmark": "Excess Return Over Own Benchmark",
        },
        "bands": [
            ("", ["S.No", "Scheme Name", "ISIN", "Benchmark", "Allocation"]),
            (f"For the day {day_label}", ["Day Scheme Return", "Day Benchmark Return", "Day Excess vs Benchmark"]),
            (f"For the Period {period_label}", ["Period Scheme Return", "Period Benchmark Return", "Period Excess vs Benchmark"]),
        ],
        "rows": mis2[mis2_cols],
        "pct_cols": ["Day Scheme Return", "Day Benchmark Return", "Day Excess vs Benchmark",
                     "Period Scheme Return", "Period Benchmark Return", "Period Excess vs Benchmark"],
        "alloc_cols": ["Allocation"],
        "num_cols": [],
        "colour_cols": ["Day Excess vs Benchmark", "Period Excess vs Benchmark"],
        "footers": [
            ("day", "Weighted Average Daily Portfolio Return", p_day, False),
            ("day", "Weighted Average Daily Benchmark Return", b_day, False),
            ("day", "Excess Return", _diff(p_day, b_day), True),
            ("period", "Weighted Average Portfolio Return", p_period, False),
            ("period", "Weighted Average Benchmark Return", b_period, False),
            ("period", "Excess Return", _diff(p_period, b_period), True),
            ("period", f"Excess Return ( since 1st April'{d_fy:%y} )", _diff(p_fy, b_fy), True),
        ],
    }

    # ── MIS 3 ────────────────────────────────────────────────────────────────
    mis3_cols = [
        "Scheme Name", "ISIN", "Allocation", "Benchmark",
        "FY Scheme Return", "FY Nifty Return", "FY Excess vs Nifty",
        "FY Benchmark Return", "FY Excess vs Benchmark",
    ]
    mis3 = df.sort_values("FY Excess vs Nifty", ascending=False, na_position="last").reset_index(drop=True)

    fy_range = f"{_long_date(d_m3_start)} - {_long_date(d_m3_end)}"
    mis3_spec = {
        "title": f"{label} — Performance {_long_date(d_m3_start)} to {_long_date(d_m3_end)}",
        "columns": mis3_cols,
        "headers": {
            "Scheme Name": "Scheme Name", "ISIN": "ISIN",
            "Allocation": "New allocation", "Benchmark": "Benchmark",
            "FY Scheme Return": "Scheme Return",
            "FY Nifty Return": "Nifty",
            "FY Excess vs Nifty": "Excess Return over Nifty",
            "FY Benchmark Return": "Benchmark Return",
            "FY Excess vs Benchmark": "Excess Return over Benchmark",
        },
        "bands": [
            ("", ["Scheme Name", "ISIN", "Allocation", "Benchmark"]),
            (fy_range, ["FY Scheme Return", "FY Nifty Return", "FY Excess vs Nifty"]),
            (fy_range + " ", ["FY Benchmark Return", "FY Excess vs Benchmark"]),
        ],
        "rows": mis3[mis3_cols],
        "pct_cols": ["FY Scheme Return", "FY Nifty Return", "FY Excess vs Nifty",
                     "FY Benchmark Return", "FY Excess vs Benchmark"],
        "alloc_cols": ["Allocation"],
        "num_cols": [],
        "colour_cols": ["FY Excess vs Nifty", "FY Excess vs Benchmark"],
        "total_row": {
            "Scheme Name": "Portfolio weighted average",
            "ISIN": "", "Allocation": df["Allocation"].sum(), "Benchmark": "",
            "FY Scheme Return": p_fy,
            "FY Nifty Return": n_fy,
            "FY Excess vs Nifty": _diff(p_fy, n_fy),
            "FY Benchmark Return": b_fy,
            "FY Excess vs Benchmark": _diff(p_fy, b_fy),
        },
        "footers": [],
    }

    coverage_warnings = []

    if n_metrics["day_stale"]:
        coverage_warnings.append(
            f"{label}: the latest Nifty 50 observation ({n_metrics['date_end']:%d-%b-%Y}) and the one "
            f"before it are {n_metrics['day_gap_days']} days apart, so no genuine one-day return exists. "
            f"The 'for the day' columns show N/A rather than a multi-day move."
        )

    stale_rows = df.loc[df["_day_stale"], "Scheme Name"].tolist()
    if stale_rows:
        coverage_warnings.append(
            f"{label}: NAV history has a hole before the report date for {', '.join(stale_rows)}, so their "
            f"one-day returns show N/A instead of a multi-day move mislabelled as daily."
        )

    if cov_day < 0.9999:
        missing = df.loc[df["Day Scheme Return"].isna(), "Scheme Name"].tolist()
        coverage_warnings.append(
            f"{label}: daily return covers {cov_day * 100:.1f}% of allocation. "
            f"No NAV move available for: {', '.join(missing)}. Weights were renormalized over the rest."
        )
    if cov_period < 0.9999:
        missing = df.loc[df["Period Scheme Return"].isna(), "Scheme Name"].tolist()
        coverage_warnings.append(
            f"{label}: period return covers {cov_period * 100:.1f}% of allocation. "
            f"No NAV history available for: {', '.join(missing)}. Weights were renormalized over the rest."
        )

    return {
        "label": label,
        "mis1": mis1_spec,
        "mis2": mis2_spec,
        "mis3": mis3_spec,
        "detail": df,
        "warnings": coverage_warnings,
    }


def generate_mis_reports_data(
    portfolio_df: pd.DataFrame,
    start_date: date,
    end_date: date,
    previous_portfolio_df: Optional[pd.DataFrame] = None,
    include_flows: bool = False,
    skip_saturdays: bool = True,
    progress=None,
    supplied_levels_df: Optional[pd.DataFrame] = None,
    mis3_start: Optional[date] = None,
    mis3_end: Optional[date] = None,
) -> Dict[str, Any]:
    """Build the full MIS pack for one (optionally two) portfolios."""
    global SKIP_WEEKEND_MARKS
    SKIP_WEEKEND_MARKS = bool(skip_saturdays)
    d_end = end_date
    d_start = start_date
    d_fy = get_fy_start_date(end_date)
    d_mtd = get_mtd_start_date(end_date)
    d_m3_start = mis3_start or d_fy
    d_m3_end = mis3_end or d_end
    earliest = min(d_start, d_fy, d_mtd, d_m3_start, d_end - timedelta(days=10))

    frames = [portfolio_df]
    if previous_portfolio_df is not None and not previous_portfolio_df.empty:
        frames.append(previous_portfolio_df)

    all_isins = sorted({str(i).strip().upper() for f in frames for i in f["ISIN"]})
    all_benchmarks = sorted({str(b).strip() for f in frames for b in f["Benchmark"]})
    if NIFTY_KEY not in all_benchmarks:
        all_benchmarks.append(NIFTY_KEY)

    # One AMFI pass covers both the schemes and the benchmark proxy funds.
    proxy_isins = required_proxy_isins(all_benchmarks)
    fetch_isins = list(dict.fromkeys(all_isins + proxy_isins))

    df_nav_raw, nav_gaps = fetch_mis_nav_history(
        fetch_isins, earliest, max(d_end, d_m3_end), progress=progress
    )
    if df_nav_raw.empty:
        raise ValueError(
            "AMFI returned no NAV history for the portfolio ISINs. Check the ISINs and the date range."
        )

    nav_series = build_nav_series(df_nav_raw, all_isins)

    missing_navs = [i for i, s in nav_series.items() if not s]
    if len(missing_navs) == len(all_isins):
        raise ValueError(
            "None of the portfolio ISINs matched an AMFI record. Check that the ISINs are correct "
            "and that the date range covers dates on which these schemes reported a NAV."
        )

    proxy_series = nav_frame_to_isin_series(df_nav_raw)
    proxy_series = overlay_live_navs(proxy_series, proxy_isins, d_end)
    bm_series = build_benchmark_series(all_benchmarks, proxy_series, earliest, describe_gaps(nav_gaps))

    # Official NSE total-return levels replace the proxy funds where available.
    # NIFTY_KEY is excluded on purpose: the Nifty comparison column reconciles
    # against the reference on the *price* index, and switching it to total
    # return would quietly move a figure that already matches.
    bm_series, bm_notes = apply_tri_levels(
        bm_series, earliest, max(d_end, d_m3_end), exclude_keys=(NIFTY_KEY,)
    )

    # User-supplied index levels win over any proxy, including an EXACT one:
    # the report is meant to reconcile to the user's own benchmark source.
    if supplied_levels_df is not None and not supplied_levels_df.empty:
        supplied, supply_problems = parse_supplied_levels(supplied_levels_df)
        bm_notes.extend(supply_problems)
        bm_series, supply_notes = apply_supplied_levels(bm_series, supplied)
        bm_notes.extend(supply_notes)

    # The Nifty column is the headline PRICE index, unlike the schemes' own
    # "TR INR" benchmarks. An index-fund proxy answers the total-return
    # question and reads ~0.5pp high over a quarter, so source it directly.
    nifty_levels = fetch_nifty_price_index(earliest, d_end)
    nifty_fallback_note = ""
    if nifty_levels:
        bm_series[NIFTY_KEY] = BenchmarkSeries(
            NIFTY_KEY, NIFTY_KEY, nifty_levels, STATUS_EXACT,
            proxy_name="Nifty 50 price index (dividends excluded)",
        )
    else:
        nifty_fallback_note = (
            "The Nifty 50 price index could not be fetched, so the Nifty column falls back to a "
            "total-return index fund. It will read slightly high — total return includes dividends."
        )

    flows: Dict[str, Dict[str, Optional[float]]] = {
        i: {"day": None, "mtd": None, "ytd": None} for i in all_isins
    }
    flow_notes: List[str] = []
    flow_date: Optional[date] = None
    if include_flows:
        flows, flow_date = compute_scheme_flows(
            df_nav_raw, all_isins, d_end, d_mtd, d_fy, notes=flow_notes
        )

    current = _build_reports(portfolio_df, nav_series, bm_series, flows,
                             d_end, d_start, d_fy, d_mtd, "14 Fund AR Model Portfolio", d_flow=flow_date,
                             d_mis3_start=d_m3_start, d_mis3_end=d_m3_end)

    previous = None
    if previous_portfolio_df is not None and not previous_portfolio_df.empty:
        previous = _build_reports(previous_portfolio_df, nav_series, bm_series, flows,
                                  d_end, d_start, d_fy, d_mtd, "Previous 14 Fund AR Model Portfolio", d_flow=flow_date,
                                  d_mis3_start=d_m3_start, d_mis3_end=d_m3_end)

    warnings: List[str] = list(bm_notes) + list(current["warnings"])
    if nifty_fallback_note:
        warnings.append(nifty_fallback_note)
    warnings.extend(flow_notes)
    if previous:
        warnings.extend(previous["warnings"])

    if missing_navs:
        warnings.append(
            "No AMFI NAV history found for these ISINs (shown as N/A): " + ", ".join(missing_navs)
        )

    if nav_gaps:
        spans = ", ".join(f"{a:%d-%b-%Y} to {b:%d-%b-%Y}" for a, b in nav_gaps)
        warnings.append(
            f"AMFI did not serve NAV history for {spans}. Returns whose baseline falls in those windows "
            f"are measured from the nearest earlier NAV instead — re-run in a few minutes for complete data."
        )

    approx = [s.requested for s in bm_series.values() if s.status == STATUS_APPROX]
    if approx:
        warnings.append(
            "These benchmarks have no matching index fund, so the closest available index was used as a "
            "stand-in — treat their excess returns as indicative: " + ", ".join(approx)
        )

    unavailable = [s.requested for s in bm_series.values() if s.status == STATUS_NONE]
    if unavailable:
        warnings.append(
            "No index data could be sourced for these benchmarks; their columns show N/A: "
            + ", ".join(unavailable)
        )

    return {
        "current": current,
        "previous": previous,
        "benchmark_report": describe_resolution(bm_series),
        "warnings": warnings,
        "dates": {
            "start_date": d_start,
            "end_date": d_end,
            "fy_start": d_fy,
            "mis3_start": d_m3_start,
            "mis3_end": d_m3_end,
            "mtd_start": d_mtd,
        },
        "include_flows": include_flows,
    }


# ─── Excel Workbook Export Engine ─────────────────────────────────────────────

FONT = "Segoe UI"
GREEN_FILL = PatternFill("solid", fgColor="92D050")
RED_FILL = PatternFill("solid", fgColor="FF0000")
HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
BAND_FILL = PatternFill("solid", fgColor="FFFFFF")
TITLE_FILL = PatternFill("solid", fgColor="FCE4D6")

THIN = Side(style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_L = Alignment(horizontal="left", vertical="center")
ALIGN_R = Alignment(horizontal="center", vertical="center")


def _write_mis_block(ws, spec: Dict[str, Any], start_row: int) -> int:
    """Write one MIS table (band header, column header, rows, footers).

    Returns the next free row.
    """
    cols = spec["columns"]
    headers = spec["headers"]
    bands = spec["bands"]
    df = spec["rows"]
    pct = set(spec.get("pct_cols", []))
    alloc = set(spec.get("alloc_cols", []))
    nums = set(spec.get("num_cols", []))
    colour = set(spec.get("colour_cols", []))

    col_index = {c: i + 1 for i, c in enumerate(cols)}

    band_row = start_row
    head_row = start_row + 1

    # Band header: the merged "For the day …" / "For the Period …" captions.
    for caption, band_cols in bands:
        first = col_index[band_cols[0]]
        last = col_index[band_cols[-1]]
        if last > first:
            ws.merge_cells(start_row=band_row, start_column=first, end_row=band_row, end_column=last)
        cell = ws.cell(row=band_row, column=first, value=caption.strip() or None)
        cell.font = Font(name=FONT, size=10, bold=True)
        cell.alignment = ALIGN_C
        if caption.strip():
            cell.fill = HEADER_FILL
        for c in range(first, last + 1):
            ws.cell(row=band_row, column=c).border = BORDER

    # The leading identity band carries the report title.
    title_first = col_index[bands[0][1][0]]
    title_last = col_index[bands[0][1][-1]]
    ws.merge_cells(start_row=band_row, start_column=title_first, end_row=band_row, end_column=title_last)
    tcell = ws.cell(row=band_row, column=title_first, value=spec["title"])
    tcell.font = Font(name=FONT, size=10, bold=True)
    tcell.alignment = ALIGN_C
    tcell.fill = TITLE_FILL

    ws.row_dimensions[band_row].height = 24
    ws.row_dimensions[head_row].height = 42

    for c in cols:
        cell = ws.cell(row=head_row, column=col_index[c], value=headers[c])
        cell.font = Font(name=FONT, size=9, bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = ALIGN_C
        cell.border = BORDER

    r = head_row + 1
    for _, row in df.iterrows():
        for c in cols:
            val = row.get(c)
            cell = ws.cell(row=r, column=col_index[c])

            if c in pct or c in nums:
                if val is None or pd.isna(val):
                    cell.value = "N/A"
                    cell.alignment = ALIGN_R
                else:
                    cell.value = round(float(val), STORE_DP)
                    cell.number_format = DISPLAY_FMT
                    cell.alignment = ALIGN_R
                    if c in colour:
                        cell.fill = GREEN_FILL if float(val) >= 0 else RED_FILL
            elif c in alloc:
                cell.value = round(float(val), STORE_DP) / 100.0 if pd.notna(val) else 0.0
                cell.number_format = "0.00%"
                cell.alignment = ALIGN_R
            elif c == "S.No":
                cell.value = int(val)
                cell.alignment = ALIGN_R
            else:
                cell.value = str(val) if pd.notna(val) else ""
                cell.alignment = ALIGN_L

            cell.font = Font(name=FONT, size=9)
            cell.border = BORDER
        ws.row_dimensions[r].height = 16
        r += 1

    # Optional total row (MIS 3 renders its portfolio average inline).
    total = spec.get("total_row")
    if total:
        for c in cols:
            val = total.get(c)
            cell = ws.cell(row=r, column=col_index[c])
            if c in pct or c in nums:
                if val is None or pd.isna(val):
                    cell.value = "N/A"
                else:
                    cell.value = round(float(val), STORE_DP)
                    cell.number_format = DISPLAY_FMT
                    if c in colour:
                        cell.fill = GREEN_FILL if float(val) >= 0 else RED_FILL
                cell.alignment = ALIGN_R
            elif c in alloc:
                cell.value = (round(float(val), STORE_DP) / 100.0) if val not in (None, "") and pd.notna(val) else None
                cell.number_format = "0.00%"
                cell.alignment = ALIGN_R
            else:
                cell.value = str(val) if val not in (None, "") else ""
                cell.alignment = ALIGN_L
            cell.font = Font(name=FONT, size=10, bold=True)
            cell.border = BORDER
        ws.row_dimensions[r].height = 26
        r += 1

    # Footers sit under the band they summarise, label then value.
    band_by_key = {}
    for caption, band_cols in bands:
        if caption.lower().startswith("for the day"):
            band_by_key["day"] = band_cols
        elif caption.lower().startswith("for the period"):
            band_by_key["period"] = band_cols

    if spec.get("footers"):
        r += 1
        offsets = {"day": 0, "period": 0}
        base_row = r
        for band_key, label, value, coloured in spec["footers"]:
            band_cols = band_by_key.get(band_key)
            if not band_cols:
                continue
            row_i = base_row + offsets[band_key]
            offsets[band_key] += 1

            lab_c = col_index[band_cols[0]]
            val_first = col_index[band_cols[1]]
            val_last = col_index[band_cols[-1]]

            lcell = ws.cell(row=row_i, column=lab_c, value=label)
            lcell.font = Font(name=FONT, size=9, bold=True)
            lcell.alignment = ALIGN_C
            lcell.border = BORDER

            if val_last > val_first:
                ws.merge_cells(start_row=row_i, start_column=val_first, end_row=row_i, end_column=val_last)
            vcell = ws.cell(row=row_i, column=val_first)
            if value is None or pd.isna(value):
                vcell.value = "N/A"
            else:
                vcell.value = round(float(value), STORE_DP)
                vcell.number_format = DISPLAY_FMT
                if coloured:
                    vcell.fill = GREEN_FILL if float(value) >= 0 else RED_FILL
            vcell.font = Font(name=FONT, size=9, bold=True)
            vcell.alignment = ALIGN_C
            for c in range(val_first, val_last + 1):
                ws.cell(row=row_i, column=c).border = BORDER
            ws.row_dimensions[row_i].height = 30

        r = base_row + max(offsets.values() or [0])

    return r + 2


def _autosize(ws, specs: List[Dict[str, Any]]):
    widths: Dict[int, int] = {}
    for spec in specs:
        for i, c in enumerate(spec["columns"], 1):
            hdr = spec["headers"][c]
            base = 34 if c == "Scheme Name" else (30 if c == "Benchmark" else 14)
            widths[i] = max(widths.get(i, 10), min(max(len(hdr) // 2 + 6, base), 40))
    for i, w in widths.items():
        ws.column_dimensions[get_column_letter(i)].width = w


def export_mis_to_excel(mis_data: Dict[str, Any]) -> bytes:
    """Excel workbook with one sheet per MIS, laid out like the reference report."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    current = mis_data["current"]
    previous = mis_data.get("previous")

    sheet_defs = [
        ("MIS 1", "mis1", "MIS 1 : Performance of Funds w.r.t Nifty"),
        ("MIS 2", "mis2", "MIS 2 : Performance of Funds w.r.t Benchmark"),
        ("MIS 3", "mis3", "Performance of fund basket since financial year start"),
    ]

    for sheet_name, key, caption in sheet_defs:
        ws = wb.create_sheet(title=sheet_name)
        ws.sheet_view.showGridLines = False

        specs = [current[key]]
        if previous:
            specs.append(previous[key])

        row = 2
        for spec in specs:
            row = _write_mis_block(ws, spec, row)

        note = ws.cell(row=row + 1, column=2, value=caption)
        note.font = Font(name=FONT, size=10, bold=True)

        _autosize(ws, specs)

    # Provenance sheet: every benchmark, its proxy fund, and coverage.
    ws_src = wb.create_sheet(title="Benchmark Sources")
    ws_src.sheet_view.showGridLines = False
    ws_src.cell(row=1, column=1, value="Benchmark index levels used in this report").font = \
        Font(name=FONT, size=12, bold=True)
    ws_src.cell(row=2, column=1, value=(
        "Each benchmark is represented by a passive index fund; its NAV history (AMFI) stands in for the "
        "index level. Rows marked APPROX have no matching index fund and use the closest available index."
    )).font = Font(name=FONT, size=9, italic=True)

    bench_df = mis_data.get("benchmark_report")
    if bench_df is not None and not bench_df.empty:
        for ci, col in enumerate(bench_df.columns, 1):
            c = ws_src.cell(row=4, column=ci, value=col)
            c.font = Font(name=FONT, size=9, bold=True)
            c.fill = HEADER_FILL
            c.border = BORDER
        for ri, (_, r) in enumerate(bench_df.iterrows(), 5):
            for ci, col in enumerate(bench_df.columns, 1):
                c = ws_src.cell(row=ri, column=ci, value=str(r[col]))
                c.font = Font(name=FONT, size=9)
                c.border = BORDER
        for ci, col in enumerate(bench_df.columns, 1):
            ws_src.column_dimensions[get_column_letter(ci)].width = 46 if col in ("Proxy Fund", "Note") else 20

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── PDF Report Export Engine ──────────────────────────────────────────────────

def export_mis_to_pdf(mis_data: Dict[str, Any]) -> bytes:
    """Landscape PDF containing MIS 1, MIS 2 and MIS 3."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    except ImportError as exc:
        raise RuntimeError("reportlab is not installed. Add reportlab to requirements.txt") from exc

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=16, rightMargin=16, topMargin=18, bottomMargin=18)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("T", parent=styles["Heading1"], fontName="Helvetica-Bold",
                                 fontSize=15, leading=19, textColor=colors.HexColor("#1F497D"), spaceAfter=3)
    sub_style = ParagraphStyle("S", parent=styles["Normal"], fontName="Helvetica", fontSize=8.5,
                               leading=11, textColor=colors.HexColor("#555555"), spaceAfter=10)
    sec_style = ParagraphStyle("H", parent=styles["Heading2"], fontName="Helvetica-Bold", fontSize=11,
                               leading=14, textColor=colors.HexColor("#1F497D"), spaceBefore=8, spaceAfter=5)
    hdr_style = ParagraphStyle("CH", parent=styles["Normal"], fontName="Helvetica-Bold",
                               fontSize=6.5, leading=8, alignment=1)
    txt_style = ParagraphStyle("CT", parent=styles["Normal"], fontName="Helvetica", fontSize=6.5, leading=8)
    num_style = ParagraphStyle("CN", parent=styles["Normal"], fontName="Helvetica", fontSize=6.5,
                               leading=8, alignment=1)

    d = mis_data["dates"]
    story = [
        Paragraph("MUTUAL FUND PORTFOLIO MIS REPORT", title_style),
        Paragraph(
            f"<b>Period:</b> {d['start_date']:%d-%b-%Y} to {d['end_date']:%d-%b-%Y} &nbsp;|&nbsp; "
            f"<b>Financial year start:</b> {d['fy_start']:%d-%b-%Y} &nbsp;|&nbsp; "
            f"Benchmark levels sourced from AMFI index-fund NAVs.",
            sub_style),
    ]

    def build_table(spec):
        cols = spec["columns"]
        pct = set(spec.get("pct_cols", []))
        nums = set(spec.get("num_cols", []))
        alloc = set(spec.get("alloc_cols", []))
        colour = set(spec.get("colour_cols", []))

        data = [[Paragraph(spec["headers"][c], hdr_style) for c in cols]]
        style = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9D9D9")),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#999999")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]

        body = list(spec["rows"].to_dict("records"))
        if spec.get("total_row"):
            body.append(spec["total_row"])

        for ri, row in enumerate(body, 1):
            cells = []
            for ci, c in enumerate(cols):
                v = row.get(c)
                if c in pct or c in nums:
                    if v is None or pd.isna(v):
                        cells.append(Paragraph("N/A", num_style))
                    else:
                        cells.append(Paragraph(f"{float(v):.2f}", num_style))
                        if c in colour:
                            bg = colors.HexColor("#92D050") if float(v) >= 0 else colors.HexColor("#FF0000")
                            style.append(("BACKGROUND", (ci, ri), (ci, ri), bg))
                elif c in alloc:
                    cells.append(Paragraph(f"{float(v):.2f}%" if pd.notna(v) else "", num_style))
                elif c == "S.No":
                    cells.append(Paragraph(str(int(v)), num_style))
                else:
                    cells.append(Paragraph(str(v) if pd.notna(v) else "", txt_style))
            data.append(cells)

        if spec.get("total_row"):
            style.append(("BACKGROUND", (0, len(data) - 1), (-1, len(data) - 1), colors.HexColor("#EAECEE")))

        n = len(cols)
        wide = {"Scheme Name": 3.4, "Benchmark": 2.8, "ISIN": 1.5, "S.No": 0.5}
        units = [wide.get(c, 1.0) for c in cols]
        total_units = sum(units)
        avail = landscape(A4)[0] - 32
        widths = [avail * u / total_units for u in units]

        t = Table(data, colWidths=widths, repeatRows=1)
        t.setStyle(TableStyle(style))
        return t

    def footer_table(spec):
        rows = [[Paragraph(f"<b>{lab}</b>", txt_style),
                 Paragraph("N/A" if (v is None or pd.isna(v)) else f"<b>{float(v):.2f}</b>", num_style)]
                for _band, lab, v, _col in spec.get("footers", [])]
        if not rows:
            return None
        t = Table(rows, colWidths=[260, 70])
        t.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#999999")),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        return t

    blocks = [("current", mis_data["current"])]
    if mis_data.get("previous"):
        blocks.append(("previous", mis_data["previous"]))

    for key, caption in [("mis1", "MIS 1 : Performance w.r.t Nifty"),
                         ("mis2", "MIS 2 : Performance w.r.t Own Benchmark"),
                         ("mis3", "MIS 3 : Performance since financial year start")]:
        story.append(Paragraph(caption, sec_style))
        for _which, block in blocks:
            spec = block[key]
            story.append(Paragraph(f"<b>{spec['title']}</b>", sub_style))
            story.append(build_table(spec))
            ft = footer_table(spec)
            if ft is not None:
                story.append(Spacer(1, 5))
                story.append(ft)
            story.append(Spacer(1, 10))

    doc.build(story)
    return buf.getvalue()


# ─── Streamlit UI View Renderer ───────────────────────────────────────────────

def _display_headers(spec: Dict[str, Any]) -> Dict[str, str]:
    """Map internal column names to on-screen headers.

    The report headers repeat across bands — two columns are both called
    "Scheme Return in %" — which is fine in Excel under a merged band caption
    but ambiguous in a flat table, so the band is appended here.
    """
    band_of = {c: caption.strip() for caption, band_cols in spec["bands"] for c in band_cols}

    out = {}
    for c in spec["columns"]:
        hdr = spec["headers"][c]
        band = band_of.get(c, "").lower()
        if band.startswith("for the day"):
            hdr = f"{hdr} — Day"
        elif band.startswith("for the period"):
            hdr = f"{hdr} — Period"
        out[c] = hdr
    return out


def _display_frame(spec: Dict[str, Any]) -> pd.DataFrame:
    """Rename internal columns to their report headers for on-screen display."""
    df = spec["rows"].copy()
    if spec.get("total_row"):
        df = pd.concat([df, pd.DataFrame([spec["total_row"]])], ignore_index=True)
    return df[spec["columns"]].rename(columns=_display_headers(spec))


def _render_block(block: Dict[str, Any], key: str):
    spec = block[key]
    st.markdown(f"**{spec['title']}**")
    disp = _display_frame(spec)
    headers = _display_headers(spec)

    styler = disp.style.format(
        {c: "{:.2f}" for c in disp.columns if disp[c].dtype.kind in "fc"}, na_rep="N/A"
    )

    def colour_by_sign(v):
        if pd.isna(v) or not isinstance(v, (int, float)):
            return ""
        return "background-color:#92D050;color:#111" if v >= 0 else "background-color:#FF6B6B;color:#111"

    existing = [headers[c] for c in spec.get("colour_cols", []) if headers.get(c) in disp.columns]
    if existing:
        # Styler.map replaced applymap in pandas 2.1; keep working on both.
        apply_elementwise = getattr(styler, "map", None) or styler.applymap
        styler = apply_elementwise(colour_by_sign, subset=existing)

    st.dataframe(styler, use_container_width=True, hide_index=True)

    if spec.get("footers"):
        f_day = [(l, v) for b, l, v, _ in spec["footers"] if b == "day"]
        f_per = [(l, v) for b, l, v, _ in spec["footers"] if b == "period"]
        c1, c2 = st.columns(2)
        for col, items, heading in ((c1, f_day, "For the day"), (c2, f_per, "For the period")):
            if not items:
                continue
            with col:
                st.caption(heading)
                foot = pd.DataFrame(
                    [{"Measure": l, "Value": (None if v is None or pd.isna(v) else round(v, STORE_DP))}
                     for l, v in items]
                )
                st.dataframe(
                    foot.style.format({"Value": "{:.2f}"}, na_rep="N/A"),
                    use_container_width=True, hide_index=True,
                )


_PREV_MODES = [
    "Auto (carry over previous version)",
    "Manual Entry / Edit",
    "Upload Excel File",
]


def render_mis_generator_page():
    """Render the full interactive MIS Generator Streamlit page."""
    render_section_header(
        "📊",
        "MIS Generator Module",
        "Portfolio vs Nifty and vs own benchmark — for the day, the period, and since 1 April",
    )

    render_info_card(
        "<strong>MIS Generator:</strong> Enter your portfolio or upload an Excel file. Scheme NAVs come from "
        "AMFI. Benchmark index levels come from a passive index fund tracking each benchmark, read from the "
        "same AMFI feed — so every figure traces to real data. Where a benchmark cannot be sourced the report "
        "shows <code>N/A</code> rather than a substituted number."
    )

    # Restore the portfolios from disk before falling back to the sample
    # template. Streamlit's session state is memory-only, so without this a
    # refresh or a redeploy silently replaces a hand-entered portfolio with the
    # sample -- and the MIS portfolio is edited about once a year, so that edit
    # is exactly the thing that must not be lost.
    if "mis_workspace_loaded" not in st.session_state:
        saved = mis_history.load_workspace()
        st.session_state["portfolio_input_df"] = (
            saved["current"] if saved["current"] is not None
            else pd.DataFrame(DEFAULT_SAMPLE_PORTFOLIO)
        )
        st.session_state["prev_portfolio_input_df"] = (
            saved["previous"] if saved["previous"] is not None
            else pd.DataFrame(columns=["Scheme Name", "ISIN", "Allocation (%)", "Benchmark"])
        )
        if saved["auto_previous"] is not None:
            st.session_state["mis_auto_prev_df"] = saved["auto_previous"]
        # Seed the radio's key before the widget exists so the saved choice is
        # what renders; setting it afterwards would be ignored for this run.
        if saved["prev_mode"] in _PREV_MODES:
            st.session_state["mis_prev_mode"] = saved["prev_mode"]
        st.session_state["mis_workspace_restored_at"] = saved["saved_at"]
        st.session_state["mis_workspace_loaded"] = True

    if "portfolio_input_df" not in st.session_state:
        st.session_state["portfolio_input_df"] = pd.DataFrame(DEFAULT_SAMPLE_PORTFOLIO)
    if "prev_portfolio_input_df" not in st.session_state:
        st.session_state["prev_portfolio_input_df"] = pd.DataFrame(columns=["Scheme Name", "ISIN", "Allocation (%)", "Benchmark"])
    if "mis_results" not in st.session_state:
        st.session_state["mis_results"] = None

    col_cfg = {
        "Scheme Name": st.column_config.TextColumn("Scheme Name", width="large", required=True),
        "ISIN": st.column_config.TextColumn("ISIN", width="medium", required=True),
        "Allocation (%)": st.column_config.NumberColumn("Allocation (%)", min_value=0.0, max_value=100.0, format="%.2f"),
        "Benchmark": st.column_config.TextColumn("Benchmark", width="large"),
    }

    with finance_panel("1. Portfolio Input"):
        input_mode = st.radio("Portfolio Entry Method", ["Manual Entry / Edit", "Upload Excel File"], horizontal=True)

        if input_mode == "Manual Entry / Edit":
            if st.button("🔄 Reset Sample Template"):
                mis_history.clear_workspace()
                st.session_state["portfolio_input_df"] = pd.DataFrame(DEFAULT_SAMPLE_PORTFOLIO)
                st.session_state["mis_results"] = None
                st.session_state.pop("mis_manual_editor", None)
                st.session_state["mis_current_sig"] = None
                st.rerun()
            current_portfolio_df = st.data_editor(
                st.session_state["portfolio_input_df"], num_rows="dynamic",
                use_container_width=True, column_config=col_cfg, key="mis_manual_editor",
            )
        else:
            uploaded_file = st.file_uploader(
                "Upload Portfolio Excel (.xlsx)", type=["xlsx", "xls"],
                help="Columns: Scheme Name, ISIN, Allocation (%), Benchmark",
            )
            if uploaded_file is not None:
                try:
                    current_portfolio_df = read_portfolio_excel(uploaded_file)
                    st.success(f"Loaded {len(current_portfolio_df)} rows.")
                except Exception as exc:
                    st.error(f"Error reading uploaded Excel file: {exc}")
                    current_portfolio_df = st.session_state["portfolio_input_df"]
            else:
                st.info("Upload an Excel file above, or switch to Manual Entry.")
                current_portfolio_df = st.session_state["portfolio_input_df"]

        clean_df, warns, errs = validate_and_normalize_portfolio(current_portfolio_df)
        for w in warns:
            st.warning(w)
        for e in errs:
            st.error(e)
        if not clean_df.empty:
            st.caption(
                f"📋 **Valid schemes:** {len(clean_df)} | "
                f"**Total allocation:** {clean_df['Allocation (%)'].sum():.2f}%"
            )

        # Remember the version of the current portfolio that was on screen before
        # this edit, so "Previous Portfolio" can carry it over automatically.
        cur_sig = _portfolio_signature(current_portfolio_df)
        if st.session_state.get("mis_current_sig") is None:
            st.session_state["mis_current_sig"] = cur_sig
            st.session_state["mis_current_df"] = current_portfolio_df.copy()
            # Seed the previous portfolio with the current one so every report
            # carries both blocks from the first run. They stay identical until
            # the current portfolio actually changes, at which point the prior
            # version takes over below.
            st.session_state["mis_auto_prev_df"] = current_portfolio_df.copy()
        elif cur_sig != st.session_state["mis_current_sig"]:
            # The current portfolio just changed: yesterday's current is today's previous.
            st.session_state["mis_auto_prev_df"] = st.session_state.get("mis_current_df")
            st.session_state["mis_current_sig"] = cur_sig
            st.session_state["mis_current_df"] = current_portfolio_df.copy()

        # Keep the on-disk copy in step with the editors. Written only when the
        # content actually differs, so a plain rerun does not rewrite the file.
        st.session_state["portfolio_input_df"] = current_portfolio_df.copy()
        # The previous portfolio and its source mode belong in the signature too.
        # Switching source without touching the current portfolio is still a
        # change worth persisting, and without them here it would not be saved
        # until something else happened to move.
        _ws_sig = (
            cur_sig,
            _portfolio_signature(st.session_state.get("mis_auto_prev_df")),
            _portfolio_signature(st.session_state.get("prev_portfolio_input_df")),
            st.session_state.get("mis_prev_mode", ""),
        )
        if st.session_state.get("mis_workspace_sig") != _ws_sig:
            if mis_history.save_workspace(
                current_portfolio_df,
                st.session_state.get("prev_portfolio_input_df"),
                st.session_state.get("mis_auto_prev_df"),
                st.session_state.get("mis_prev_mode", ""),
            ):
                st.session_state["mis_workspace_sig"] = _ws_sig

        # Report what is genuinely on disk, read back rather than assumed, so a
        # failed write shows up here instead of at the next reload.
        _disk = mis_history.load_workspace()
        if _disk["current"] is not None:
            _match = _portfolio_signature(_disk["current"]) == cur_sig
            _when = str(_disk.get("saved_at") or "")[:16].replace("T", " ")
            st.caption(
                f"💾 Saved to disk at {_when} — survives refresh and restart."
                if _match else
                "⚠️ On-disk copy is out of step with the editor. Click "
                "**Generate MIS Reports** to force a save."
            )
        else:
            st.caption("⚠️ Nothing saved to disk yet — this portfolio will not survive a reload.")

    with finance_panel("2. Previous Portfolio (optional)"):
        st.caption(
            "Renders the second comparison block that appears beneath each MIS table in the "
            "reference report. Leave empty to skip it."
        )
        prev_mode = st.radio(
            "Previous Portfolio Source", _PREV_MODES,
            horizontal=True, key="mis_prev_mode",
        )

        auto_prev = st.session_state.get("mis_auto_prev_df")

        if prev_mode == "Auto (carry over previous version)":
            if auto_prev is not None and not auto_prev.empty:
                unchanged = _portfolio_signature(auto_prev) == _portfolio_signature(current_portfolio_df)
                if unchanged:
                    st.info(
                        f"Same as the current portfolio ({len(auto_prev)} schemes). Both blocks are "
                        f"written to every report; change section 1 and this holds the prior version."
                    )
                else:
                    st.success(
                        f"Carried over the {len(auto_prev)}-scheme portfolio that was in section 1 "
                        f"before your last change."
                    )
                st.dataframe(auto_prev, use_container_width=True, hide_index=True)
                prev_raw = auto_prev
            else:
                st.info("Add schemes in section 1 and they appear here automatically.")
                prev_raw = pd.DataFrame(columns=["Scheme Name", "ISIN", "Allocation (%)", "Benchmark"])

        elif prev_mode == "Manual Entry / Edit":
            seed = st.session_state["prev_portfolio_input_df"]
            if seed.empty and auto_prev is not None and not auto_prev.empty:
                seed = auto_prev  # start from the carried-over version rather than blank
            prev_raw = st.data_editor(
                seed, num_rows="dynamic",
                use_container_width=True, column_config=col_cfg, key="mis_prev_editor",
            )
            # Persist edits: the Generate click reruns the script, and without
            # this the editor is reseeded from the carry-over and the edits are
            # silently discarded before the report is built.
            st.session_state["prev_portfolio_input_df"] = prev_raw.copy()
            mis_history.save_workspace(
                st.session_state.get("portfolio_input_df"), prev_raw,
                st.session_state.get("mis_auto_prev_df"), prev_mode,
            )

        else:
            prev_file = st.file_uploader(
                "Upload Previous Portfolio Excel (.xlsx)", type=["xlsx", "xls"],
                help="Columns: Scheme Name, ISIN, Allocation (%), Benchmark",
                key="mis_prev_upload",
            )
            if prev_file is not None:
                try:
                    prev_raw = read_portfolio_excel(prev_file)
                    st.session_state["prev_portfolio_input_df"] = prev_raw.copy()
                    mis_history.save_workspace(
                        st.session_state.get("portfolio_input_df"), prev_raw,
                        st.session_state.get("mis_auto_prev_df"), prev_mode,
                    )
                    st.success(f"Loaded {len(prev_raw)} rows.")
                except Exception as exc:
                    st.error(f"Error reading previous portfolio file: {exc}")
                    prev_raw = pd.DataFrame(columns=["Scheme Name", "ISIN", "Allocation (%)", "Benchmark"])
            else:
                st.info("Upload an Excel file above, or switch to another source.")
                prev_raw = pd.DataFrame(columns=["Scheme Name", "ISIN", "Allocation (%)", "Benchmark"])

        prev_clean, prev_warns, _prev_errs = validate_and_normalize_portfolio(prev_raw)
        for w in prev_warns:
            st.warning(f"Previous portfolio: {w}")
        if not prev_clean.empty:
            st.caption(
                f"📋 **Previous schemes:** {len(prev_clean)} | "
                f"**Total allocation:** {prev_clean['Allocation (%)'].sum():.2f}%"
            )

        # Same read-back confirmation the current portfolio gets. Which frame is
        # the one that must persist depends on the source: Auto carries a
        # snapshot, the other two carry what was typed or uploaded.
        _prev_live = auto_prev if prev_mode.startswith("Auto") else prev_raw
        _prev_sig = _portfolio_signature(_prev_live)
        _pdisk = mis_history.load_workspace()
        _pstored = _pdisk["auto_previous"] if prev_mode.startswith("Auto") else _pdisk["previous"]

        if _prev_sig == ():
            st.caption("Nothing to save — the previous portfolio is empty.")
        elif _pstored is None:
            st.caption("⚠️ Nothing saved to disk yet — this previous portfolio will not survive a reload.")
        elif _portfolio_signature(_pstored) == _prev_sig:
            _pwhen = str(_pdisk.get("saved_at") or "")[:16].replace("T", " ")
            st.caption(f"💾 Saved to disk at {_pwhen} — survives refresh and restart.")
        else:
            st.caption(
                "⚠️ On-disk copy is out of step with the editor. Click "
                "**Generate MIS Reports** to force a save."
            )

    with finance_panel("2b. Saved MIS History"):
        # Clear anything an interrupted save left behind before counting.
        mis_history.prune_orphans()
        saved = mis_history.list_reports()
        n_saved, mb = mis_history.history_size()

        # Count what is actually readable, not just what is listed: a sidecar
        # can survive while its workbook does not.
        n_ok = sum(1 for m in saved if mis_history.load_report(m["id"])[1] is not None)
        if n_ok == n_saved:
            st.caption(
                f"Every report you generate is kept here — **{n_saved} saved**, {mb:.1f} MB, "
                f"all readable. " + mis_history.EPHEMERAL_NOTE
            )
        else:
            st.warning(
                f"{n_saved} report(s) saved but only {n_ok} have a readable workbook. "
                f"The rest were interrupted mid-save and can be deleted."
            )

        if not saved:
            st.info("No saved reports yet. Generate one below and it appears here.")
        else:
            index = {
                f"{m.get('label', m['id'])}  ·  {m.get('scheme_count', 0)} schemes  "
                f"·  saved {str(m.get('saved_at', ''))[:16].replace('T', ' ')}": m["id"]
                for m in saved
            }
            chosen_label = st.selectbox("Saved reports", list(index), key="mis_hist_pick")
            chosen_id = index[chosen_label]
            meta, xlsx = mis_history.load_report(chosen_id)

            if meta is None:
                st.error("That saved report could not be read.")
            else:
                c1, c2, c3 = st.columns(3)
                c1.metric("Period", f"{meta.get('period_start')} → {meta.get('period_end')}")
                c2.metric("Schemes", meta.get("scheme_count", 0))
                c3.metric("Flows", "included" if meta.get("include_flows") else "off")

                with st.expander("Portfolio as reported", expanded=False):
                    st.dataframe(
                        mis_history.portfolio_frame(meta.get("portfolio", [])),
                        use_container_width=True, hide_index=True,
                    )
                    prev_rows = meta.get("previous_portfolio") or []
                    if prev_rows:
                        st.caption("Previous portfolio")
                        st.dataframe(
                            mis_history.portfolio_frame(prev_rows),
                            use_container_width=True, hide_index=True,
                        )
                    for w in (meta.get("warnings") or [])[:4]:
                        st.caption(f"⚠️ {w[:200]}")

                b1, b2, b3 = st.columns(3)
                with b1:
                    if xlsx:
                        st.download_button(
                            "⬇️ Download this report", data=xlsx,
                            file_name=f"{meta.get('id')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True, key="mis_hist_dl",
                        )
                    else:
                        st.caption("Workbook file missing.")
                with b2:
                    # Load the saved portfolio back into the editors so an old
                    # report can be reissued against today's NAVs.
                    if st.button("↩️ Load portfolio into editor", use_container_width=True,
                                 key="mis_hist_load"):
                        st.session_state["portfolio_input_df"] = mis_history.portfolio_frame(
                            meta.get("portfolio", [])
                        )
                        prev_rows = meta.get("previous_portfolio") or []
                        if prev_rows:
                            st.session_state["prev_portfolio_input_df"] = mis_history.portfolio_frame(prev_rows)
                        # Force section 1 to re-seed from the restored frame.
                        st.session_state.pop("mis_manual_editor", None)
                        st.session_state["mis_current_sig"] = None
                        st.rerun()
                with b3:
                    if st.button("🗑️ Delete", use_container_width=True, key="mis_hist_del"):
                        mis_history.delete_report(chosen_id)
                        st.rerun()

    with finance_panel("3. Benchmark index levels (optional)"):
        st.caption(
            "Upload your own index levels to override the proxy funds. Columns: **Benchmark**, "
            "**Date**, **Close**. Use this for benchmarks a proxy can only approximate — a "
            "Morningstar TR INR series, or an index with no tracker such as S&P BSE 250 SmallCap. "
            "Supplied levels take precedence over every proxy."
        )
        levels_file = st.file_uploader(
            "Index levels (.xlsx / .csv)", type=["xlsx", "xls", "csv"], key="mis_levels_upload",
        )
        supplied_levels_df = None
        if levels_file is not None:
            try:
                supplied_levels_df = (
                    pd.read_csv(levels_file) if levels_file.name.lower().endswith(".csv")
                    else read_portfolio_excel(levels_file)
                )
                st.success(f"Loaded {len(supplied_levels_df)} level row(s).")
            except Exception as exc:
                st.error(f"Could not read the levels file: {exc}")

    with finance_panel("4. Date Range & Generate"):
        c1, c2, c3 = st.columns([2, 2, 2])
        today = date.today()
        with c1:
            start_date = st.date_input("Period Start Date", value=today - timedelta(days=270), max_value=today)
        with c2:
            end_date = st.date_input("Report Date", value=today, max_value=today)
        with c3:
            include_flows = st.checkbox(
                "Include Flows column", value=False,
                help="Derives net flows from daily AUM via the AMFI performance API. Noticeably slower.",
            )

        if start_date > end_date:
            st.error("Period Start Date must be on or before the Report Date.")

        skip_saturdays = st.checkbox(
            "Skip Saturday / Sunday NAV marks", value=True, key="mis_skip_sat",
            help="Markets are shut at weekends, but some AMCs still stamp a NAV. "
                 "On: those marks never anchor a return. Off: every published mark is used.",
        )

        use_own_m3 = st.checkbox(
            "Use a separate date range for MIS 3",
            value=False, key="mis3_own_range",
            help="MIS 3 defaults to the financial year (1 April) through the Report Date.",
        )
        mis3_start = mis3_end = None
        if use_own_m3:
            m3c1, m3c2 = st.columns(2)
            with m3c1:
                mis3_start = st.date_input(
                    "MIS 3 Start Date", value=get_fy_start_date(end_date),
                    max_value=today, key="mis3_start_date",
                )
            with m3c2:
                mis3_end = st.date_input(
                    "MIS 3 End Date", value=end_date, max_value=today, key="mis3_end_date",
                )
            if mis3_start > mis3_end:
                st.error("MIS 3 Start Date must be on or before its End Date.")

        live_sig = _report_input_signature(
            clean_df, prev_clean, start_date, end_date, mis3_start, mis3_end,
            include_flows, skip_saturdays,
        )
        # Rebuild automatically once a report exists and any input has moved --
        # editing either portfolio must reach the on-screen tables and the
        # downloads without a second click. AMFI responses are cached, so a
        # composition-only change recomputes rather than refetching.
        inputs_moved = (
            st.session_state.get("mis_results") is not None
            and st.session_state.get("mis_inputs_sig") not in (None, live_sig)
        )
        clicked = st.button("⚡ Generate MIS Reports", type="primary", use_container_width=True)

        if clicked or inputs_moved:
            if clean_df.empty:
                if clicked:
                    st.error("Cannot generate MIS reports without valid portfolio input.")
            elif start_date > end_date:
                if clicked:
                    st.error("Invalid date range.")
            else:
                if inputs_moved and not clicked:
                    st.caption("↻ Inputs changed — rebuilding the report…")
                # AMFI serves ~25s per month of range on a good day and drops
                # responses on a bad one, so this runs for minutes. Show the
                # actual position rather than an indefinite spinner.
                bar = st.progress(0.0, text="Contacting AMFI…")

                def _report(done: int, total: int, label: str) -> None:
                    frac = (done / total) if total else 1.0
                    bar.progress(
                        min(max(frac, 0.0), 1.0),
                        text=f"Fetching NAV history — window {min(done + 1, total)} of {total} ({label})",
                    )

                try:
                    # Whatever a report is generated from is by definition what
                    # the user meant to keep. Save it unconditionally, so a
                    # generated portfolio can never be lost to a reload even if
                    # an editor edit failed to persist on its own rerun.
                    mis_history.save_workspace(
                        current_portfolio_df,
                        st.session_state.get("prev_portfolio_input_df"),
                        st.session_state.get("mis_auto_prev_df"),
                        st.session_state.get("mis_prev_mode", ""),
                    )
                    st.session_state["mis_inputs_sig"] = live_sig
                    st.session_state["mis_results"] = generate_mis_reports_data(
                        clean_df, start_date, end_date,
                        previous_portfolio_df=prev_clean if not prev_clean.empty else None,
                        include_flows=include_flows,
                        skip_saturdays=skip_saturdays,
                        progress=_report,
                        supplied_levels_df=supplied_levels_df,
                        mis3_start=mis3_start,
                        mis3_end=mis3_end,
                    )
                    if clicked:
                        st.success("MIS reports generated.")
                        try:
                            entry_id = mis_history.save_report(
                                st.session_state["mis_results"],
                                export_mis_to_excel(st.session_state["mis_results"]),
                                clean_df,
                                prev_clean if not prev_clean.empty else None,
                            )
                            st.caption(f"💾 Saved to history as `{entry_id}`")
                        except Exception as exc:
                            # A history failure must not cost the user the report,
                            # but it must not pass unnoticed either.
                            st.warning(
                                f"⚠️ Report generated but NOT saved to history: {exc}. "
                                f"Download it below if you need to keep it."
                            )
                except Exception as exc:
                    st.session_state["mis_results"] = None
                    st.error(f"Failed to generate MIS reports: {exc}")
                finally:
                    bar.empty()

    res = st.session_state.get("mis_results")
    if not res:
        return

    # Inputs are rebuilt automatically above, so reaching here with a mismatch
    # means the rebuild failed. Say so rather than handing over a file that no
    # longer matches what is on screen.
    if st.session_state.get("mis_inputs_sig") not in (None, live_sig):
        st.warning(
            "⚠️ The tables and downloads below are from an earlier run — the automatic "
            "rebuild did not complete. Click **Generate MIS Reports** to retry."
        )

    for w in res.get("warnings", []):
        st.warning(w)

    with finance_panel("4. MIS Reports"):
        tab1, tab2, tab3, tab4 = st.tabs([
            "📊 MIS 1: vs Nifty 50",
            "🎯 MIS 2: vs Own Benchmark",
            "📅 MIS 3: Since 1 April",
            "🔎 Benchmark Sources",
        ])

        blocks = [res["current"]] + ([res["previous"]] if res.get("previous") else [])

        with tab1:
            for b in blocks:
                _render_block(b, "mis1")
        with tab2:
            for b in blocks:
                _render_block(b, "mis2")
        with tab3:
            for b in blocks:
                _render_block(b, "mis3")
        with tab4:
            st.caption(
                "Each benchmark resolves to a passive index fund whose NAV history stands in for the index "
                "level. EXACT = the fund tracks that index. APPROX = no fund tracks it, closest index used. "
                "UNAVAILABLE = no data, columns show N/A."
            )
            st.dataframe(res["benchmark_report"], use_container_width=True, hide_index=True)

        render_section_header("📥", "Export Reports")
        e1, e2 = st.columns(2)
        stamp = f"{res['dates']['start_date']}_to_{res['dates']['end_date']}"

        with e1:
            try:
                st.download_button(
                    "Download Excel MIS Report (.xlsx)",
                    data=export_mis_to_excel(res),
                    file_name=f"MIS_Report_{stamp}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            except Exception as exc:
                st.error(f"Excel export failed: {exc}")

        with e2:
            try:
                st.download_button(
                    "Download PDF MIS Report (.pdf)",
                    data=export_mis_to_pdf(res),
                    file_name=f"MIS_Report_{stamp}.pdf",
                    mime="application/pdf",
                    use_container_width=True,
                )
            except Exception as exc:
                st.warning(f"PDF export unavailable: {exc}")
