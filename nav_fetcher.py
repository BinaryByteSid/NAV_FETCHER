"""
AMFI NAV Fetcher — Standalone UI
Fetch historical NAV data from AMFI India by ISIN(s) and/or date range.

Run with:
    streamlit run nav_fetcher.py
"""

from __future__ import annotations

import os
import re
import time
import threading
import sys
from datetime import datetime, timedelta
from io import BytesIO
from pathlib import Path
from typing import List, Optional, Tuple, Union

import pandas as pd
import requests
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ─── Robust date parsing (locale-independent) ────────────────────────────────

_MONTH_ABBR = {
    "jan": "01", "feb": "02", "mar": "03", "apr": "04",
    "may": "05", "jun": "06", "jul": "07", "aug": "08",
    "sep": "09", "oct": "10", "nov": "11", "dec": "12",
}


def _parse_amfi_date_str(date_str: str) -> str | None:
    """Convert a single AMFI date string like '01-Jun-2026' → '2026-06-01' (ISO).

    Returns None if unparseable. This avoids reliance on pandas locale settings.
    """
    if not isinstance(date_str, str) or not date_str.strip():
        return None
    parts = date_str.strip().split("-")
    if len(parts) == 3:
        day, month_abbr, year = parts
        month_num = _MONTH_ABBR.get(month_abbr.lower()[:3])
        if month_num and day.isdigit() and year.isdigit():
            return f"{year}-{month_num}-{day.zfill(2)}"
    return None


def parse_amfi_date_series(series: pd.Series) -> pd.Series:
    """Parse a Series of AMFI date strings ('dd-Mon-YYYY' or 'dd-mm-YYYY') to datetime.

    Uses a manual month-abbreviation lookup so it works identically on all
    locales / pandas versions (including Streamlit Cloud).
    """
    # Fast path: try explicit format first (works when locale is OK)
    result = pd.to_datetime(series, format="%d-%b-%Y", errors="coerce")
    if not result.isna().all():
        # Fill any remaining NaTs with manual parsing
        mask = result.isna() & series.notna()
        if mask.any():
            iso_strs = series[mask].apply(_parse_amfi_date_str)
            result[mask] = pd.to_datetime(iso_strs, errors="coerce")
        return result

    # Fallback: try numeric format dd-mm-YYYY (for pre-converted dates)
    result = pd.to_datetime(series, format="%d-%m-%Y", errors="coerce")
    if not result.isna().all():
        return result

    # Last resort: manual element-wise parsing for AMFI dates
    iso_strs = series.apply(_parse_amfi_date_str)
    result = pd.to_datetime(iso_strs, errors="coerce")
    if not result.isna().all():
        return result

    # Absolute fallback
    return pd.to_datetime(series, errors="coerce")


def get_fund_seed(name: str) -> int:
    hash_val = 0
    for char in name:
        hash_val = ord(char) + ((hash_val << 5) - hash_val)
        hash_val = hash_val & 0xFFFFFFFF
    if hash_val > 0x7FFFFFFF:
        hash_val = hash_val - 0x100000000
    return abs(hash_val) % 100


def load_portfolio_aum_data() -> pd.DataFrame:
    paths = [
        # Preferred: scheme-level AUM only. The full portfolio workbook carries
        # instrument holdings this function never reads, and shipping those to a
        # deployment would publish position-level data to serve three columns.
        "scheme_aum_monthly.xlsx",
        "../scheme_aum_monthly.xlsx",
        "../portfolio last 6 months.xlsx",
        "portfolio last 6 months.xlsx",
        "../Misc/portfolio last 6 months.xlsx",
        "c:/Users/sidha/OneDrive/Desktop/portfolio last 6 months.xlsx",
        "c:/Users/sidha/OneDrive/Desktop/Misc/portfolio last 6 months.xlsx",
    ]
    for path in paths:
        if os.path.exists(path):
            try:
                df = pd.read_excel(path, header=3)
                df['SD_Scheme ISIN'] = df['SD_Scheme ISIN'].astype(str).str.strip().str.upper()
                df['PD_Month End'] = pd.to_numeric(df['PD_Month End'], errors='coerce')
                df['PD_Scheme AUM'] = pd.to_numeric(df['PD_Scheme AUM'], errors='coerce')
                return df
            except Exception:
                pass
    return pd.DataFrame()


def calculate_aum_for_row(row, df_port: pd.DataFrame) -> float:
    scheme_name = row.get("Scheme Name", "")
    isin_growth = row.get("ISIN Div Payout / ISIN Growth") or row.get("ISIN Div Payout/ ISIN Growth")
    isin_reinvestment = row.get("ISIN Div Reinvestment")
    date_str = row.get("Date") or row.get("NAV Date")
    
    try:
        if isinstance(date_str, pd.Timestamp):
            year = date_str.year
            month = date_str.month
        else:
            parts = str(date_str).split("-")
            month_map = {
                "Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6,
                "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12
            }
            if len(parts) == 3:
                month_str = parts[1]
                year = int(parts[2][:4])
                month = month_map.get(month_str[:3], 4)
            else:
                parsed_dt = pd.to_datetime(date_str)
                year = parsed_dt.year
                month = parsed_dt.month
    except Exception:
        year = 2026
        month = 4
        
    m_val = year * 100 + month
    
    match_port = pd.DataFrame()
    if not df_port.empty:
        isins = []
        if isin_growth and pd.notna(isin_growth) and isin_growth != "-":
            isins.append(str(isin_growth).strip().upper())
        if isin_reinvestment and pd.notna(isin_reinvestment) and isin_reinvestment != "-":
            isins.append(str(isin_reinvestment).strip().upper())
            
        if isins:
            match_port = df_port[df_port['SD_Scheme ISIN'].isin(isins)]
            
    aum_monthly = None
    if not match_port.empty:
        match_month = match_port[match_port['PD_Month End'] == m_val]
        if not match_month.empty:
            aum_monthly = float(match_month['PD_Scheme AUM'].iloc[0])
        else:
            latest_row = match_port.sort_values('PD_Month End', ascending=False).iloc[0]
            aum_monthly = float(latest_row['PD_Scheme AUM'])
            
    # No real disclosed AUM for this scheme/month → return None rather than
    # fabricating a value. The live perf-API pass may still fill it; otherwise the
    # AUM/flows stay blank so nothing is invented.
    return aum_monthly


def map_section_to_ids(sec):
    sec_lower = str(sec).lower()
    
    # Maturity Type
    maturity_id = 1  # Open ended default
    if "close" in sec_lower:
        maturity_id = 2
    elif "interval" in sec_lower:
        maturity_id = 2
        
    # Category
    cat_id = 1  # Equity default
    if "debt" in sec_lower:
        cat_id = 2
    elif "hybrid" in sec_lower:
        cat_id = 3
    elif "solution" in sec_lower:
        cat_id = 4
    elif "other" in sec_lower:
        cat_id = 5
    elif "gilt" in sec_lower or "money market" in sec_lower or "income" in sec_lower:
        cat_id = 2
        
    # Subcategory defaults
    if cat_id == 1:
        sub_id = 1
    elif cat_id == 2:
        sub_id = 15
    elif cat_id == 3:
        sub_id = 30
    elif cat_id == 4:
        sub_id = 36
    elif cat_id == 5:
        sub_id = 38
    else:
        sub_id = 1
    
    # Subcategory mapping rules
    if cat_id == 1:  # Equity
        if "large & mid" in sec_lower:
            sub_id = 2
        elif "large cap" in sec_lower:
            sub_id = 1
        elif "flexi cap" in sec_lower:
            sub_id = 3
        elif "multi cap" in sec_lower:
            sub_id = 4
        elif "mid cap" in sec_lower:
            sub_id = 5
        elif "small cap" in sec_lower:
            sub_id = 6
        elif "value" in sec_lower:
            sub_id = 7
        elif "elss" in sec_lower:
            sub_id = 8
        elif "contra" in sec_lower:
            sub_id = 9
        elif "dividend yield" in sec_lower:
            sub_id = 10
        elif "focused" in sec_lower:
            sub_id = 11
        elif "sectoral" in sec_lower or "thematic" in sec_lower:
            sub_id = 12
    elif cat_id == 2:  # Debt
        if "gilt with 10" in sec_lower or "10 year constant" in sec_lower:
            sub_id = 29
        elif "gilt" in sec_lower:
            sub_id = 28
        elif "medium to long" in sec_lower:
            sub_id = 14
        elif "long duration" in sec_lower:
            sub_id = 13
        elif "ultra short" in sec_lower:
            sub_id = 19
        elif "short duration" in sec_lower:
            sub_id = 15
        elif "medium duration" in sec_lower:
            sub_id = 16
        elif "money market" in sec_lower:
            sub_id = 17
        elif "low duration" in sec_lower:
            sub_id = 18
        elif "liquid" in sec_lower:
            sub_id = 20
        elif "overnight" in sec_lower:
            sub_id = 21
        elif "dynamic bond" in sec_lower:
            sub_id = 22
        elif "corporate bond" in sec_lower:
            sub_id = 23
        elif "credit risk" in sec_lower:
            sub_id = 24
        elif "banking" in sec_lower or "psu" in sec_lower:
            sub_id = 25
        elif "floater" in sec_lower:
            sub_id = 26
        elif "fmp" in sec_lower:
            sub_id = 27
    elif cat_id == 3:  # Hybrid
        if "aggressive hybrid" in sec_lower:
            sub_id = 30
        elif "conservative hybrid" in sec_lower or "conservative hyrbid" in sec_lower:
            sub_id = 31
        elif "equity savings" in sec_lower:
            sub_id = 32
        elif "arbitrage" in sec_lower:
            sub_id = 33
        elif "multi asset" in sec_lower:
            sub_id = 34
        elif "dynamic asset" in sec_lower or "balanced advantage" in sec_lower:
            sub_id = 35
        elif "balanced hybrid" in sec_lower:
            sub_id = 40
    elif cat_id == 4:  # Solution Oriented
        if "children" in sec_lower:
            sub_id = 36
        elif "retirement" in sec_lower:
            sub_id = 37
    elif cat_id == 5:  # Other
        if "index fund" in sec_lower or "index" in sec_lower:
            sub_id = 38
        elif "etf" in sec_lower:
            sub_id = 38
        elif "fof" in sec_lower or "fund of funds" in sec_lower:
            sub_id = 39
        
    return maturity_id, cat_id, sub_id


def clean_name(name: str) -> str:
    n = str(name).lower()
    n = n.replace("flexicap", "flexi cap")
    n = n.replace("multicap", "multi cap")
    n = n.replace("midcap", "mid cap")
    n = n.replace("smallcap", "small cap")
    n = n.replace("largecap", "large cap")
    n = n.replace("focussed", "focused")
    
    n = n.replace("-", " ").replace("/", " ").replace("(", " ").replace(")", " ")
    n = n.replace(" sl ", " sun life ")
    tokens = n.split()
    suffixes_to_remove = {
        "direct", "regular", "retail", "plan", "growth", "option", "idcw", "dividend", 
        "payout", "reinvestment", "annual", "monthly", "weekly", "quarterly", "fortnightly",
        "bonus", "fund"
    }
    cleaned_tokens = [t for t in tokens if t not in suffixes_to_remove]
    return " ".join(cleaned_tokens)


API_CACHE = {}


# Cached: the MIS re-runs whenever a portfolio is edited, and a financial year
# of flows is hundreds of these calls. Without this every edit re-fetches them.
# Thread-local sessions: populate_actual_aum fans these calls out across a
# pool, and a shared Session is not thread-safe.
def _is_cloud_env() -> bool:
    """Detect Streamlit Cloud (ephemeral FS / ~1 GB RAM) vs a local machine.

    Used to size worker pools: local runs can afford far more concurrency,
    which is the main speed lever for multi-year date ranges.
    """
    if os.environ.get("STREAMLIT_SHARING_MODE") or os.environ.get("STREAMLIT_RUNTIME_ENV"):
        return True
    # No writable local amfi_cache dir next to this file ⇒ treat as cloud.
    return not os.path.exists(
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "amfi_cache")
    )



def _perf_api_workers(n_calls: int) -> int:
    """Concurrent daily-AUM performance-API POSTs. Moderate to avoid the AMFI
    gateway rate-limiting (which returns empty responses → blank AUM)."""
    if n_calls <= 0:
        return 1
    cap = 6 if _is_cloud_env() else 8
    return max(1, min(cap, n_calls))



def _nav_chunk_workers(n_chunks: int) -> int:
    """Concurrent NAV-history downloads. Kept moderate: AMFI's portal throttles
    (503/timeouts) when hit with too many parallel connections, which shows up as
    'can't fetch even 6 months'. 6 local / 3 cloud is a safe, reliable balance."""
    if n_chunks <= 0:
        return 1
    cap = 3 if _is_cloud_env() else 6
    return max(1, min(cap, n_chunks))



_thread_local = threading.local()

# Responses persist here so a re-run does not refetch a financial year of calls.
API_CACHE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "api_cache")
os.makedirs(API_CACHE_DIR, exist_ok=True)

# Populated by populate_actual_aum each run so the UI can report how much AUM
# came from the live feed vs was carried or left blank.
_LAST_AUM_STATS: dict = {}

def get_api_session():
    if not hasattr(_thread_local, "session"):
        _thread_local.session = requests.Session()
    return _thread_local.session


def fetch_performance_data_from_api(date_str: str, maturity_id: int, category_id: int, subcategory_id: int) -> list:
    """Daily AUM for one (category, date) from AMFI's fund-performance API.

    Single implementation shared by Fund Performance and the MIS. This is
    app.py's version, kept over nav_fetcher's for three reasons that all
    showed up as blank or wrong AUM:

      * it sends the browser Referer/Origin/X-Requested-With headers, without
        which AMFI's gateway returns an empty or non-SUCCESS response;
      * it never caches a failure, whereas the old version wrote an empty list
        into its cache on error, turning one transient failure into a
        permanently blank AUM for the rest of the session;
      * responses persist to disk, so a re-run does not refetch a financial
        year of calls.
    """
    import json
    import os
    import time
    import random

    # 1. Skip querying dates before 2018-01-01 (SEBI categorization and AMFI performance dataset starts around 2018)
    # This prevents thousands of useless requests that return 0 rows.
    try:
        parsed_date = datetime.strptime(date_str, "%d-%b-%Y")
        if parsed_date.year < 2018:
            return []
    except Exception:
        pass

    key_str = f"{date_str}_{maturity_id}_{category_id}_{subcategory_id}"
    cache_path = os.path.join(API_CACHE_DIR, f"{key_str}.json")
    
    # Check if we can use the cached file
    if os.path.exists(cache_path):
        try:
            parsed_date = datetime.strptime(date_str, "%d-%b-%Y")
            yesterday = datetime.today() - timedelta(days=1)
            if parsed_date < yesterday:
                with open(cache_path, "r", encoding="utf-8") as f:
                    return json.load(f)
            else:
                # 1 hour TTL for recent dates
                file_age = time.time() - os.path.getmtime(cache_path)
                if file_age < 3600:
                    with open(cache_path, "r", encoding="utf-8") as f:
                        return json.load(f)
        except Exception:
            pass
            
    url = "https://www.amfiindia.com/gateway/pollingsebi/api/amfi/fundperformance"
    # AMFI's gateway API rejects/blanks requests that don't look like they came
    # from the fund-performance web page. Sending the browser Referer/Origin and an
    # explicit JSON Accept header is what makes it return data instead of an empty
    # or non-SUCCESS response — a common cause of blank AUM.
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Content-Type": "application/json",
        "Accept": "application/json, text/plain, */*",
        "Origin": "https://www.amfiindia.com",
        "Referer": "https://www.amfiindia.com/research-information/other-data/fund-performance",
        "X-Requested-With": "XMLHttpRequest",
    }
    payload = {
        "maturityType": maturity_id,
        "category": category_id,
        "subCategory": subcategory_id,
        "mfid": 0,
        "reportDate": date_str
    }

    max_retries = 4
    delay = 1.0

    session = get_api_session()

    for attempt in range(max_retries):
        try:
            resp = session.post(url, json=payload, headers=headers, timeout=(10, 40))
            if resp.status_code == 200:
                try:
                    res_data = resp.json()
                except Exception:
                    res_data = None
                if isinstance(res_data, dict):
                    # Accept the data whether or not validationMsg is present/SUCCESS —
                    # the important thing is that a data array came back.
                    rows = res_data.get("data")
                    if rows is None and res_data.get("validationMsg") == "SUCCESS":
                        rows = []
                    if isinstance(rows, list):
                        try:
                            with open(cache_path, "w", encoding="utf-8") as f:
                                json.dump(rows, f)
                        except Exception:
                            pass
                        return rows
            elif resp.status_code in (429, 503):
                # Throttled — back off harder before retrying.
                time.sleep(delay + random.uniform(0, 1.0))
                delay = min(delay * 2, 20)
                continue
            time.sleep(delay + random.uniform(0, 0.5))
            delay = min(delay * 2, 20)
        except Exception:
            time.sleep(delay + random.uniform(0, 0.5))
            delay = min(delay * 2, 20)

    return []


def fetch_performance_data_from_api(date_str: str, maturity_id: int, category_id: int, subcategory_id: int) -> list:
    key = (date_str, maturity_id, category_id, subcategory_id)
    if key in API_CACHE:
        return API_CACHE[key]
        
    url = "https://www.amfiindia.com/gateway/pollingsebi/api/amfi/fundperformance"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Content-Type": "application/json"
    }
    payload = {
        "maturityType": maturity_id,
        "category": category_id,
        "subCategory": subcategory_id,
        "mfid": 0,
        "reportDate": date_str
    }
    
    import time
    max_retries = 3
    backoff = 1.0
    
    for attempt in range(max_retries):
        try:
            resp = requests.post(url, json=payload, headers=headers, timeout=20)
            if resp.status_code == 200:
                res_data = resp.json()
                if res_data.get("validationMsg") == "SUCCESS":
                    rows = res_data.get("data", [])
                    API_CACHE[key] = rows
                    return rows
                else:
                    print(f"AMFI API Validation failed for {key}: {res_data.get('validationMsg')}")
            else:
                print(f"AMFI API returned HTTP code {resp.status_code} for {key}")
        except Exception as e:
            print(f"Attempt {attempt+1} failed for {key} with error: {e}")
        
        if attempt < max_retries - 1:
            time.sleep(backoff)
            backoff *= 2
            
    API_CACHE[key] = []
    return []


def find_matching_perf_row(nav_name: str, perf_rows: list) -> Optional[dict]:
    if not perf_rows:
        return None
    cleaned_nav = clean_name(nav_name)
    if not cleaned_nav:
        return None
        
    # 1. Substring match
    for p_row in perf_rows:
        p_name = p_row.get("schemeName") or ""
        cleaned_perf = clean_name(p_name)
        if cleaned_perf and (cleaned_perf in cleaned_nav or cleaned_nav in cleaned_perf):
            return p_row
            
    # 2. Token overlap fallback
    nav_tokens = set(cleaned_nav.split())
    best_row = None
    best_score = 0.0
    for p_row in perf_rows:
        p_name = p_row.get("schemeName") or ""
        cleaned_perf = clean_name(p_name)
        if not cleaned_perf:
            continue
        perf_tokens = set(cleaned_perf.split())
        intersection = nav_tokens.intersection(perf_tokens)
        if intersection:
            score = len(intersection) / len(nav_tokens.union(perf_tokens))
            if score > best_score:
                best_score = score
                best_row = p_row
                
    if best_score > 0.4:
        return best_row
    return None


def populate_actual_aum(df: pd.DataFrame, df_port: pd.DataFrame, want_aum: bool = True,
                        fetch_live_aum: bool = False, budget_seconds: float | None = None,
                        max_workers: int | None = None, on_incomplete=None,
                        gap_fill: str = "carry") -> pd.DataFrame:
    """Attach a daily AUM to every row, live from AMFI where available.

    Single implementation shared by Fund Performance and the MIS. They want
    different things from the days the live feed does not cover, which is the
    one behaviour that genuinely differed between the old copies:

      gap_fill="carry"     the nearest real AUM is carried across the gap, so a
                           continuous series can be charted. Right for an AUM
                           report, wrong for a flow one -- a carried figure is
                           indistinguishable from AMFI repeating a stale value
                           and every carried day would score as a zero flow.

      gap_fill="fallback"  uncovered days fall back to the NAV-scaled monthly
                           AUM, leaving real movement between consecutive days
                           for the flow formula to measure.
    """
    if df.empty:
        return df.copy()

    if not want_aum:
        df_res = df.copy()
        df_res["AUM"] = None
        return df_res
        
    # 1. Calculate the fallback AUM (NAV-scaled monthly AUM) for every row.
    # This ensures we always have a meaningful, date-varying AUM estimate.
    # Pre-build dictionary lookup for target ISINs
    parsed_isins_set = set()
    col_growth = "ISIN Div Payout / ISIN Growth" if "ISIN Div Payout / ISIN Growth" in df.columns else "ISIN Div Payout/ ISIN Growth"
    if col_growth in df.columns:
        parsed_isins_set.update(df[col_growth].dropna().astype(str).str.strip().str.upper())
    if "ISIN Div Reinvestment" in df.columns:
        parsed_isins_set.update(df["ISIN Div Reinvestment"].dropna().astype(str).str.strip().str.upper())
        
    aum_lookup = {}
    if not df_port.empty:
        for _, row in df_port.iterrows():
            isin = row.get('SD_Scheme ISIN')
            m_end = row.get('PD_Month End')
            aum = row.get('PD_Scheme AUM')
            if pd.notna(isin) and pd.notna(m_end) and pd.notna(aum):
                isin_str = str(isin).strip().upper()
                if isin_str in parsed_isins_set:
                    if isin_str not in aum_lookup:
                        aum_lookup[isin_str] = []
                    aum_lookup[isin_str].append((int(m_end), float(aum)))

        for isin_str in aum_lookup:
            aum_lookup[isin_str].sort(key=lambda x: x[0], reverse=True)

    def get_aum_from_lookup(isin_growth, isin_reinvestment, m_val, scheme_name, year, month):
        # Real monthly AUM from the uploaded portfolio disclosure file, matched by
        # ISIN. Prefer the exact month; else fall back to the latest disclosed month.
        for isin in [isin_growth, isin_reinvestment]:
            if isin and isin != "-":
                isin_upper = str(isin).strip().upper()
                records = aum_lookup.get(isin_upper)
                if records:
                    for m_end, aum in records:
                        if m_end == m_val:
                            return aum
                    return records[0][1]

        # No real disclosure available for this scheme/month. Return None instead of
        # fabricating a value — the daily perf-API pass (below) may still fill it, and
        # anything left blank stays blank so AUM/flows are never invented.
        return None

    df_res = df.copy()
    scheme_names = df_res["Scheme Name"].tolist()
    isin_growths = df_res[col_growth].tolist()
    isin_reinvs = df_res["ISIN Div Reinvestment"].tolist()
    dates = (df_res["NAV Date"] if "NAV Date" in df_res.columns else df_res["Date"]).tolist()

    month_map = {
        "Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6,
        "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12
    }

    monthly_aums = []
    for i in range(len(df_res)):
        scheme_name = scheme_names[i]
        isin_g = isin_growths[i]
        isin_r = isin_reinvs[i]
        date_val = dates[i]
        
        try:
            if isinstance(date_val, pd.Timestamp):
                year = date_val.year
                month = date_val.month
            else:
                parts = str(date_val).split("-")
                if len(parts) == 3:
                    month_str = parts[1]
                    year = int(parts[2][:4])
                    month = month_map.get(month_str[:3], 4)
                else:
                    parsed_dt = pd.to_datetime(date_val)
                    year = parsed_dt.year
                    month = parsed_dt.month
        except Exception:
            year = 2026
            month = 4
            
        m_val = year * 100 + month
        monthly_aums.append(get_aum_from_lookup(isin_g, isin_r, m_val, scheme_name, year, month))

    # Monthly_AUM may contain None where no real disclosure exists — keep it
    # numeric (NaN) so downstream math propagates blanks instead of erroring.
    df_res["Monthly_AUM"] = pd.to_numeric(pd.Series(monthly_aums, index=df_res.index), errors="coerce")

    # Scale the (real) monthly AUM by each day's NAV relative to the scheme mean NAV
    # so that AUM varies day-to-day proportional to NAV movement. Rows without a real
    # monthly AUM stay NaN (blank) rather than being back-filled with a guess.
    mean_navs = df_res.groupby("Scheme Code")["NAV"].transform("mean")
    mean_navs = mean_navs.fillna(1.0).replace(0.0, 1.0)
    df_res["Fallback_AUM"] = (df_res["Monthly_AUM"] * (df_res["NAV"] / mean_navs)).round(4)
    
    if not fetch_live_aum:
        df_res["AUM"] = df_res["Fallback_AUM"]
        return df_res
        
    # Initialize AUM with the fallback so every row already has a value.
    # Rows where the live API succeeds will get overwritten below.
    df_res["AUM"] = df_res["Fallback_AUM"]
    
    # 2. Try to fetch real AUM from the AMFI performance API for each unique (date, asset-class) pair.
    def get_date_str(dt):
        try:
            if isinstance(dt, pd.Timestamp) or hasattr(dt, "strftime"):
                if pd.notna(dt):
                    return dt.strftime("%d-%b-%Y")
                return None
            parsed = pd.to_datetime(dt, format="%d-%b-%Y", errors="coerce")
            if pd.isna(parsed):
                parsed = pd.to_datetime(dt, errors="coerce")
            if pd.notna(parsed):
                return parsed.strftime("%d-%b-%Y")
        except Exception:
            pass
        return str(dt)
        
    date_col = "NAV Date" if "NAV Date" in df_res.columns else "Date"
    df_res["Date_Str_Temp"] = df_res[date_col].apply(get_date_str)

    perf_lookup = {}

    # Run API calls in parallel using ThreadPoolExecutor for high performance fetching
    from concurrent.futures import ThreadPoolExecutor, as_completed

    # Determine if running inside Streamlit context to show status
    in_streamlit = False
    try:
        from streamlit.runtime import exists as runtime_exists
        in_streamlit = runtime_exists()
    except ImportError:
        pass

    if in_streamlit:
        status_text = st.empty()
        status_text.text("Fetching daily AUM from AMFI fund performance website...")

    # Map every distinct asset-class string to its (maturity, category, subcategory)
    # ID tuple once, so we can deduplicate network calls: many different asset-class
    # labels collapse to the same performance-API query, and over a multi-year range
    # this cuts the number of POSTs dramatically (the dominant cost of live AUM).
    asset_classes = [ac for ac in df_res["Asset Class"].dropna().unique() if ac]
    ac_to_ids = {ac: map_section_to_ids(ac) for ac in asset_classes}

    unique_groups = df_res[["Asset Class", "Date_Str_Temp"]].drop_duplicates()

    # Collapse to unique (ids, date) queries; remember which (date, asset_class)
    # pairs each query feeds so we can rebuild the original lookup shape after.
    ids_date_to_pairs = {}
    for _, grp in unique_groups.iterrows():
        asset_class = grp["Asset Class"]
        date_str = grp["Date_Str_Temp"]
        if not asset_class or not date_str:
            continue
        ids = ac_to_ids.get(asset_class)
        if not ids:
            continue
        query_key = (ids[0], ids[1], ids[2], date_str)
        ids_date_to_pairs.setdefault(query_key, []).append((date_str, asset_class))

    def fetch_task(query_key):
        m_id, c_id, s_id, date_str = query_key
        rows = fetch_performance_data_from_api(date_str, m_id, c_id, s_id)
        return query_key, rows

    perf_query_ok = 0
    _workers = max_workers or _perf_api_workers(len(ids_date_to_pairs))
    _deadline = (time.monotonic() + budget_seconds) if budget_seconds else None
    with ThreadPoolExecutor(max_workers=_workers) as executor:
        futures = [executor.submit(fetch_task, qk) for qk in ids_date_to_pairs]

        completed_count = 0
        total_count = len(futures)
        for future in as_completed(futures):
            if _deadline is not None and time.monotonic() > _deadline:
                # Out of budget. Whatever has not landed stays unfetched; the
                # rows it would have covered fall back below rather than the
                # report hanging on a slow AMFI.
                future.cancel()
                continue
            try:
                query_key, rows = future.result()
                if rows:
                    perf_query_ok += 1
                    # Fan the shared result back out to every (date, asset_class)
                    # pair that maps to this query, preserving downstream logic.
                    for pair in ids_date_to_pairs.get(query_key, []):
                        perf_lookup[pair] = rows
            except Exception:
                pass
            completed_count += 1
            if in_streamlit and completed_count % 10 == 0:
                status_text.text(f"Fetching daily AUM from AMFI fund performance website ({completed_count}/{total_count})...")

    if in_streamlit:
        status_text.empty()
            
    # Pre-compute the scheme-to-performance-scheme matching
    unique_schemes = df_res[["Scheme Name", "Asset Class"]].drop_duplicates()
    scheme_match_cache = {}  # (scheme_name, asset_class) -> matched_perf_schemeName

    matched_by_name = 0
    unmatched = 0

    for _, row in unique_schemes.iterrows():
        s_name = row["Scheme Name"]
        a_class = row["Asset Class"]
        matched_perf_name = None

        # Matched by name only. NAV looked like the stronger key -- it is an
        # exact value from the same feed, where the name is fuzzy -- but the
        # API's navRegular is stamped with its own navDate, which does not
        # always correspond to the report date requested. Matching on it moved
        # the MIS day flows from 14/14 against the reference to 4/14 while
        # reporting every scheme as confidently NAV-matched, so an apparently
        # unique NAV match can land on the wrong scheme.
        #
        # It was briefly kept as a fallback for schemes the name could not
        # resolve, but that trades "no AUM" for "possibly wrong AUM" on exactly
        # the schemes least able to be checked. A missing figure is reported as
        # such elsewhere in this report; a wrong one is not visible at all.
        for key, p_rows in perf_lookup.items():
            if key[1] == a_class and p_rows:
                match_row = find_matching_perf_row(s_name, p_rows)
                if match_row:
                    matched_perf_name = match_row.get("schemeName")
                    matched_by_name += 1
                    break

        if not matched_perf_name:
            unmatched += 1
        scheme_match_cache[(s_name, a_class)] = matched_perf_name

    if matched_by_name or unmatched:
        print(f"AUM row matching: {matched_by_name} matched by name, {unmatched} unmatched "
              f"(those schemes report no AUM rather than a guessed one).")

    # Build a flat O(1) lookup: (date_str, asset_class, perf_name) -> daily_aum
    # Only store AUMs for schemes we actually matched — this keeps the dict small.
    needed_perf_names = set(v for v in scheme_match_cache.values() if v)
    flat_perf_lookup = {}
    for (date_str, asset_class), p_rows in perf_lookup.items():
        for p_row in p_rows:
            p_name = p_row.get("schemeName")
            if p_name and p_name in needed_perf_names:
                flat_perf_lookup[(date_str, asset_class, p_name)] = p_row.get("dailyAUM")
    # Free the large perf_lookup dict — flat_perf_lookup is far smaller
    del perf_lookup

    # Overwrite AUM with real API values using flat O(1) lookup
    a_classes = df_res["Asset Class"].tolist()
    d_strs = df_res["Date_Str_Temp"].tolist()
    s_names = df_res["Scheme Name"].tolist()
    aums = df_res["AUM"].tolist()
    
    # Track which rows actually received a figure from the performance feed.
    # AUM starts out holding the derived fallback, so counting non-null values
    # afterwards reports derived numbers as live -- which is how a run with zero
    # answered queries still claimed 86% of rows came from the live feed.
    live_flags = [False] * len(df_res)

    for i in range(len(df_res)):
        asset_class = a_classes[i]
        date_str = d_strs[i]
        scheme_name = s_names[i]
        
        perf_name = scheme_match_cache.get((scheme_name, asset_class))
        if perf_name:
            daily_aum = flat_perf_lookup.get((date_str, asset_class, perf_name))
            if daily_aum is not None and daily_aum != "":
                try:
                    aums[i] = float(daily_aum)
                    live_flags[i] = True
                except Exception:
                    pass
    
    del flat_perf_lookup, scheme_match_cache
    df_res["AUM"] = pd.to_numeric(pd.Series(aums, index=df_res.index), errors="coerce")

    live_rows = int(sum(live_flags))
    # Rows standing on the derived monthly-AUM fallback rather than a feed value.
    if len(df_res):
        _not_live = ~pd.Series(live_flags, index=df_res.index)
        derived_rows = int((_not_live & df_res["AUM"].notna()).sum())
    else:
        derived_rows = 0

    # Carry real AUM across gaps so no day is left blank when the scheme has at
    # least one real reading. Non-trading days (weekends/holidays, where NAV is
    # also carried) and the occasional failed/slow API call get filled with the
    # nearest real AUM instead of showing "None". Nothing is fabricated — every
    # filled value is an actual disclosed AUM carried to an adjacent day.
    if gap_fill == "fallback":
        df_res["AUM"] = df_res["AUM"].fillna(df_res["Fallback_AUM"])
    else:
        try:
            _dcol = "NAV Date" if "NAV Date" in df_res.columns else "Date"
            df_res["_aum_sort"] = parse_amfi_date_series(df_res[_dcol])
            df_res = df_res.sort_values(["Scheme Code", "_aum_sort"]).reset_index(drop=True)
            df_res["AUM"] = df_res.groupby("Scheme Code")["AUM"].ffill()
            df_res["AUM"] = df_res.groupby("Scheme Code")["AUM"].bfill()
            df_res = df_res.drop(columns=["_aum_sort"])
        except Exception:
            pass

    # Record fetch health so the UI can flag when the live AMFI feed under-delivered.
    total_rows = len(df_res)
    filled_rows = int(df_res["AUM"].notna().sum())
    _LAST_AUM_STATS.clear()
    _LAST_AUM_STATS.update({
        "perf_queries": int(len(ids_date_to_pairs)),
        "perf_ok": int(perf_query_ok),
        "rows_total": total_rows,
        "rows_live_aum": live_rows,
        "rows_derived_aum": derived_rows,
        "rows_carried": max(filled_rows - live_rows - derived_rows, 0),
        "rows_filled": filled_rows,
        "rows_blank": total_rows - filled_rows,
    })

    if on_incomplete and perf_query_ok < len(ids_date_to_pairs):
        on_incomplete(perf_query_ok, len(ids_date_to_pairs))

    df_res = df_res.drop(columns=["Date_Str_Temp"])
    return df_res




# A daily NAV move smaller than this cannot distinguish a stale AUM from a
# genuinely flat one, so staleness is only called above it.
STALE_AUM_MIN_MOVE_PCT = 0.05


def calculate_flows_for_dataframe(df: pd.DataFrame, start_date, meta_cols: list) -> pd.DataFrame:
    """Calculate the flows format columns for a vertical Mutual Fund DataFrame."""
    df = df.copy()
    
    # Standardize columns
    if "NAV" not in df.columns and "NAVs" in df.columns:
        df = df.rename(columns={"NAVs": "NAV"})
    if "NAV" not in df.columns:
        df["NAV"] = None
    if "AUM" not in df.columns:
        df["AUM"] = None
        
    df["NAV"] = pd.to_numeric(df["NAV"], errors="coerce")
    df["AUM"] = pd.to_numeric(df["AUM"], errors="coerce")
    
    # Standardize date column to 'NAV Date'
    date_col = None
    for c in ["NAV Date", "AUM Date", "Date"]:
        if c in df.columns:
            date_col = c
            break
            
    if date_col and date_col != "NAV Date":
        df = df.rename(columns={date_col: "NAV Date"})
        
    if "NAV Date" not in df.columns:
        df["NAV Date"] = None
        
    df["NAV Date_parsed"] = parse_amfi_date_series(df["NAV Date"])
        
    df = df.sort_values(by=["Scheme Code", "NAV Date_parsed"]).reset_index(drop=True)
    
    df["Closing AUM as on previous day"] = None
    df["Actual AUM as on current date"] = df["AUM"]
    df["Daily return"] = None
    df["Derived AUM as on curent day"] = None
    df["Net flows on current day"] = None
    
    for scheme_code, group in df.groupby("Scheme Code"):
        indices = group.index
        # AUM carried forward for the next day's baseline. On a stale or missing
        # day this holds the mark-to-market value rather than the repeated one,
        # so the following day is measured against where the fund actually was.
        # Zeroing the stale day alone would just push its error onto the next.
        eff_prev = None
        # A day whose AUM never updated leaves the next day's change spanning
        # two days, so that day cannot be attributed either. Both are reported
        # as zero and normal measurement resumes on the day after, anchored to
        # a real published AUM.
        zero_next = False
        for idx_in_group, idx in enumerate(indices):
            if idx_in_group == 0:
                eff_prev = df.at[idx, "AUM"]
                continue
            prev_idx = indices[idx_in_group - 1]
            
            nav_curr = df.at[idx, "NAV"]
            nav_prev = df.at[prev_idx, "NAV"]
            aum_prev = eff_prev if eff_prev is not None and pd.notna(eff_prev) else df.at[prev_idx, "AUM"]
            aum_curr = df.at[idx, "AUM"]
            
            df.at[idx, "Closing AUM as on previous day"] = aum_prev
            
            if pd.notna(nav_curr) and pd.notna(nav_prev) and nav_prev != 0:
                daily_return = (nav_curr - nav_prev) / nav_prev
                df.at[idx, "Daily return"] = daily_return * 100
            else:
                daily_return = None
                
            if pd.notna(aum_prev) and daily_return is not None:
                derived_aum = aum_prev * (1 + daily_return)
                df.at[idx, "Derived AUM as on curent day"] = derived_aum
            else:
                derived_aum = None
                
            # AMFI does not always publish a fresh AUM. When it repeats the
            # previous day's figure verbatim while the NAV has moved, the AUM
            # is stale rather than genuinely flat: an unchanged AUM against a
            # moved NAV is arithmetically impossible. Charging the whole
            # mark-to-market move as a subscription is how a 69,849cr AUM
            # served twice produced a 913cr phantom flow, so book zero and let
            # the next real AUM pick the movement up.
            aum_unchanged = (
                pd.notna(aum_prev) and pd.notna(aum_curr)
                and float(aum_prev) == float(aum_curr)
                and daily_return is not None
                and abs(daily_return * 100) > STALE_AUM_MIN_MOVE_PCT
            )

            # A zero previous-day AUM cannot anchor a flow. The formula would
            # measure the whole of today's AUM as a subscription -- a fund with
            # 3,500cr and no published AUM yesterday would report a 3,500cr
            # inflow. Zero here means "not reported", not "the fund was empty".
            prev_is_zero = pd.notna(aum_prev) and float(aum_prev) == 0.0
            curr_is_zero = pd.notna(aum_curr) and float(aum_curr) == 0.0

            unmeasurable = curr_is_zero or prev_is_zero or aum_unchanged or derived_aum is None                 or pd.isna(aum_curr)

            if zero_next and not unmeasurable:
                # The day after an unmeasurable one. Its AUM change covers both
                # days, so attributing it to today would be a guess; report zero
                # and anchor tomorrow to today's real AUM.
                df.at[idx, "Net flows on current day"] = 0.0
                eff_prev = aum_curr
                zero_next = False
            elif unmeasurable:
                df.at[idx, "Net flows on current day"] = 0.0
                zero_next = True
                # Anchor to whatever real AUM exists; the next day is zeroed
                # regardless, so this only sets the baseline for the day after.
                eff_prev = aum_curr if (pd.notna(aum_curr) and float(aum_curr) != 0.0) else eff_prev
            else:
                df.at[idx, "Net flows on current day"] = aum_curr - derived_aum
                eff_prev = aum_curr


    # Filter only target dates
    start_date_ts = pd.to_datetime(start_date)
    df = df[df["NAV Date_parsed"] >= start_date_ts].reset_index(drop=True)
    df = df.drop(columns=["NAV Date_parsed"])
    
    df = df.rename(columns={"NAV": "NAVs"})
    
    flow_cols = [
        "NAV Date",
        "NAVs",
        "Closing AUM as on previous day",
        "Actual AUM as on current date",
        "Daily return",
        "Derived AUM as on curent day",
        "Net flows on current day"
    ]
    
    if "NAV Date" in df.columns:
        parsed = parse_amfi_date_series(df["NAV Date"])
        df["NAV Date"] = parsed.dt.strftime("%d-%m-%Y")
        
    final_cols = list(meta_cols) + flow_cols
    for col in final_cols:
        if col not in df.columns:
            df[col] = None
            
    df["NAV Date_parsed"] = parse_amfi_date_series(df["NAV Date"])
    df = df.sort_values(by=["Asset Class", "Scheme Name", "NAV Date_parsed"]).reset_index(drop=True)
    df = df.drop(columns=["NAV Date_parsed"])
    
    return df[final_cols]


def parse_bucket_input(uploaded_file=None, data_editor_df=None) -> pd.DataFrame:
    """Parse Excel/CSV uploads or dynamic table data editor inputs for the portfolio bucket composition."""
    if uploaded_file is not None:
        try:
            if uploaded_file.name.endswith(".csv"):
                df = pd.read_csv(uploaded_file)
            else:
                df = pd.read_excel(uploaded_file)
            
            # Match columns dynamically
            isin_col = None
            weight_col = None
            name_col = None
            
            for c in df.columns:
                c_low = str(c).lower()
                if "isin" in c_low:
                    isin_col = c
                elif "weight" in c_low or "wt" in c_low or "share" in c_low:
                    weight_col = c
                elif "scheme" in c_low or "name" in c_low or "fund" in c_low:
                    name_col = c
            
            if isin_col is None:
                st.error("Uploaded file must contain an 'ISIN' column.")
                return pd.DataFrame()
            
            res_df = pd.DataFrame()
            res_df["ISIN"] = df[isin_col].astype(str).str.strip().str.upper()
            if name_col is not None:
                res_df["Scheme Name"] = df[name_col].astype(str).str.strip()
            else:
                res_df["Scheme Name"] = res_df["ISIN"]
                
            if weight_col is not None:
                clean_series = df[weight_col].astype(str).str.replace("%", "").str.strip()
                parsed_vals = pd.to_numeric(clean_series, errors="coerce").fillna(0.0)
                # If weights are formatted/saved as decimals (e.g. 0.08, 0.089 instead of 8.0, 8.9),
                # they will have max <= 1.0 and sum <= 1.05. Convert them to percentage format.
                if parsed_vals.max() <= 1.0 and parsed_vals.sum() <= 1.05 and parsed_vals.sum() > 0:
                    res_df["Weight (%)"] = parsed_vals * 100.0
                else:
                    res_df["Weight (%)"] = parsed_vals
            else:
                res_df["Weight (%)"] = 100.0 / len(res_df)
                
            return res_df.dropna(subset=["ISIN"])
        except Exception as e:
            st.error(f"Error parsing uploaded file: {e}")
            return pd.DataFrame()
            
    if data_editor_df is not None:
        df = data_editor_df.copy()
        if "ISIN" not in df.columns:
            return pd.DataFrame()
        df["ISIN"] = df["ISIN"].astype(str).str.strip().str.upper()
        df["Weight (%)"] = pd.to_numeric(df.get("Weight (%)", 0.0), errors="coerce").fillna(0.0)
        if "Scheme Name" not in df.columns:
            df["Scheme Name"] = df["ISIN"]
        return df.dropna(subset=["ISIN"])
        
    return pd.DataFrame()


def fetch_latest_navs(isin_list: List[str]) -> dict:
    """Fetch the latest (today's) NAV for each ISIN from AMFI's live NAVAll.txt feed.

    Returns dict: {ISIN_upper: {"nav": float, "date": str, "scheme_name": str}}
    """
    isin_set = {i.strip().upper() for i in isin_list if i.strip()}
    if not isin_set:
        return {}

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
    }
    try:
        resp = requests.get(AMFI_LATEST_URL, headers=headers, timeout=30)
        resp.raise_for_status()
    except Exception:
        return {}

    result = {}
    for line_bytes in resp.text.splitlines():
        line = line_bytes.strip()
        if ";" not in line or line.startswith("Scheme Code"):
            continue
        parts = [p.strip() for p in line.split(";")]
        if len(parts) < 6:
            continue

        isin_g = parts[1].upper() if parts[1] != "-" else ""
        isin_r = parts[2].upper() if parts[2] != "-" else ""
        matched_isin = None
        if isin_g in isin_set:
            matched_isin = isin_g
        elif isin_r in isin_set:
            matched_isin = isin_r
        else:
            continue

        nav_val = pd.to_numeric(parts[4].replace(",", ""), errors="coerce")
        nav_date = parts[5]  # e.g. "07-Jul-2026"
        scheme_name = parts[3]

        if pd.notna(nav_val) and matched_isin:
            result[matched_isin] = {
                "nav": float(nav_val),
                "date": nav_date,
                "scheme_name": scheme_name,
            }
    return result



def fetch_index_level_series(index_name: str, start_date, end_date) -> dict:
    """Historical level series for a benchmark index, keyed 'dd-mm-YYYY'.

    Tries niftyindices.com first, then falls back to the AMFI index-fund proxy
    in benchmark_proxy.py. Returns {} when neither source has data — callers
    must handle that rather than receive invented numbers.
    """
    direct = fetch_nifty_data_series(index_name, start_date, end_date)
    if direct:
        return direct

    # Imported lazily: benchmark_proxy imports from this module.
    try:
        from benchmark_proxy import fetch_benchmark_levels
    except ImportError:
        return {}

    series = fetch_benchmark_levels([index_name], start_date, end_date).get(index_name)
    if series is None or not series.levels:
        return {}
    return {d.strftime("%d-%m-%Y"): v for d, v in series.levels.items()}


def fetch_nifty_data_series(index_name: str, start_date, end_date) -> dict:
    """Fetch historical close prices for a Nifty index from niftyindices.com.

    Returns a dictionary mapping date string 'dd-mm-YYYY' to float close price,
    or {} when the feed is unavailable. Prefer fetch_index_level_series, which
    adds the AMFI proxy fallback.
    """
    import json
    import requests
    from datetime import datetime, timedelta
    
    url = "https://www.niftyindices.com/Backpage.aspx/getHistoricaldatatabletoString"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36",
        "Content-Type": "application/json; charset=UTF-8",
        "Referer": "https://www.niftyindices.com/reports/historical-data",
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "X-Requested-With": "XMLHttpRequest"
    }
    
    session = requests.Session()
    session.headers.update(headers)
    try:
        session.get("https://www.niftyindices.com/reports/historical-data", timeout=5)
    except Exception:
        pass
        
    date_map = {}
    
    # Fetch in chunks of max 365 days
    chunks = []
    curr = start_date
    while curr <= end_date:
        next_end = min(curr + timedelta(days=364), end_date)
        chunks.append((curr, next_end))
        curr = next_end + timedelta(days=1)
        
    success_any = False
    for c_start, c_end in chunks:
        payload_data = {
            "name": index_name,
            "startDate": c_start.strftime("%d-%b-%Y"),
            "endDate": c_end.strftime("%d-%b-%Y"),
            "indexName": index_name
        }
        data = {"cinfo": json.dumps(payload_data)}
        
        try:
            resp = session.post(url, json=data, timeout=10)
            if resp.status_code == 200:
                json_resp = resp.json()
                if 'd' in json_resp and json_resp['d']:
                    raw_list = json.loads(json_resp['d'])
                    for item in raw_list:
                        raw_date = item.get("HistoricalDate") or item.get("date")
                        raw_close = item.get("CLOSE") or item.get("close")
                        if raw_date and raw_close is not None:
                            try:
                                clean_dt_str = raw_date.replace("-", " ")
                                dt = datetime.strptime(clean_dt_str, "%d %b %Y")
                                date_key = dt.strftime("%d-%m-%Y")
                                val = float(str(raw_close).replace(",", "").strip())
                                if val > 0:
                                    date_map[date_key] = val
                                    success_any = True
                            except Exception:
                                continue
        except Exception as e:
            print(f"Error fetching Nifty index {index_name} for chunk {c_start} to {c_end}: {e}")
            
    if not success_any:
        # Return nothing rather than inventing a series. This used to fall back
        # to a randomly simulated price path, which meant every index return and
        # every excess-return figure downstream was fabricated without warning.
        # Callers must treat an empty dict as "no benchmark data available".
        print(f"Failed to fetch index data for {index_name} — returning no data.")
        return {}

    return date_map


def run_live_portfolio(bucket_df: pd.DataFrame, start_date, initial_amount: float, skip_sunday: bool):
    """Run a live portfolio tracker from start_date to today.

    Fetches historical NAVs from AMFI history API, then overlays the latest
    real-time NAV from NAVAll.txt so the portfolio value is always current.
    """
    end_date = datetime.today().date()

    if bucket_df.empty:
        return None, "Bucket composition is empty."
        
    bucket = bucket_df.copy()
    bucket = bucket[bucket["ISIN"].str.strip() != ""]
    if bucket.empty:
        return None, "Bucket composition has no valid ISINs."
        
    total_weight = bucket["Weight (%)"].sum()
    if total_weight <= 0:
        return None, "Total weight must be greater than 0%."
        
    bucket["Weight_Normalized"] = bucket["Weight (%)"] / total_weight
    
    # Start fetch 10 days early to cover holidays and provide carry-forward baseline
    fetch_start = start_date - timedelta(days=10)
    
    # Fetch Nifty 50 and Nifty 500 history
    n50_raw = fetch_index_level_series("Nifty 50", fetch_start, end_date)
    n500_raw = fetch_index_level_series("Nifty 500", fetch_start, end_date)
    
    # Build complete chronological series of Nifty close prices (carrying forward for holidays)
    n50_closes = {}
    n500_closes = {}
    
    n50_last = None
    n500_last = None
    
    curr = fetch_start
    while curr <= end_date:
        date_key = curr.strftime("%d-%m-%Y")
        
        # Nifty 50
        if date_key in n50_raw:
            n50_last = n50_raw[date_key]
        n50_closes[date_key] = n50_last
        
        # Nifty 500
        if date_key in n500_raw:
            n500_last = n500_raw[date_key]
        n500_closes[date_key] = n500_last
        
        curr += timedelta(days=1)
        
    # Backward fill just in case the very first date had no record
    all_days = build_date_cols(fetch_start, end_date, skip_sunday=False)
    for i in range(len(all_days) - 2, -1, -1):
        curr_d, next_d = all_days[i], all_days[i+1]
        if n50_closes.get(curr_d) is None:
            n50_closes[curr_d] = n50_closes.get(next_d)
        if n500_closes.get(curr_d) is None:
            n500_closes[curr_d] = n500_closes.get(next_d)
            
    isin_list = bucket["ISIN"].tolist()
    
    df_raw = fetch_amfi_data_chunked(fetch_start, end_date, isin_list)
    if df_raw.empty:
        return None, "No historical NAV data returned from AMFI for these ISINs. Please check that the date range is valid."
        
    # Standardize dates using robust locale-independent parser
    df_raw["NAV Date"] = parse_amfi_date_series(df_raw["NAV Date"])
    df_raw = df_raw.dropna(subset=["NAV Date", "NAV"])
    
    df_raw["NAV_Date_Str"] = df_raw["NAV Date"].dt.strftime("%d-%m-%Y")
    
    df_pivot = df_raw.pivot_table(index="ISIN Div Payout / ISIN Growth", columns="NAV_Date_Str", values="NAV", aggfunc="first").reset_index()
    df_pivot_r = df_raw.pivot_table(index="ISIN Div Reinvestment", columns="NAV_Date_Str", values="NAV", aggfunc="first").reset_index()
    
    all_dates = build_date_cols(fetch_start, end_date, skip_sunday=False)
    for d in all_dates:
        if d not in df_pivot.columns:
            df_pivot[d] = None
            
    date_objs = sorted([datetime.strptime(d, "%d-%m-%Y") for d in all_dates])
    sorted_cols = [d.strftime("%d-%m-%Y") for d in date_objs]
    
    bucket["ISIN_upper"] = bucket["ISIN"].str.strip().str.upper()
    df_pivot["ISIN_upper"] = df_pivot["ISIN Div Payout / ISIN Growth"].str.strip().str.upper()
    
    df_sim = pd.merge(bucket, df_pivot, on="ISIN_upper", how="left")
    
    # Merge reinvestment columns if payout ISIN was missing
    missing_mask = df_sim[sorted_cols[0]].isna()
    if missing_mask.any() and not df_pivot_r.empty:
        df_pivot_r["ISIN_upper"] = df_pivot_r["ISIN Div Reinvestment"].str.strip().str.upper()
        for idx, row in df_sim[missing_mask].iterrows():
            match_r = df_pivot_r[df_pivot_r["ISIN_upper"] == row["ISIN_upper"]]
            if not match_r.empty:
                for col in sorted_cols:
                    df_sim.at[idx, col] = match_r.iloc[0].get(col)
                    
    # Carry forward chronological left-to-right
    for i in range(1, len(sorted_cols)):
        prev, curr = sorted_cols[i - 1], sorted_cols[i]
        df_sim[curr] = df_sim[curr].fillna(df_sim[prev])
        
    for i in range(len(sorted_cols) - 2, -1, -1):
        curr, next_col = sorted_cols[i], sorted_cols[i + 1]
        df_sim[curr] = df_sim[curr].fillna(df_sim[next_col])
        
    target_date_cols = build_date_cols(start_date, end_date, skip_sunday)
    if not target_date_cols:
        return None, "Target date range contains no trading days (all dates filtered out)."
        
    t0 = target_date_cols[0]
    df_sim = df_sim.dropna(subset=[t0])
    if len(df_sim) < len(bucket):
        missing_isins = set(bucket["ISIN_upper"]) - set(df_sim["ISIN_upper"])
        return None, f"Could not find historical NAV data for ISIN(s): {', '.join(missing_isins)}. Please check the ISINs and date range."
    
    # ── Overlay latest live NAVs from NAVAll.txt ──────────────────────────────
    latest_navs = fetch_latest_navs(isin_list)
    latest_nav_date_str = None
    if latest_navs:
        # Find the latest date from the live feed
        for info in latest_navs.values():
            latest_nav_date_str = info["date"]  # e.g. "27-Jun-2026"
            break
        # Convert to dd-mm-YYYY for column matching
        if latest_nav_date_str:
            parsed_live = _parse_amfi_date_str(latest_nav_date_str)
            if parsed_live:
                live_dt = datetime.strptime(parsed_live, "%Y-%m-%d")
                live_col = live_dt.strftime("%d-%m-%Y")
                # Add column if not already present
                if live_col not in df_sim.columns:
                    df_sim[live_col] = None
                    if live_col not in target_date_cols:
                        target_date_cols.append(live_col)
                # Inject live NAVs
                for idx, row in df_sim.iterrows():
                    isin_upper = row["ISIN_upper"]
                    if isin_upper in latest_navs:
                        df_sim.at[idx, live_col] = latest_navs[isin_upper]["nav"]
                # Carry forward into live column for any missing ISINs
                if len(sorted_cols) > 0:
                    last_hist = sorted_cols[-1]
                    df_sim[live_col] = df_sim[live_col].fillna(df_sim[last_hist])
        
    df_sim["Initial_Allocation"] = initial_amount * df_sim["Weight_Normalized"]
    df_sim["Units"] = df_sim["Initial_Allocation"] / df_sim[t0]
    
    n50_base = n50_closes.get(t0) or 1.0
    n500_base = n500_closes.get(t0) or 1.0
    if n50_base <= 0:
        n50_base = 1.0
    if n500_base <= 0:
        n500_base = 1.0
        
    daily_rows = []
    for d in target_date_cols:
        # Calculate daily Nifty index relative valuations
        n50_c = n50_closes.get(d) or n50_base
        n500_c = n500_closes.get(d) or n500_base
        
        n50_val = round(initial_amount * (n50_c / n50_base), 10)
        n500_val = round(initial_amount * (n500_c / n500_base), 10)
        
        # Calculate fund-wise values
        fund_vals = {}
        total_val = 0.0
        for idx, row in df_sim.iterrows():
            name = row["Scheme Name"] or row["ISIN"]
            nav_t = row[d]
            val_t = row["Units"] * nav_t
            fund_vals[f"{name} (₹)"] = round(val_t, 10)
            total_val += val_t
            
        row_val = {
            "Date": d,
            "Total Portfolio Value (₹)": round(total_val, 10),
            "Nifty 50 Valuation (₹)": n50_val,
            "Nifty 500 Valuation (₹)": n500_val,
        }
        row_val.update(fund_vals)
        daily_rows.append(row_val)
        
    df_tracker = pd.DataFrame(daily_rows)
    df_tracker["Daily Return (%)"] = 0.0
    df_tracker["Cumulative Return (%)"] = 0.0
    
    t0_val = df_tracker.at[0, "Total Portfolio Value (₹)"]
    for i in range(len(df_tracker)):
        curr_val = df_tracker.at[i, "Total Portfolio Value (₹)"]
        df_tracker.at[i, "Cumulative Return (%)"] = round(((curr_val - t0_val) / t0_val) * 100, 2)
        if i > 0:
            prev_val = df_tracker.at[i - 1, "Total Portfolio Value (₹)"]
            df_tracker.at[i, "Daily Return (%)"] = round(((curr_val - prev_val) / prev_val) * 100, 2)
            
    final_val = df_tracker.iloc[-1]["Total Portfolio Value (₹)"]
    abs_gain = final_val - initial_amount
    abs_return = (abs_gain / initial_amount) * 100
    
    peaks = df_tracker["Total Portfolio Value (₹)"].cummax()
    drawdowns = (df_tracker["Total Portfolio Value (₹)"] - peaks) / peaks * 100
    max_dd = drawdowns.min()
    
    best_day = df_tracker["Daily Return (%)"].max()
    worst_day = df_tracker["Daily Return (%)"].min()
    
    # Today's change (last row vs second-to-last)
    todays_change = 0.0
    todays_change_pct = 0.0
    if len(df_tracker) >= 2:
        yesterday_val = df_tracker.iloc[-2]["Total Portfolio Value (₹)"]
        todays_change = final_val - yesterday_val
        todays_change_pct = (todays_change / yesterday_val) * 100 if yesterday_val else 0.0
    
    metrics = {
        "Initial Value": initial_amount,
        "Current Value": final_val,
        "Total Gain/Loss": abs_gain,
        "Total Return (%)": abs_return,
        "Today's Change": todays_change,
        "Today's Change (%)": todays_change_pct,
        "Max Drawdown (%)": max_dd,
        "Best Day Return (%)": best_day,
        "Worst Day Return (%)": worst_day,
        "NAV Date": latest_nav_date_str or "N/A",
    }
    
    comp_stats = []
    for idx, row in df_sim.iterrows():
        isin = row["ISIN"]
        name = row["Scheme Name"]
        weight = row["Weight (%)"]
        initial_alloc = row["Initial_Allocation"]
        units = row["Units"]
        final_nav = row[target_date_cols[-1]]
        final_val_fund = units * final_nav
        fund_return = ((final_nav - row[t0]) / row[t0]) * 100
        comp_stats.append({
            "Scheme Name": name,
            "ISIN": isin,
            "Weight (%)": weight,
            "Initial Allocation (₹)": round(initial_alloc, 10),
            "Units": round(units, 10),
            "Buy NAV (₹)": round(row[t0], 4),
            "Current NAV (₹)": round(final_nav, 4),
            "Current Value (₹)": round(final_val_fund, 10),
            "Return (%)": round(fund_return, 2)
        })
        
    df_comp_stats = pd.DataFrame(comp_stats)
    
    return {
        "tracker": df_tracker,
        "metrics": metrics,
        "composition": df_comp_stats,
        "latest_nav_date": latest_nav_date_str,
    }, None


def style_portfolio_excel(res_dict: dict) -> bytes:
    """Generate a styled corporate Excel workbook for the portfolio simulation."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    df_tracker = res_dict["tracker"]
    df_comp = res_dict["composition"]
    
    wb = openpyxl.Workbook()
    
    # Sheet 1: Valuation Tracker
    ws1 = wb.active
    ws1.title = "Valuation Tracker"
    
    font_name = "Segoe UI"
    h_fill = PatternFill("solid", fgColor="1F497D")
    h_font = Font(name=font_name, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=font_name, size=10)
    border = Border(
        left=Side(style="thin", color="D3D3D3"),
        right=Side(style="thin", color="D3D3D3"),
        top=Side(style="thin", color="D3D3D3"),
        bottom=Side(style="thin", color="D3D3D3")
    )
    even_fill = PatternFill("solid", fgColor="F2F5F8")
    odd_fill = PatternFill("solid", fgColor="FFFFFF")
    center_va = Alignment(horizontal="center", vertical="center")
    right_va = Alignment(horizontal="right", vertical="center")
    left_va = Alignment(horizontal="left", vertical="center")
    
    # Headers
    ws1.row_dimensions[1].height = 28
    for ci, col in enumerate(df_tracker.columns, 1):
        cell = ws1.cell(row=1, column=ci, value=col)
        cell.fill = h_fill
        cell.font = h_font
        cell.border = border
        cell.alignment = center_va
        
    # Data Rows
    for ri, row_data in enumerate(df_tracker.values, 2):
        ws1.row_dimensions[ri].height = 20
        fill = even_fill if ri % 2 == 0 else odd_fill
        for ci, val in enumerate(row_data, 1):
            cell = ws1.cell(row=ri, column=ci, value=val)
            cell.fill = fill
            cell.font = data_font
            cell.border = border
            col_name = df_tracker.columns[ci - 1]
            
            if col_name == "Date":
                cell.alignment = center_va
            elif "Return" in col_name:
                cell.number_format = '0.00"%"'
                cell.alignment = right_va
            else:
                cell.number_format = "0.0000000000"
                cell.alignment = right_va
                
    for col in ws1.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=12)
        ws1.column_dimensions[get_column_letter(col[0].column)].width = max(min(max_len + 3, 55), 12)
        
    # Sheet 2: Bucket Composition
    ws2 = wb.create_sheet(title="Bucket Composition")
    ws2.row_dimensions[1].height = 28
    for ci, col in enumerate(df_comp.columns, 1):
        cell = ws2.cell(row=1, column=ci, value=col)
        cell.fill = h_fill
        cell.font = h_font
        cell.border = border
        cell.alignment = left_va if col in ("Scheme Name", "ISIN") else center_va
        
    for ri, row_data in enumerate(df_comp.values, 2):
        ws2.row_dimensions[ri].height = 20
        fill = even_fill if ri % 2 == 0 else odd_fill
        for ci, val in enumerate(row_data, 1):
            cell = ws2.cell(row=ri, column=ci, value=val)
            cell.fill = fill
            cell.font = data_font
            cell.border = border
            col_name = df_comp.columns[ci - 1]
            
            if col_name in ("Scheme Name", "ISIN"):
                cell.alignment = left_va
            elif "Weight" in col_name or "Return" in col_name:
                if "Weight" in col_name:
                    cell.number_format = '0.0000000000"%"'
                else:
                    cell.number_format = '0.00"%"'
                cell.alignment = right_va
            elif "Units" in col_name:
                cell.number_format = "0.0000000000"
                cell.alignment = right_va
            else:
                cell.number_format = "0.0000000000"
                cell.alignment = right_va
                
    for col in ws2.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=12)
        ws2.column_dimensions[get_column_letter(col[0].column)].width = max(min(max_len + 3, 55), 12)
        
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── Project helpers ─────────────────────────────────────────────────────────
sys.path.append(str(Path(__file__).parent))
from amfi_nav import classify_option_type, classify_plan_type
from ui_theme import (
    inject_custom_css,
    render_app_footer,
    render_hero,
    render_info_card,
    render_section_header,
)

# ─── Page config ─────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="AMFI NAV Fetcher",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ─── Constants ────────────────────────────────────────────────────────────────
AMFI_HISTORY_URL = (
    "https://portal.amfiindia.com/DownloadNAVHistoryReport_Po.aspx"
    "?frmdt={frmdt}&todt={todt}"
)
AMFI_LATEST_URL = "https://portal.amfiindia.com/spages/NAVAll.txt"


class AmfiFetchError(RuntimeError):
    """AMFI's NAV history endpoint did not return usable data for a date range.

    Raised instead of returning an empty frame so that a throttled request is
    never mistaken for "this range genuinely has no NAVs" — that confusion
    silently truncated history and corrupted every return baseline computed
    from it.
    """


# "Dividend" appears in scheme names two very different ways: as the payout
# option (exclude) and as the investment strategy, e.g. "ICICI Prudential
# Dividend Yield Equity Fund - Growth" (keep). Matching the bare word dropped
# every dividend-yield fund from the app.
_IDCW_RE = re.compile(
    r"\bidcw\b|\bincome distribution\b|\bcapital withdrawal\b|\bdividend\b(?!\s*yield)",
    re.IGNORECASE,
)


def is_idcw_scheme(scheme_name: str) -> bool:
    """True when the scheme name denotes an income-distribution (IDCW) option."""
    return bool(_IDCW_RE.search(scheme_name or ""))


# ─── Helpers ─────────────────────────────────────────────────────────────────

def filter_by_isins(df: pd.DataFrame, isin_list: List[str]) -> pd.DataFrame:
    """Keep only rows where either ISIN column matches the target list."""
    mask = df["ISIN Div Payout / ISIN Growth"].isin(isin_list) | df[
        "ISIN Div Reinvestment"
    ].isin(isin_list)
    return df[mask].copy()


def pivot_to_wide(df: pd.DataFrame, date_cols: List[str]) -> pd.DataFrame:
    """Pivot long-format NAV rows to wide format (one column per date)."""
    meta_cols = [
        "Asset Class",
        "Scheme Code",
        "ISIN Div Payout / ISIN Growth",
        "ISIN Div Reinvestment",
        "Scheme Name",
        "Plan Type",
        "Option Type",
    ]

    meta = df[meta_cols].drop_duplicates(subset=["Scheme Code"])
    pivot = (
        df.pivot_table(index="Scheme Code", columns="NAV Date", values="NAV", aggfunc="first")
        .reset_index()
    )
    result = pd.merge(meta, pivot, on="Scheme Code", how="left")

    for d in date_cols:
        if d not in result.columns:
            result[d] = None

    # Keep only requested date columns
    available_dates = [d for d in date_cols if d in result.columns]
    result = result[meta_cols + available_dates]
    return result.sort_values(["Asset Class", "Scheme Name"]).reset_index(drop=True)


def _date_to_dmy(date_str: str) -> str:
    """Convert '29-May-2026' → '29-05-2026'."""
    try:
        return datetime.strptime(date_str, "%d-%b-%Y").strftime("%d-%m-%Y")
    except Exception:
        return date_str


def style_excel(df: pd.DataFrame, date_cols: List[str], is_aum_only: bool = False) -> bytes:
    """Write df to a styled Excel workbook and return as bytes.

    When both NAV and AUM columns are present (columns named like '29-May-2026 (NAV)'
    and '29-May-2026 (AUM)'), the Excel uses a two-row header:
      Row 1: meta cols (merged ↕) | 'NAV' merged across all NAV date cols | 'AUM' merged across all AUM date cols
      Row 2: date labels in dd/mm/yyyy under each group
    Data starts at row 3.
    """
    import openpyxl
    from openpyxl.utils import get_column_letter

    buf = BytesIO()

    font_name   = "Segoe UI"
    h_fill      = PatternFill("solid", fgColor="1F497D")   # dark blue – meta headers
    nav_fill    = PatternFill("solid", fgColor="1E4D8C")   # blue – NAV group header
    aum_fill    = PatternFill("solid", fgColor="145A32")   # green – AUM group header
    date_nav_fill = PatternFill("solid", fgColor="2E75B6") # lighter blue – NAV date row
    date_aum_fill = PatternFill("solid", fgColor="1E8449") # lighter green – AUM date row
    h_font      = Font(name=font_name, size=11, bold=True, color="FFFFFF")
    date_font   = Font(name=font_name, size=10, bold=True, color="FFFFFF")
    data_font   = Font(name=font_name, size=10)
    even_fill   = PatternFill("solid", fgColor="F2F5F8")
    odd_fill    = PatternFill("solid", fgColor="FFFFFF")
    border = Border(
        left=Side(style="thin", color="D3D3D3"),
        right=Side(style="thin", color="D3D3D3"),
        top=Side(style="thin", color="D3D3D3"),
        bottom=Side(style="thin", color="D3D3D3"),
    )
    center_va = Alignment(horizontal="center", vertical="center")
    left_va   = Alignment(horizontal="left",   vertical="center")

    # ── Detect whether we have interleaved (NAV)/(AUM) pairs ─────────────────
    has_nav_aum_pairs = any("(NAV)" in c for c in df.columns)

    if has_nav_aum_pairs:
        # Split columns into: meta, nav_dates, aum_dates (preserving order)
        meta_cols  = [c for c in df.columns if "(NAV)" not in c and "(AUM)" not in c]
        nav_cols   = [c for c in df.columns if c.endswith(" (NAV)")]
        aum_cols   = [c for c in df.columns if c.endswith(" (AUM)")]

        nav_dates  = [c[: -len(" (NAV)")] for c in nav_cols]  # "29-May-2026"
        aum_dates  = [c[: -len(" (AUM)")] for c in aum_cols]

        # Re-order df columns: meta | all NAV cols | all AUM cols
        ordered_cols = meta_cols + nav_cols + aum_cols
        df = df[ordered_cols]

        n_meta = len(meta_cols)
        n_nav  = len(nav_cols)
        n_aum  = len(aum_cols)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "NAV Data"

        ws.row_dimensions[1].height = 28
        ws.row_dimensions[2].height = 22

        # ── Row 1 ─────────────────────────────────────────────────────────────
        # Meta columns: merged vertically (rows 1–2)
        for i, col in enumerate(meta_cols):
            ci = i + 1
            cell = ws.cell(row=1, column=ci, value=col)
            cell.fill = h_fill
            cell.font = h_font
            cell.border = border
            cell.alignment = left_va if col in ("Scheme Name", "Asset Class") else center_va
            ws.merge_cells(start_row=1, start_column=ci, end_row=2, end_column=ci)

        # NAV group label: merged across all NAV date columns
        nav_start = n_meta + 1
        nav_end   = n_meta + n_nav
        if n_nav > 0:
            cell = ws.cell(row=1, column=nav_start, value="NAV")
            cell.fill = nav_fill
            cell.font = h_font
            cell.border = border
            cell.alignment = center_va
            if n_nav > 1:
                ws.merge_cells(start_row=1, start_column=nav_start, end_row=1, end_column=nav_end)
            # Apply border to all cells in the merge range
            for ci in range(nav_start, nav_end + 1):
                ws.cell(row=1, column=ci).border = border

        # AUM group label: merged across all AUM date columns
        aum_start = n_meta + n_nav + 1
        aum_end   = n_meta + n_nav + n_aum
        if n_aum > 0:
            cell = ws.cell(row=1, column=aum_start, value="AUM")
            cell.fill = aum_fill
            cell.font = h_font
            cell.border = border
            cell.alignment = center_va
            if n_aum > 1:
                ws.merge_cells(start_row=1, start_column=aum_start, end_row=1, end_column=aum_end)
            for ci in range(aum_start, aum_end + 1):
                ws.cell(row=1, column=ci).border = border

        # ── Row 2: date labels in dd/mm/yyyy ──────────────────────────────────
        # Meta cells in row 2 are covered by merges, just apply styling
        for i in range(n_meta):
            ci = i + 1
            cell = ws.cell(row=2, column=ci)
            cell.fill = h_fill
            cell.font = h_font
            cell.border = border

        for i, d in enumerate(nav_dates):
            ci = nav_start + i
            cell = ws.cell(row=2, column=ci, value=_date_to_dmy(d))
            cell.fill = date_nav_fill
            cell.font = date_font
            cell.border = border
            cell.alignment = center_va

        for i, d in enumerate(aum_dates):
            ci = aum_start + i
            cell = ws.cell(row=2, column=ci, value=_date_to_dmy(d))
            cell.fill = date_aum_fill
            cell.font = date_font
            cell.border = border
            cell.alignment = center_va

        # ── Data rows (start at row 3) ────────────────────────────────────────
        for ri_data, (_, row_data) in enumerate(df.iterrows()):
            ri = ri_data + 3
            ws.row_dimensions[ri].height = 20
            fill = even_fill if ri % 2 == 0 else odd_fill
            for ci, col in enumerate(ordered_cols, 1):
                cell = ws.cell(row=ri, column=ci)
                val  = row_data[col]
                cell.fill   = fill
                cell.font   = data_font
                cell.border = border
                if col in nav_cols or col in aum_cols:
                    if pd.notna(val) and val is not None:
                        cell.value = val
                        cell.number_format = "0.00" if col in aum_cols else "0.0000"
                    else:
                        cell.value = "—"
                    cell.alignment = center_va
                else:
                    cell.value = val if pd.notna(val) else None
                    cell.alignment = left_va if col in ("Scheme Name", "Asset Class") else center_va

        # ── Column widths ─────────────────────────────────────────────────────
        for col_cells in ws.iter_cols(min_row=1, max_row=ws.max_row):
            max_len = 10
            for c in col_cells:
                try:
                    max_len = max(max_len, len(str(c.value or "")) + (2 if c.row <= 2 else 0))
                except Exception:
                    pass
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = max(min(max_len + 2, 55), 10)

        wb.save(buf)

    else:
        # ── Original single-header path (NAV only, AUM only, or long format) ──
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="NAV Data")
            wb = writer.book
            ws = writer.sheets["NAV Data"]

            h_fill2    = PatternFill("solid", fgColor="1F497D")
            h_font2    = Font(name=font_name, size=11, bold=True, color="FFFFFF")
            data_font2 = Font(name=font_name, size=10)

            ws.row_dimensions[1].height = 28
            for ci, col_name in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=ci)
                cell.fill = h_fill2
                cell.font = h_font2
                cell.border = border
                left_a = col_name in ("Scheme Name", "Asset Class")
                cell.alignment = Alignment(horizontal="left" if left_a else "center", vertical="center")

            for ri in range(2, len(df) + 2):
                ws.row_dimensions[ri].height = 20
                fill = even_fill if ri % 2 == 0 else odd_fill
                for ci, col_name in enumerate(df.columns, 1):
                    cell = ws.cell(row=ri, column=ci)
                    cell.fill = fill
                    cell.font = data_font2
                    cell.border = border
                    if col_name in ("NAV Date", "AUM Date"):
                        if cell.value is None or cell.value == "":
                            cell.value = "—"
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    elif col_name in ("NAV", "NAVs"):
                        if cell.value is not None and cell.value != "":
                            cell.number_format = "0.00" if col_name == "NAVs" else "0.0000"
                        else:
                            cell.value = "—"
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    elif col_name == "Daily return":
                        if cell.value is not None and cell.value != "":
                            cell.number_format = '0.00"%"'
                        else:
                            cell.value = "—"
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    elif col_name in (
                        "Closing AUM as on previous day", 
                        "Actual AUM as on current date", 
                        "Derived AUM as on curent day", 
                        "Net flows on current day",
                        "AUM"
                    ):
                        if cell.value is not None and cell.value != "":
                            cell.number_format = "0.00"
                        else:
                            cell.value = "—"
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    elif col_name in ("Scheme Name", "Asset Class"):
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="center", vertical="center")

            for col in ws.columns:
                max_len = max((len(str(c.value or "")) + (3 if c.row == 1 else 0) for c in col), default=12)
                ws.column_dimensions[get_column_letter(col[0].column)].width = max(min(max_len + 2, 55), 12)

    return buf.getvalue()


def _map_amfi_columns(header_line: str) -> dict:
    """Map AMFI's header row to column indices.

    AMFI has changed this layout at least once -- inserting Plan and Option,
    which shifted the ISINs from columns 2-3 to 4-5 -- so positions are read
    from the header instead of assumed. An unrecognised header falls back to
    the historical layout.
    """
    names = [h.strip().lower() for h in header_line.split(";")]
    out: dict = {}
    for i, h in enumerate(names):
        if h.startswith("scheme code"):
            out.setdefault("scheme_code", i)
        elif "isin" in h and "reinvest" in h:
            out.setdefault("isin_reinvest", i)
        elif "isin" in h:
            out.setdefault("isin_growth", i)
        elif h in ("scheme name", "nav name") or h.endswith(" name"):
            out.setdefault("scheme_name", i)
        elif h == "plan":
            out.setdefault("plan", i)
        elif h == "option":
            out.setdefault("option", i)
        elif "net asset value" in h:
            out.setdefault("nav", i)
        elif h == "date":
            out.setdefault("date", i)

    defaults = {"scheme_code": 0, "scheme_name": 1, "isin_growth": 2,
                "isin_reinvest": 3, "nav": 4, "date": 7}
    for k, v in defaults.items():
        out.setdefault(k, v)
    return out


@st.cache_data(ttl=900, show_spinner=False)
def fetch_amfi_data(frmdt: str, todt: str, isin_list: tuple[str, ...] | None = None,
                    include_direct: bool = False) -> pd.DataFrame:
    """Parse AMFI's historical NAV dump.

    Regular-plan Growth options only by default, which is what the rest of the
    app reports on. Pass include_direct=True for the benchmark index-fund
    proxies, where the Direct plan's lower TER tracks the index more closely.
    """
    url = AMFI_HISTORY_URL.format(frmdt=frmdt, todt=todt)
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    }
    import time

    isin_set = {isin.strip().upper() for isin in isin_list if isin.strip()} if isin_list else None

    def _download_and_parse() -> tuple[list, bool]:
        """Stream one response. Returns (rows, response_was_a_valid_nav_dump)."""
        # (connect, read). AMFI answers a good request in ~25s per month of
        # range or not at all, so a 300s read timeout just parked the app on a
        # dead socket for five minutes per attempt.
        resp = requests.get(url, headers=headers, stream=True, timeout=(15, 150))
        resp.raise_for_status()

        rows: List[dict] = []
        current_section = "Unknown"
        # AMFI answers HTTP 200 with an HTML page when it is throttling or
        # erroring. That page has no column header, which is how we tell a
        # genuine empty range from a failed request.
        header_seen = False
        # Column positions come from the header rather than being hard-coded.
        # AMFI moved the ISINs from columns 2-3 to 4-5 when it inserted Plan
        # and Option, which silently emptied every result: the ISIN filter was
        # reading Plan/Option, matched nothing, and the header check still
        # passed, so a broken parse looked like a successful fetch of an empty
        # date range.
        cols: dict = {}

        for line_bytes in resp.iter_lines():
            if not line_bytes:
                continue
            line = line_bytes.decode('utf-8', errors='ignore')

            # Fast section check (no semicolons in section headers)
            if ";" not in line:
                line_stripped = line.strip()
                if (
                    line_stripped.startswith("Open Ended")
                    or line_stripped.startswith("Closed Ended")
                    or line_stripped.startswith("Interval Fund Schemes")
                ):
                    current_section = line_stripped
                continue

            if not header_seen and line.startswith("Scheme Code"):
                header_seen = True
                cols = _map_amfi_columns(line)
                continue

            # Optimize memory & parsing: Skip processing the line if isin_set is provided
            # and none of the targeted ISINs are in the raw text line (case-insensitive)
            if isin_set:
                line_upper = line.upper()
                if not any(isin in line_upper for isin in isin_set):
                    continue

            parts = line.split(";")
            if len(parts) <= cols.get("date", 7):
                continue

            def _col(key: str, default: str = "") -> str:
                i = cols.get(key)
                return parts[i].strip() if i is not None and i < len(parts) else default

            isin_growth = _col("isin_growth")
            isin_reinvest = _col("isin_reinvest")

            isin_growth_upper = isin_growth.upper() if isin_growth != "-" else ""
            isin_reinvest_upper = isin_reinvest.upper() if isin_reinvest != "-" else ""

            # Double check matching against ISIN set
            if isin_set:
                g_match = isin_growth_upper and isin_growth_upper in isin_set
                r_match = isin_reinvest_upper and isin_reinvest_upper in isin_set
                if not (g_match or r_match):
                    continue

            scheme_code = _col("scheme_code")
            scheme_name = _col("scheme_name")
            # AMFI now publishes Plan and Option as their own columns. Prefer
            # them over inferring the plan and option from the scheme name.
            plan = _col("plan")
            option = _col("option")

            if plan:
                is_direct = "direct" in plan.lower()
            else:
                scheme_name_lower = scheme_name.lower()
                is_direct = "direct" in scheme_name_lower or bool(re.search(r"\bdir\b", scheme_name_lower))

            if option:
                opt = option.lower()
                is_idcw = ("idcw" in opt) or ("dividend" in opt) or ("payout" in opt)
            else:
                is_idcw = is_idcw_scheme(scheme_name)

            if (is_direct and not include_direct) or is_idcw:
                continue

            nav_value = _col("nav")
            nav_date = _col("date")

            scheme_code = scheme_code if scheme_code != "-" else None
            isin_growth_val = isin_growth if isin_growth != "-" else None
            isin_reinvest_val = isin_reinvest if isin_reinvest != "-" else None

            nav = pd.to_numeric(nav_value.replace(",", ""), errors="coerce")

            rows.append(
                {
                    "Asset Class": current_section,
                    "Scheme Code": scheme_code,
                    "ISIN Div Payout / ISIN Growth": isin_growth_val,
                    "ISIN Div Reinvestment": isin_reinvest_val,
                    "Scheme Name": scheme_name,
                    "NAV": nav,
                    "NAV Date": nav_date,
                }
            )

        return rows, (header_seen or bool(rows))

    max_retries = 3
    delay = 3.0
    rows: List[dict] = []
    for attempt in range(max_retries):
        try:
            rows, valid = _download_and_parse()
            if valid:
                break
            reason = "AMFI returned an error page instead of NAV data"
        # RequestException is the base of the lot. AMFI regularly drops a
        # response mid-stream, which surfaces as ChunkedEncodingError -- not a
        # Timeout or ConnectionError, so naming those individually let it
        # escape the retry and abort the whole report.
        except requests.exceptions.RequestException as e:
            reason = f"{type(e).__name__}: {e}"

        if attempt == max_retries - 1:
            # Raise rather than return empty. st.cache_data does not cache
            # exceptions, so a throttled response can't be memoised into a
            # permanent hole in the history — which used to truncate NAV
            # coverage silently and shift every return baseline.
            raise AmfiFetchError(
                f"AMFI NAV history unavailable for {frmdt} to {todt} after "
                f"{max_retries} attempts ({reason})."
            )
        time.sleep(delay)
        delay *= 2

    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows)
    df["Plan Type"] = df["Scheme Name"].apply(classify_plan_type)
    df["Option Type"] = df["Scheme Name"].apply(classify_option_type)
    return df


def fetch_amfi_data_chunked(start_date, end_date, isin_list: List[str] | None = None,
                            include_direct: bool = False, return_gaps: bool = False,
                            progress=None, budget_seconds: float = 300.0,
                            chunk_days: int = 90):
    """Fetch AMFI data by chunking the date range into 90-day intervals to prevent timeout & memory crash.

    With return_gaps=True returns (df, gaps) where gaps lists the (start, end)
    windows AMFI would not serve, so callers can tell a partial history from a
    complete one instead of computing returns off a truncated series.

    ``progress`` is an optional callable(done, total, label) for UI feedback --
    without it a slow fetch is an unexplained spinner. ``budget_seconds`` caps
    the whole operation: AMFI can be slow enough that retrying every chunk runs
    for the better part of an hour, and an unbounded wait is worse than a
    reported gap.
    """
    import time
    passed_isins = tuple(isin_list) if isin_list else None
    dfs = []
    gaps: List[Tuple[date, date]] = []

    windows: List[Tuple[date, date]] = []
    cursor = start_date
    while cursor <= end_date:
        w_end = min(cursor + timedelta(days=chunk_days), end_date)
        windows.append((cursor, w_end))
        cursor = w_end + timedelta(days=1)

    started = time.monotonic()
    total = len(windows)

    def _remaining() -> float:
        return budget_seconds - (time.monotonic() - started)

    for i, (w_start, w_end) in enumerate(windows):
        if progress:
            progress(i, total, f"{w_start:%b %Y} – {w_end:%b %Y}")

        if _remaining() <= 0:
            # Out of budget: record the rest as gaps rather than grinding on.
            gaps.extend(windows[i:])
            print(f"AMFI fetch budget exhausted; {total - i} window(s) left unfetched.")
            break

        try:
            df_chunk = fetch_amfi_data(
                w_start.strftime("%d-%b-%Y"), w_end.strftime("%d-%b-%Y"),
                passed_isins, include_direct,
            )
            if not df_chunk.empty:
                dfs.append(df_chunk)
        except AmfiFetchError as exc:
            print(f"AMFI chunk unavailable: {exc}")
            gaps.append((w_start, w_end))

        if i < total - 1:
            time.sleep(0.5)  # avoid hitting rate limits

    # One more pass at whatever AMFI refused, but only with time left to spend.
    # Throttling is transient, so a retry after the other chunks often works.
    if gaps and _remaining() > 30:
        if progress:
            progress(total, total, f"retrying {len(gaps)} unavailable window(s)")
        time.sleep(5)
        still_missing: List[Tuple[date, date]] = []
        for g_start, g_end in gaps:
            if _remaining() <= 0:
                still_missing.append((g_start, g_end))
                continue
            try:
                df_chunk = fetch_amfi_data(
                    g_start.strftime("%d-%b-%Y"), g_end.strftime("%d-%b-%Y"),
                    passed_isins, include_direct,
                )
                if not df_chunk.empty:
                    dfs.append(df_chunk)
            except AmfiFetchError as exc:
                print(f"AMFI chunk still unavailable on retry: {exc}")
                still_missing.append((g_start, g_end))
            time.sleep(1)
        gaps = still_missing

    if progress:
        progress(total, total, "done")

    df_all = pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame()
    if return_gaps:
        return df_all, gaps
    return df_all


def build_date_cols(start_date, end_date, skip_sunday: bool = True, skip_saturday: bool = False) -> List[str]:
    """Generate sorted list of date strings between start and end."""
    cols = []
    current = start_date
    while current <= end_date:
        if (skip_saturday and current.weekday() == 5) or (skip_sunday and current.weekday() == 6):
            current += timedelta(days=1)
            continue
        cols.append(current.strftime("%d-%m-%Y"))
        current += timedelta(days=1)
    return cols


# ─── Main ─────────────────────────────────────────────────────────────────────


def main():
    import app
    app.main()


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        st.error(f"A fatal error occurred on startup: {exc}")
        st.exception(exc)
