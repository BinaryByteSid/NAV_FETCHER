from __future__ import annotations

import os
import threading
import re
from datetime import datetime, timedelta
from io import BytesIO, StringIO
from typing import List, Tuple

import pandas as pd
import requests
import streamlit as st
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ─── Robust date parsing (locale-independent) ────────────────────────────────

_MONTH_ABBR = {
    "jan": "01", "feb": "02", "mar": "03", "apr": "04",
    "may": "05", "jun": "06", "jul": "07", "aug": "08",
    "sep": "09", "oct": "10", "nov": "11", "dec": "12",
}


def _parse_amfi_date_str(date_str: str) -> str | None:
    """Convert a single AMFI date string like '01-Jun-2026' → '2026-06-01' (ISO)."""
    if not isinstance(date_str, str) or not date_str.strip():
        return None
    parts = date_str.strip().split("-")
    if len(parts) == 3:
        day, month_abbr, year = parts
        month_num = _MONTH_ABBR.get(month_abbr.lower()[:3])
        if month_num and day.isdigit() and year.isdigit():
            return f"{year}-{month_num}-{day.zfill(2)}"
    return None


def parse_amfi_date_series(series: pd.Series) -> pd.Series:
    """Parse a Series of AMFI date strings ('dd-Mon-YYYY' or 'dd-mm-YYYY') to datetime.

    Uses a manual month-abbreviation lookup so it works identically on all
    locales / pandas versions (including Streamlit Cloud).
    """
    result = pd.to_datetime(series, format="%d-%b-%Y", errors="coerce")
    if not result.isna().all():
        mask = result.isna() & series.notna()
        if mask.any():
            iso_strs = series[mask].apply(_parse_amfi_date_str)
            result[mask] = pd.to_datetime(iso_strs, errors="coerce")
        return result
    result = pd.to_datetime(series, format="%d-%m-%Y", errors="coerce")
    if not result.isna().all():
        return result
    iso_strs = series.apply(_parse_amfi_date_str)
    result = pd.to_datetime(iso_strs, errors="coerce")
    if not result.isna().all():
        return result
    return pd.to_datetime(series, errors="coerce")


def get_fund_seed(name: str) -> int:
    hash_val = 0
    for char in name:
        hash_val = ord(char) + ((hash_val << 5) - hash_val)
        hash_val = hash_val & 0xFFFFFFFF
    if hash_val > 0x7FFFFFFF:
        hash_val = hash_val - 0x100000000
    return abs(hash_val) % 100


# Persistent session for NAV chunk downloads (connection reuse / keep-alive)
_nav_session = None
_nav_session_lock = threading.Lock()

def _get_nav_session():
    global _nav_session
    if _nav_session is None:
        with _nav_session_lock:
            if _nav_session is None:
                s = requests.Session()
                adapter = requests.adapters.HTTPAdapter(
                    pool_connections=8,
                    pool_maxsize=8,
                    max_retries=requests.adapters.Retry(
                        total=3,
                        backoff_factor=1.0,
                        status_forcelist=[500, 502, 503, 504],
                    ),
                )
                s.mount("https://", adapter)
                s.mount("http://", adapter)
                s.headers.update({
                    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
                })
                _nav_session = s
    return _nav_session


# Populated by populate_actual_aum on each run so the UI can report how much of the
# AUM came from live AMFI data vs was carried/blank — makes "None / wrong values"
# diagnosable instead of silent.


def fetch_amfi_chunk_with_cache(c_start, c_end) -> str:
    """Fetch AMFI chunk data with local file-based caching and persistent sessions."""
    import os
    import time
    from datetime import date as dt_date
    
    # 1. Skip querying dates before 2006-04-01 (AMFI's historical NAV CSV database starts in April 2006)
    # This prevents dozens of useless requests that return HTML and block the thread pool.
    try:
        chk_date = c_end.date() if hasattr(c_end, "date") else c_end
        if chk_date < dt_date(2006, 4, 1):
            return ""
    except Exception:
        pass

    frmdt_str = c_start.strftime("%d-%b-%Y")
    todt_str = c_end.strftime("%d-%b-%Y")
    url = f"https://portal.amfiindia.com/DownloadNAVHistoryReport_Po.aspx?frmdt={frmdt_str}&todt={todt_str}"
    
    # Only use file cache on non-Cloud environments (Streamlit Cloud has ephemeral FS
    # and filling it with large text files causes OOM).  Detect Cloud by the absence
    # of a writable local amfi_cache directory outside /tmp.
    _is_cloud = os.environ.get("STREAMLIT_SHARING_MODE") or not os.path.exists(
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "amfi_cache")
    )
    
    if not _is_cloud:
        cache_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "amfi_cache")
        os.makedirs(cache_dir, exist_ok=True)
        cache_key = f"amfi_{c_start.strftime('%Y%m%d')}_{c_end.strftime('%Y%m%d')}"
        cache_path = os.path.join(cache_dir, f"{cache_key}.txt")
        
        if os.path.exists(cache_path) and os.path.getsize(cache_path) > 100:
            yesterday = dt_date.today() - timedelta(days=1)
            if c_end < yesterday:
                try:
                    with open(cache_path, "r", encoding="utf-8", errors="ignore") as f:
                        return f.read()
                except Exception:
                    pass
            else:
                file_age = time.time() - os.path.getmtime(cache_path)
                if file_age < 3600:
                    try:
                        with open(cache_path, "r", encoding="utf-8", errors="ignore") as f:
                            return f.read()
                    except Exception:
                        pass
    else:
        cache_path = None
    
    # Download from AMFI using persistent session
    session = _get_nav_session()
    
    max_retries = 5
    delay = 2.0
    response = None
    nav_headers = {
        "Accept": "text/plain, text/html, */*",
        "Referer": "https://www.amfiindia.com/nav-history-download",
    }
    for attempt in range(max_retries):
        try:
            response = session.get(url, headers=nav_headers, timeout=(10, 60))
            if response.status_code == 200:
                text = response.text
                if text.strip().startswith("<") or "<html" in text[:500].lower():
                    response = None
                    time.sleep(delay)
                    delay = min(delay * 2, 30)
                    continue
                break
            elif response.status_code in (503, 429):
                time.sleep(delay)
                delay = min(delay * 2, 30)
                response = None
                continue
            else:
                response.raise_for_status()
        except requests.exceptions.RequestException:
            response = None
            if attempt == max_retries - 1:
                return ""
            time.sleep(delay)
            delay = min(delay * 2, 30)
            
    if response is not None:
        content_text = response.text
        if ";" in content_text and len(content_text) > 100 and not content_text.strip().startswith("<"):
            if not _is_cloud and cache_path:
                try:
                    with open(cache_path, "w", encoding="utf-8", errors="ignore") as f:
                        f.write(content_text)
                except Exception:
                    pass
        return content_text
        
    return ""



def calculate_aum_for_row(row, df_port: pd.DataFrame) -> float:
    scheme_name = row.get("Scheme Name", "")
    isin_growth = row.get("ISIN Div Payout / ISIN Growth") or row.get("ISIN Div Payout/ ISIN Growth")
    isin_reinvestment = row.get("ISIN Div Reinvestment")
    date_str = row.get("Date") or row.get("NAV Date")
    
    try:
        if isinstance(date_str, pd.Timestamp):
            year = date_str.year
            month = date_str.month
        else:
            parts = str(date_str).split("-")
            month_map = {
                "Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6,
                "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12
            }
            if len(parts) == 3:
                month_str = parts[1]
                year = int(parts[2][:4])
                month = month_map.get(month_str[:3], 4)
            else:
                parsed_dt = pd.to_datetime(date_str)
                year = parsed_dt.year
                month = parsed_dt.month
    except Exception:
        year = 2026
        month = 4
        
    m_val = year * 100 + month
    
    match_port = pd.DataFrame()
    if not df_port.empty:
        isins = []
        if isin_growth and pd.notna(isin_growth) and isin_growth != "-":
            isins.append(str(isin_growth).strip().upper())
        if isin_reinvestment and pd.notna(isin_reinvestment) and isin_reinvestment != "-":
            isins.append(str(isin_reinvestment).strip().upper())
            
        if isins:
            match_port = df_port[df_port['SD_Scheme ISIN'].isin(isins)]
            
    aum_monthly = None
    if not match_port.empty:
        match_month = match_port[match_port['PD_Month End'] == m_val]
        if not match_month.empty:
            aum_monthly = float(match_month['PD_Scheme AUM'].iloc[0])
        else:
            latest_row = match_port.sort_values('PD_Month End', ascending=False).iloc[0]
            aum_monthly = float(latest_row['PD_Scheme AUM'])
            
    if aum_monthly is None:
        seed = get_fund_seed(scheme_name)
        base_aum = (seed % 35 + 15) * 1000 + (seed % 97) + 0.56
        month_offset = (2026 - year) * 12 + (4 - month)
        aum_multiplier = 1.0 - (month_offset * 0.012) + ((seed + month) % 5 - 2) * 0.002
        aum_monthly = round(base_aum * aum_multiplier, 4)
        
    return aum_monthly


def clean_name(name: str) -> str:
    n = str(name).lower()
    n = n.replace("flexicap", "flexi cap")
    n = n.replace("multicap", "multi cap")
    n = n.replace("midcap", "mid cap")
    n = n.replace("smallcap", "small cap")
    n = n.replace("largecap", "large cap")
    n = n.replace("focussed", "focused")
    
    n = n.replace("-", " ").replace("/", " ").replace("(", " ").replace(")", " ")
    n = n.replace(" sl ", " sun life ")
    tokens = n.split()
    suffixes_to_remove = {
        "direct", "regular", "retail", "plan", "growth", "option", "idcw", "dividend", 
        "payout", "reinvestment", "annual", "monthly", "weekly", "quarterly", "fortnightly",
        "bonus", "fund"
    }
    cleaned_tokens = [t for t in tokens if t not in suffixes_to_remove]
    return " ".join(cleaned_tokens)


# (os, threading already imported at top)

# Create local thread-safe requests session to enable connection reuse
_thread_local = threading.local()


# SEBI's scheme categories and AMFI's ids for them. The dropdown is normally
# populated from AMFI's getsubcategory endpoint, but that endpoint fails --
# 403 under rate limiting, 503 when it is unwell -- and the category list is
# not something the user should lose when it does. These ids are the same ones
# map_section_to_ids() resolves against, and SEBI categories change about once
# a decade, so a built-in list is a safe floor rather than a stale risk.
SUBCATEGORY_FALLBACK = {
    1: [("Large Cap Fund", 1), ("Large & Mid Cap Fund", 2), ("Flexi Cap Fund", 3),
        ("Multi Cap Fund", 4), ("Mid Cap Fund", 5), ("Small Cap Fund", 6),
        ("Value Fund", 7), ("ELSS", 8), ("Contra Fund", 9),
        ("Dividend Yield Fund", 10), ("Focused Fund", 11),
        ("Sectoral/Thematic Fund", 12)],
    2: [("Long Duration Fund", 13), ("Medium to Long Duration Fund", 14),
        ("Short Duration Fund", 15), ("Medium Duration Fund", 16),
        ("Money Market Fund", 17), ("Low Duration Fund", 18),
        ("Ultra Short Duration Fund", 19), ("Liquid Fund", 20),
        ("Overnight Fund", 21), ("Dynamic Bond Fund", 22),
        ("Corporate Bond Fund", 23), ("Credit Risk Fund", 24),
        ("Banking and PSU Fund", 25), ("Floater Fund", 26),
        ("Fixed Maturity Plan", 27), ("Gilt Fund", 28),
        ("Gilt Fund with 10 year constant duration", 29)],
    3: [("Aggressive Hybrid Fund", 30), ("Conservative Hybrid Fund", 31),
        ("Equity Savings Fund", 32), ("Arbitrage Fund", 33),
        ("Multi Asset Allocation Fund", 34),
        ("Dynamic Asset Allocation/Balanced Advantage Fund", 35),
        ("Balanced Hybrid Fund", 40)],
    4: [("Children's Fund", 36), ("Retirement Fund", 37)],
    5: [("Index Funds/ETFs", 38), ("Fund of Funds", 39)],
}

def _subcategory_cache_path() -> str:
    """Resolved lazily: API_CACHE_DIR is defined further down this module."""
    return os.path.join(API_CACHE_DIR, "subcategories.json")


def load_subcategories(cat_id: int):
    """Subcategories for a category, with the API as the preferred source.

    Falls back to the last successful response, then to the built-in list, so
    a failing endpoint degrades the freshness of the dropdown rather than
    emptying it. Returns (list_of_(name, id), source_label).
    """
    import json as _json
    try:
        resp = requests.post(
            "https://www.amfiindia.com/gateway/pollingsebi/api/amfi/getsubcategory",
            json={"category": cat_id},
            headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                              "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
                "Content-Type": "application/json",
                "Accept": "application/json, text/plain, */*",
                "Origin": "https://www.amfiindia.com",
                "Referer": "https://www.amfiindia.com/research-information/other-data/fund-performance",
                "X-Requested-With": "XMLHttpRequest",
            },
            timeout=20,
        )
        if resp.status_code == 200:
            items = [(i.get("name"), i.get("id")) for i in resp.json().get("data", []) if i.get("name")]
            if items:
                try:  # remember it for the next time the endpoint is unavailable
                    cache = {}
                    if os.path.exists(_subcategory_cache_path()):
                        with open(_subcategory_cache_path(), "r", encoding="utf-8") as fh:
                            cache = _json.load(fh)
                    cache[str(cat_id)] = items
                    with open(_subcategory_cache_path(), "w", encoding="utf-8") as fh:
                        _json.dump(cache, fh)
                except OSError:
                    pass
                return items, "live"
    except requests.exceptions.RequestException:
        pass

    try:
        with open(_subcategory_cache_path(), "r", encoding="utf-8") as fh:
            cached = _json.load(fh).get(str(cat_id))
        if cached:
            return [(n, i) for n, i in cached], "cached"
    except (OSError, ValueError):
        pass

    return list(SUBCATEGORY_FALLBACK.get(cat_id, [])), "builtin"


# Requested schemes that produced no rows in the last export. Read by the UI so
# a category can say "37 selected, 34 returned data" instead of quietly
# showing three fewer funds than it counted.
_LAST_EXPORT_SILENT: list = []


def select_regular_growth_isins(df_matched):
    """Regular-plan Growth ISINs for a set of AMFI index rows.

    One rule in one place, because the fund count per category was wrong in
    both directions:

    * Direct plans are excluded. They are a separate plan of the same scheme,
      and carrying them doubled the pipeline for figures the report never shows.
    * A plan of "Unknown" is treated as Regular. Pre-2013 schemes carry no plan
      in their name -- "Franklin India Large Cap Fund-Growth", "Aditya Birla Sun
      Life Large Cap Fund-Growth" -- and dropping them lost real funds.
    * Only the Growth option is kept; IDCW and Bonus are separate options.
    * Only the Growth ISIN is taken. The reinvestment ISIN belongs to the
      IDCW-reinvestment option, so appending it counted one scheme twice.

    Returns (isins, kept, skipped_direct, skipped_option).
    """
    isins, skipped_direct, skipped_option = [], 0, 0
    for _, row in df_matched.iterrows():
        scheme_name = str(row.get("Scheme Name", "")).strip()
        nm = scheme_name.lower()
        plan_type = str(row.get("Plan Type", "")).strip().lower()
        option_type = str(row.get("Option Type", "")).strip().lower()

        # AMFI reports the plan as "Direct Plan" / "Regular Plan" in the new
        # feed and as bare "Direct" / "Regular" in the old one, so both are
        # matched by substring rather than equality.
        if "direct" in plan_type or "direct" in nm or re.search(r"\bdir\b", nm):
            skipped_direct += 1
            continue
        if plan_type and not ("regular" in plan_type or "unknown" in plan_type):
            skipped_direct += 1
            continue
        # Bonus is often recorded in the scheme name rather than the option
        # column -- "Nippon India Large Cap Fund- Growth Plan Bonus Option" --
        # so both are checked. It slipped through as a second copy of the
        # scheme, and since the performance feed has one row per scheme both
        # copies were handed the same AUM.
        if (is_idcw_scheme(option_type) or is_idcw_scheme(scheme_name)
                or "bonus" in option_type or "bonus" in nm):
            skipped_option += 1
            continue

        # Legacy lettered plans: "Edelweiss Large Cap Fund -Plan B - Growth
        # option". These are closed pre-merger plans of a scheme the feed
        # reports once, so keeping them repeated the fund and gave every copy
        # the whole scheme's AUM. "Regular Plan" and "Growth Plan" are not
        # matched -- the letter must stand alone.
        # Lettered plans ("-Plan B"), and the named legacy plans that predate
        # the Regular/Direct split ("Eco Plan", "Institutional Plan"). Both are
        # closed plans of a scheme the feed reports once, under the Regular
        # plan's NAV, so keeping them duplicates the fund and misstates its AUM.
        # ECO is matched on a word boundary, never as a substring: "eco"
        # appears inside Canara Rob(eco), and a plain containment test drops
        # all 158 of their rows -- Canara Robeco Multi Cap among them.
        #
        # Wealth and Blended must carry the word "plan"; the others stand alone.
        # "Wealth" on its own catches 47 funds of which only 4 are a Wealth
        # Plan -- the rest belong to The Wealth Company, an AMC with funds in
        # every category, and to Long Term Wealth Enhancement Fund. That is the
        # Canara Rob(eco) trap again: a share-class word that is also part of a
        # legitimate name.
        #
        # Institutional, Retail and Super are closed share classes AMFI still
        # lists but no longer prices:
        # "ICICI Prudential Bluechip Fund - Institutional Option - I",
        # "PGIM India Large Cap Fund Wealth Plan", "ICICI Prudential Blended
        # Plan A". All four found across the equity categories returned zero
        # NAVs over an eight-month window, so they only inflated the fund count
        # and then showed up as schemes with no data.
        if (re.search(r"\bplan\s*[b-z]\b", nm)
                or re.search(r"\beco\b", nm)
                or re.search(r"\b(institutional|retail|super)\b", nm)
                or re.search(r"\b(wealth|blended)\s+plan\b", nm)):
            skipped_option += 1
            continue
        # Option is "Growth", "Growth Option", "IDCW" or "Other" depending on
        # which feed built the index. "Other" covers schemes whose option the
        # feed does not state, which the name test above has already vetted, so
        # only an explicitly non-Growth option is rejected here.
        # "Other" means the feed states an option that is neither Growth nor
        # IDCW, so it is not a Growth share class and must go. Only a genuinely
        # absent option ("unknown") falls through to the name test above.
        if option_type and option_type != "unknown" and "growth" not in option_type:
            skipped_option += 1
            continue

        isin_g = row.get("ISIN Div Payout/ ISIN Growth") or row.get("ISIN Div Payout / ISIN Growth")
        if isin_g and pd.notna(isin_g) and str(isin_g).strip() not in ("", "-"):
            isins.append(str(isin_g).strip().upper())

    isins = sorted(set(isins))
    return isins, len(isins), skipped_direct, skipped_option


def run_historical_export(
    parsed_isins: List[str],
    start_date: datetime.date,
    end_date: datetime.date,
    carry_forward: bool,
    skip_sundays: bool,
    want_nav: bool,
    want_aum: bool,
    want_flows: bool,
    fetch_live_aum: bool = False,
    skip_saturdays: bool = True
) -> Tuple[pd.DataFrame, bool, str | None]:
    if want_flows:
        want_nav = True
        want_aum = True
        fetch_live_aum = True
        
    fetch_start_date = start_date
    if want_flows:
        fetch_start_date = start_date - timedelta(days=10)
        
    try:
        # Determine if running inside Streamlit context
        in_streamlit = False
        try:
            from streamlit.runtime import exists as runtime_exists
            in_streamlit = runtime_exists()
        except ImportError:
            pass

        # Split target date range into chunks of up to 90 days
        # (AMFI portal reliably returns CSV for 90-day windows)
        chunks = []
        curr_start = fetch_start_date
        while curr_start <= end_date:
            curr_end = min(curr_start + timedelta(days=90), end_date)
            chunks.append((curr_start, curr_end))
            curr_start = curr_end + timedelta(days=1)

        rows = []
        df_port = load_portfolio_aum_data()
        parsed_isins_set = {x.upper() for x in parsed_isins}

        if in_streamlit:
            progress_bar = st.progress(0.0, text=f"Fetching historical data (0/{len(chunks)} chunks)...")

        # ── Concurrent chunk fetching & parsing ────────────────────────
        from concurrent.futures import ThreadPoolExecutor, as_completed
        import time

        chunk_results = {}  # idx -> parsed rows list

        def _download_and_parse_chunk(idx, c_start, c_end, isins_set):
            try:
                content_text = fetch_amfi_chunk_with_cache(c_start, c_end)
                if not content_text:
                    return idx, []
                
                chunk_rows = []
                current_section = "Unknown"
                # AMFI inserted Plan and Option into the dump, moving the ISINs
                # from columns 2-3 to 4-5. Read positions from the header rather
                # than assuming them, or the ISIN filter matches Plan/Option and
                # silently discards every row.
                cols = {}
                for line in content_text.splitlines():
                    if not line:
                        continue
                    if ";" not in line:
                        line_stripped = line.strip()
                        if (
                            line_stripped.startswith("Open Ended")
                            or line_stripped.startswith("Closed Ended")
                            or line_stripped.startswith("Interval Fund Schemes")
                        ):
                            current_section = line_stripped
                        continue

                    if not cols and line.startswith("Scheme Code"):
                        cols = _map_amfi_columns(line)
                        continue

                    parts = line.split(";")
                    if len(parts) <= cols.get("date", 7):
                        continue

                    def _col(key, default=""):
                        i = cols.get(key)
                        return parts[i].strip() if i is not None and i < len(parts) else default

                    isin_growth = _col("isin_growth")
                    isin_reinvestment = _col("isin_reinvest")
                    isin_growth_upper = isin_growth.upper() if isin_growth != "-" else ""
                    isin_reinvest_upper = isin_reinvestment.upper() if isin_reinvestment != "-" else ""

                    g_match = isin_growth_upper and isin_growth_upper in isins_set
                    r_match = isin_reinvest_upper and isin_reinvest_upper in isins_set

                    if g_match or r_match:
                        scheme_name = _col("scheme_name")
                        # Plan and Option are their own columns now; prefer them
                        # over inferring from the scheme name.
                        plan = _col("plan")
                        option = _col("option")
                        if plan:
                            is_direct = "direct" in plan.lower()
                        else:
                            scheme_name_lower = scheme_name.lower()
                            is_direct = "direct" in scheme_name_lower or bool(re.search(r"\bdir\b", scheme_name_lower))
                        if option:
                            opt = option.lower()
                            is_idcw = ("idcw" in opt) or ("dividend" in opt) or ("payout" in opt)
                        else:
                            is_idcw = is_idcw_scheme(scheme_name)
                        if is_direct or is_idcw:
                            continue

                        scheme_code = _col("scheme_code")
                        nav_value = _col("nav")
                        nav_date = _col("date")
                        
                        scheme_code = scheme_code if scheme_code != "-" else None
                        isin_growth = isin_growth if isin_growth != "-" else None
                        isin_reinvestment = isin_reinvestment if isin_reinvestment != "-" else None
                        
                        nav = pd.to_numeric(nav_value.replace(",", ""), errors="coerce")
                        chunk_rows.append({
                            "Asset Class": current_section,
                            "Scheme Code": scheme_code,
                            "ISIN Div Payout / ISIN Growth": isin_growth,
                            "ISIN Div Reinvestment": isin_reinvestment,
                            "Scheme Name": scheme_name,
                            "NAV": nav,
                            "Date": nav_date
                        })
                return idx, chunk_rows
            except Exception as exc:
                # Was a bare `return idx, []`, which made a throttled or failed
                # chunk indistinguishable from an empty date range and sent
                # everyone to "check the ISINs and the date range".
                print(f"NAV chunk {c_start:%d-%b-%Y} to {c_end:%d-%b-%Y} failed: "
                      f"{type(exc).__name__}: {exc}")
                return idx, []

        # Concurrency is the main speed lever for multi-year ranges (a 10-year
        # window is ~40 chunks). On Streamlit Cloud we cap at 3 workers to stay
        # under the ~1 GB RAM limit (each AMFI chunk is ~5 MB of raw text); on a
        # local machine we push up to 12 parallel downloads for a big speed-up.
        with ThreadPoolExecutor(max_workers=_nav_chunk_workers(len(chunks))) as executor:
            futures = {
                executor.submit(_download_and_parse_chunk, idx, c_start, c_end, parsed_isins_set): idx
                for idx, (c_start, c_end) in enumerate(chunks)
            }
            completed = 0
            for future in as_completed(futures):
                completed += 1
                idx, chunk_rows = future.result()
                chunk_results[idx] = chunk_rows
                if in_streamlit:
                    c_s, c_e = chunks[idx]
                    progress_bar.progress(
                        completed / len(chunks),
                        text=f"Fetching historical data ({completed}/{len(chunks)} chunks): {c_s.strftime('%d-%b-%Y')} to {c_e.strftime('%d-%b-%Y')}..."
                    )

        # Assemble results in chronological order
        for chunk_idx in range(len(chunks)):
            rows.extend(chunk_results.get(chunk_idx, []))

        if in_streamlit:
            progress_bar.progress(1.0, text="Fetching complete!")
            time.sleep(0.5)
            progress_bar.empty()

        if not rows:
            return pd.DataFrame(), False, "No NAV records found matching the specified criteria and date range."
            
        df_raw = pd.DataFrame(rows)
        
        df_raw["Plan Type"] = df_raw["Scheme Name"].apply(classify_plan_type)
        df_raw["Option Type"] = df_raw["Scheme Name"].apply(classify_option_type)
    
        # Infer proper Asset Class
        known_sections = ["Open Ended", "Closed Ended", "Interval Fund Schemes"]
        def infer_section_from_name(name: str) -> str:
            lowered = name.lower()
            if "flexi" in lowered:
                return "Open Ended Schemes (Equity Scheme - Flexi Cap Fund)"
            if "mid" in lowered:
                return "Open Ended Schemes (Equity Scheme - Mid Cap Fund)"
            if "small" in lowered:
                return "Open Ended Schemes (Equity Scheme - Small Cap Fund)"
            if "large" in lowered:
                return "Open Ended Schemes (Equity Scheme - Large Cap Fund)"
            if "focused" in lowered or "focus" in lowered:
                return "Open Ended Schemes (Equity Scheme - Focused Fund)"
            return "Open Ended Schemes (Equity Scheme - Large Cap Fund)"
        df_raw["Asset Class"] = df_raw.apply(
            lambda row: row["Asset Class"] if any(row["Asset Class"].startswith(s) for s in known_sections) else infer_section_from_name(row["Scheme Name"]),
            axis=1)
        df_raw = populate_actual_aum(df_raw, df_port, fetch_live_aum=fetch_live_aum)
        if not df_raw.empty:
            parsed_dates = parse_amfi_date_series(df_raw["Date"])
            df_raw["Date_parsed"] = parsed_dates
            df_raw["Date"] = parsed_dates.dt.strftime("%d-%m-%Y")
            df_raw = df_raw.sort_values(["Scheme Code", "Date_parsed"]).reset_index(drop=True)
            df_raw["Prev NAV"] = df_raw.groupby("Scheme Code")["NAV"].shift(1)
            df_raw["Daily Return %"] = (
                (df_raw["NAV"] - df_raw["Prev NAV"]) / df_raw["Prev NAV"] * 100
            ).where(df_raw["Prev NAV"].notna() & (df_raw["Prev NAV"] != 0))
            df_raw["Prev AUM"] = df_raw.groupby("Scheme Code")["AUM"].shift(1)
            df_raw["Derived AUM"] = df_raw["Prev AUM"] * (1 + df_raw["Daily Return %"] / 100)
            df_raw["Flows"] = (df_raw["AUM"] - df_raw["Derived AUM"]).where(
                df_raw["AUM"].notna() & df_raw["Derived AUM"].notna()
            )
            df_raw.drop(columns=["Prev NAV", "Prev AUM", "Derived AUM", "Date_parsed"], inplace=True)
            
        all_dates = pd.date_range(start=fetch_start_date, end=end_date)
    
        target_dates = []
        for dt in all_dates:
            if skip_saturdays and dt.weekday() == 5:
                continue
            if skip_sundays and dt.weekday() == 6:
                continue
            target_dates.append(dt.strftime("%d-%m-%Y"))
        
        fund_metadata = df_raw[[
            "Asset Class", 
            "Scheme Code", 
            "ISIN Div Payout / ISIN Growth", 
            "ISIN Div Reinvestment", 
            "Scheme Name", 
            "Plan Type", 
            "Option Type"
        ]].drop_duplicates(subset=["Scheme Code"])
    
        # Determine sorted date columns chronologically
        sorted_date_cols = target_dates
        if carry_forward and len(target_dates) > 1:
            try:
                date_objs = sorted([datetime.strptime(d, "%d-%m-%Y") for d in target_dates])
                sorted_date_cols = [d.strftime("%d-%m-%Y") for d in date_objs]
            except Exception:
                pass

        if want_nav and not want_aum:
            df_pivot = df_raw.pivot(index="Scheme Code", columns="Date", values="NAV").reset_index()
            df_pivot = df_pivot.reindex(columns=["Scheme Code"] + sorted_date_cols)
            if carry_forward and len(sorted_date_cols) > 1:
                df_pivot[sorted_date_cols] = df_pivot[sorted_date_cols].T.ffill().T
            display_date_cols = sorted_date_cols
            is_aum_only = False
        elif want_aum and not want_nav:
            df_pivot = df_raw.pivot(index="Scheme Code", columns="Date", values="AUM").reset_index()
            df_pivot = df_pivot.reindex(columns=["Scheme Code"] + sorted_date_cols)
            if carry_forward and len(sorted_date_cols) > 1:
                df_pivot[sorted_date_cols] = df_pivot[sorted_date_cols].T.ffill().T
            display_date_cols = sorted_date_cols
            is_aum_only = True
        else:
            df_pivot_nav = df_raw.pivot(index="Scheme Code", columns="Date", values="NAV").reset_index()
            df_pivot_aum = df_raw.pivot(index="Scheme Code", columns="Date", values="AUM").reset_index()
            df_pivot_flow = df_raw.pivot(index="Scheme Code", columns="Date", values="Flows").reset_index()
            
            df_pivot_nav = df_pivot_nav.reindex(columns=["Scheme Code"] + sorted_date_cols)
            df_pivot_aum = df_pivot_aum.reindex(columns=["Scheme Code"] + sorted_date_cols)
            df_pivot_flow = df_pivot_flow.reindex(columns=["Scheme Code"] + sorted_date_cols)
            
            if carry_forward and len(sorted_date_cols) > 1:
                df_pivot_nav[sorted_date_cols] = df_pivot_nav[sorted_date_cols].T.ffill().T
                df_pivot_aum[sorted_date_cols] = df_pivot_aum[sorted_date_cols].T.ffill().T
                
            nav_cols_map = {d: f"{d} (NAV)" for d in sorted_date_cols}
            aum_cols_map = {d: f"{d} (AUM)" for d in sorted_date_cols}
            flow_cols_map = {d: f"{d} (Flows)" for d in sorted_date_cols}
            df_pivot_nav = df_pivot_nav.rename(columns=nav_cols_map)
            df_pivot_aum = df_pivot_aum.rename(columns=aum_cols_map)
            df_pivot_flow = df_pivot_flow.rename(columns=flow_cols_map)
            
            df_pivot = pd.merge(df_pivot_nav, df_pivot_aum, on="Scheme Code", how="left")
            df_pivot = pd.merge(df_pivot, df_pivot_flow, on="Scheme Code", how="left")
            
            interleaved_dates = []
            for d in sorted_date_cols:
                interleaved_dates.append(f"{d} (NAV)")
                interleaved_dates.append(f"{d} (AUM)")
                interleaved_dates.append(f"{d} (Flows)")
            display_date_cols = interleaved_dates
            is_aum_only = False
            
        df_final = pd.merge(fund_metadata, df_pivot, on="Scheme Code", how="left")
        del df_pivot  # free pivot memory

        if want_aum:
            df_pivot_fallback = df_raw.pivot(index="Scheme Code", columns="Date", values="Fallback_AUM").reset_index()
            fallback_cols_map = {}
            for d in target_dates:
                if want_aum and not want_nav:
                    fallback_cols_map[d] = f"{d}_fallback_temp"
                elif want_nav and want_aum:
                    fallback_cols_map[f"{d} (AUM)"] = f"{d}_fallback_temp"
                
            if fallback_cols_map:
                df_pivot_fallback_renamed = df_pivot_fallback.rename(columns={d: f"{d}_fallback_temp" for d in target_dates if d in df_pivot_fallback.columns})
                available_temp_cols = [col for col in fallback_cols_map.values() if col in df_pivot_fallback_renamed.columns]
                df_final_temp = pd.merge(df_final, df_pivot_fallback_renamed[["Scheme Code"] + available_temp_cols], on="Scheme Code", how="left")
                for main_col, temp_col in fallback_cols_map.items():
                    if main_col in df_final.columns and temp_col in df_final_temp.columns:
                        df_final[main_col] = df_final[main_col].fillna(df_final_temp[temp_col])
            
        vertical_rows = []
        for _, row in df_final.iterrows():
            meta = {
                "Asset Class": row["Asset Class"],
                "Scheme Code": row["Scheme Code"],
                "ISIN Div Payout / ISIN Growth": row["ISIN Div Payout / ISIN Growth"],
                "ISIN Div Reinvestment": row["ISIN Div Reinvestment"],
                "Scheme Name": row["Scheme Name"],
                "Plan Type": row["Plan Type"],
                "Option Type": row["Option Type"]
            }
            for d in target_dates:
                r_item = meta.copy()
                if want_nav and not want_aum:
                    r_item["NAV Date"] = d
                    r_item["NAV"] = row[d]
                elif want_aum and not want_nav:
                    r_item["AUM Date"] = d
                    r_item["AUM"] = row[d]
                else:
                    r_item["NAV Date"] = d
                    r_item["NAV"] = row[f"{d} (NAV)"]
                    r_item["AUM Date"] = d
                    r_item["AUM"] = row[f"{d} (AUM)"]
                vertical_rows.append(r_item)

        ordered_cols = [
            "Asset Class", 
            "Scheme Code", 
            "ISIN Div Payout / ISIN Growth", 
            "ISIN Div Reinvestment", 
            "Scheme Name", 
            "Plan Type", 
            "Option Type"
        ]
        if want_nav:
            ordered_cols.extend(["NAV Date", "NAV"])
        if want_aum:
            ordered_cols.extend(["AUM Date", "AUM"])
    
        del df_raw  # free the large intermediate frame before building vertical rows
        
        if vertical_rows:
            df_res_final = pd.DataFrame(vertical_rows)
            del vertical_rows  # free list now it's in a DataFrame
            # NAV Date / AUM Date are already formatted as dd-mm-YYYY strings — skip redundant parse
            df_res_final = df_res_final[ordered_cols]

            # Drop rows carrying no observation at all. With weekend skipping
            # switched off, a row is scaffolded for every calendar date, and the
            # non-trading ones have neither a NAV nor an AUM — they render as a
            # run of "None" cells that reads like a broken report rather than a
            # closed market. Nothing is lost: these rows never held a value.
            _obs = [c for c in ("NAV", "AUM") if c in df_res_final.columns]
            if _obs:
                _empty = df_res_final[_obs].isna().all(axis=1)
                if _empty.any():
                    print(f"Dropped {int(_empty.sum())} row(s) with no NAV or AUM "
                          f"(non-trading days).")
                    df_res_final = df_res_final[~_empty].reset_index(drop=True)

            df_res_final = df_res_final.sort_values(by=["Asset Class", "Scheme Name"]).reset_index(drop=True)
            if want_flows:
                df_res_final = calculate_flows_for_dataframe(df_res_final, start_date, ["Asset Class", "Scheme Code", "ISIN Div Payout / ISIN Growth", "ISIN Div Reinvestment", "Scheme Name", "Plan Type", "Option Type"])
                is_aum_only = False
        else:
            df_res_final = pd.DataFrame(columns=ordered_cols)

        # Which requested schemes produced nothing? A fund can be selected and
        # still yield no row -- dormant share classes, or a scheme launched
        # after the start date. Silently returning fewer schemes than the count
        # promised is what made these look like a bug rather than a fact about
        # the data, so the caller is told which ones and why.
        try:
            _asked = {str(i).strip().upper() for i in (parsed_isins or [])}
            _got = set()
            for _c in ("ISIN Div Payout / ISIN Growth", "ISIN Div Reinvestment"):
                if _c in df_res_final.columns:
                    _got |= {str(v).strip().upper() for v in df_res_final[_c].dropna()}
            _silent = sorted(_asked - _got)
            if _silent:
                globals()["_LAST_EXPORT_SILENT"] = _silent
                print(f"{len(_silent)} requested scheme(s) returned no NAV in this range "
                      f"(dormant plans, or launched after the start date): "
                      f"{', '.join(_silent[:6])}")
            else:
                globals()["_LAST_EXPORT_SILENT"] = []
        except Exception:
            globals()["_LAST_EXPORT_SILENT"] = []

        return df_res_final, is_aum_only, None
    except Exception as e:
        return pd.DataFrame(), False, str(e)


def generate_historical_excel(df_final: pd.DataFrame, target_dates: List[str], is_aum_only: bool = False) -> bytes:
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df_final.to_excel(writer, index=False, sheet_name="NAV Data")
        
        workbook = writer.book
        worksheet = writer.sheets["NAV Data"]
        
        font_family = "Segoe UI"
        header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")  # Navy Blue
        header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
        
        data_font = Font(name=font_family, size=10)
        
        # Zebra striping
        even_row_fill = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
        odd_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        thin_border = Border(
            left=Side(style="thin", color="D3D3D3"),
            right=Side(style="thin", color="D3D3D3"),
            top=Side(style="thin", color="D3D3D3"),
            bottom=Side(style="thin", color="D3D3D3")
        )
        
        alignments = {
            "Asset Class": Alignment(horizontal="left", vertical="center"),
            "Scheme Code": Alignment(horizontal="center", vertical="center"),
            "ISIN Div Payout / ISIN Growth": Alignment(horizontal="center", vertical="center"),
            "ISIN Div Reinvestment": Alignment(horizontal="center", vertical="center"),
            "Scheme Name": Alignment(horizontal="left", vertical="center"),
            "Plan Type": Alignment(horizontal="center", vertical="center"),
            "Option Type": Alignment(horizontal="center", vertical="center"),
            "NAV Date": Alignment(horizontal="center", vertical="center"),
            "AUM Date": Alignment(horizontal="center", vertical="center"),
            "NAV": Alignment(horizontal="right", vertical="center"),
            "NAVs": Alignment(horizontal="right", vertical="center"),
            "AUM": Alignment(horizontal="right", vertical="center"),
            "Closing AUM as on previous day": Alignment(horizontal="right", vertical="center"),
            "Actual AUM as on current date": Alignment(horizontal="right", vertical="center"),
            "Daily return": Alignment(horizontal="right", vertical="center"),
            "Derived AUM as on curent day": Alignment(horizontal="right", vertical="center"),
            "Net flows on current day": Alignment(horizontal="right", vertical="center"),
        }
            
        worksheet.row_dimensions[1].height = 28
        for col_idx, col_name in enumerate(df_final.columns, 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center" if col_name not in ["Scheme Name", "Asset Class"] else "left", vertical="center")
            cell.border = thin_border
            
        for row_idx in range(2, len(df_final) + 2):
            worksheet.row_dimensions[row_idx].height = 20
            fill = even_row_fill if row_idx % 2 == 0 else odd_row_fill
            
            for col_idx, col_name in enumerate(df_final.columns, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.fill = fill
                cell.font = data_font
                cell.border = thin_border
                
                val = cell.value
                if "NAV" in col_name:
                    if val is not None and val != "":
                        cell.number_format = "0.00" if "NAVs" in col_name else "0.0000"
                    else:
                        cell.value = "-"
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                elif "Return" in col_name or "return" in col_name:
                    if val is not None and val != "":
                        cell.number_format = '0.00"%"'
                    else:
                        cell.value = "-"
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                elif "AUM" in col_name or "flows" in col_name or "AUM" in col_name:
                    if val is not None and val != "":
                        cell.number_format = "#,##0.00"
                    else:
                        cell.value = "-"
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                elif "Info Ratio" in col_name:
                    if val is not None and val != "":
                        cell.number_format = "0.0000"
                    else:
                        cell.value = "-"
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                elif "Date" in col_name:
                    if val is None or val == "":
                        cell.value = "-"
                
                align = alignments.get(col_name, Alignment(horizontal="left", vertical="center"))
                cell.alignment = align

        for col in worksheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if cell.row == 1:
                    val_str = val_str + "   "
                if len(val_str) > max_len:
                    max_len = len(val_str)
            worksheet.column_dimensions[col_letter].width = max(min(max_len + 2, 55), 12)
            
    return buffer.getvalue()

from amfi_nav import (
    AMFINavError,
    benchmark_delta,
    comparison_table,
    export_to_excel,
    fetch_nav_data,
    filter_family_rows,
    load_historical_snapshots,
    search_fund,
    sip_future_value,
    summarize_families,
    classify_plan_type,
    classify_option_type,
)
from nav_fetcher import (
    parse_bucket_input, run_live_portfolio, style_portfolio_excel, is_idcw_scheme,
    _map_amfi_columns,
    # Consolidated: app.py used to carry byte-identical (or near-identical)
    # copies of these, so the AMFI column change and the stale-AUM rule each
    # had to be fixed twice and the second copy was missed both times.
    map_section_to_ids, find_matching_perf_row, load_portfolio_aum_data,
    calculate_flows_for_dataframe, STALE_AUM_MIN_MOVE_PCT,
    fetch_performance_data_from_api, get_api_session, API_CACHE_DIR,
    populate_actual_aum, _LAST_AUM_STATS,
    _is_cloud_env, _perf_api_workers, _nav_chunk_workers,
)
from ui_theme import (
    finance_panel,
    inject_custom_css,
    render_app_footer,
    render_hero,
    render_info_card,
    render_section_header,
    render_ticker_strip,
    render_top_bar,
)


st.set_page_config(
    page_title="NAV Terminal · AMFI India",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed",
)


def frame_to_excel_bytes(frame: pd.DataFrame) -> bytes:
    buffer = BytesIO()
    export_to_excel(frame, buffer)
    return buffer.getvalue()


def frame_to_csv_bytes(frame: pd.DataFrame) -> bytes:
    buffer = StringIO()
    export_frame = frame.copy()
    if "NAV Date" in export_frame.columns:
        export_frame["NAV Date"] = pd.to_datetime(export_frame["NAV Date"], errors="coerce").dt.strftime("%d-%m-%Y")
    export_frame.to_csv(buffer, index=False)
    return buffer.getvalue().encode("utf-8")


def build_export_frame(frame: pd.DataFrame) -> pd.DataFrame:
    export_frame = frame.copy()
    columns = [
        "Scheme Name",
        "Scheme Code",
        "Plan Type",
        "Option Type",
        "NAV",
        "NAV Date",
        "AMC Name",
    ]
    available = [column for column in columns if column in export_frame.columns]
    export_frame = export_frame[available]
    export_frame = export_frame.rename(columns={"Option Type": "Plan Option"})
    return export_frame


def normalize_filename(value: str) -> str:
    value = re.sub(r"[^A-Za-z0-9_-]+", "_", value)
    return value.strip("_")[:80]


def render_aum_cache_page() -> None:
    """Build and share the AMFI AUM cache.

    AMFI rate-limits by IP and refuses a caller that asks for a lot at once, so
    fetching on demand fails exactly when a report is wanted. This collects the
    data slowly instead and keeps it: past-date AUM never changes, so every
    answered query is good permanently.
    """
    import aum_backfill
    import aum_cache_sync
    from datetime import date as _date, timedelta as _td

    st.markdown("### AUM Cache")
    st.caption(
        "AMFI answers a slow caller and refuses a fast one. Build the cache here over "
        "as long as it takes, then reuse it: reports read the cache and never wait on AMFI."
    )

    summary = aum_backfill.cache_summary()
    c1, c2, c3 = st.columns(3)
    c1.metric("Cached queries", f"{summary['files']:,}")
    c2.metric("Dates covered", summary["dates"])
    c3.metric("Size", f"{summary['mb']} MB")
    if summary["earliest"]:
        st.caption(f"Covers {summary['earliest']} to {summary['latest']}.")

    st.divider()
    st.markdown("**Collect**")
    st.caption(
        "Run this where AMFI answers you — your own machine. A hosted deployment fetches "
        "from a datacenter address, which is what AMFI blocks hardest, and its disk is "
        "wiped on restart so nothing accumulates."
    )

    d1, d2 = st.columns(2)
    today = _date.today()
    with d1:
        bf_start = st.date_input("From", value=today - _td(days=90), max_value=today, key="bf_start")
    with d2:
        bf_end = st.date_input("To", value=today, max_value=today, key="bf_end")

    planned = aum_backfill.plan_work(bf_start, bf_end)
    st.caption(f"{len(planned):,} queries still to fetch for this range.")

    budget_min = st.slider("Stop after (minutes)", 1, 240, 30, key="bf_budget",
                           help="The job is resumable — whatever it fetches is kept, and "
                                "re-running continues from there.")

    if st.button("Start collecting", type="primary", disabled=not planned, key="bf_go"):
        bar = st.progress(0.0, text="Starting…")
        log = st.empty()

        def on_progress(stats: dict) -> None:
            done = stats.get("fetched", 0)
            frac = done / max(stats.get("planned", 1), 1)
            bar.progress(min(max(frac, 0.0), 1.0),
                         text=f"{done:,} of {stats.get('planned', 0):,} fetched")
            if stats.get("waiting"):
                log.warning(
                    f"AMFI refused the request ({stats['reason']}). Waiting "
                    f"{stats['waiting']}s, then retrying {stats['current']}. "
                    f"Nothing already fetched is lost."
                )
            else:
                log.caption(f"Fetching {stats.get('current', '')}")

        with st.spinner("Collecting — leave this tab open."):
            stats = aum_backfill.backfill(
                bf_start, bf_end, progress=on_progress, total_budget=budget_min * 60
            )
        bar.empty()
        log.empty()

        msg = (f"Fetched {stats['fetched']:,} queries "
               f"({stats['empty']:,} had no data), waited out {stats['blocked_waits']:,} "
               f"block(s) in {stats['elapsed']:.0f}s.")
        if stats["stopped_early"]:
            st.warning(msg + " Stopped at the time limit — press Start again to continue.")
        else:
            st.success(msg + " Range complete.")
        st.rerun()

    st.divider()
    st.markdown("**Share with the deployed app**")
    st.caption(
        "Publishes the cache to a Hugging Face Dataset the Space reads at start-up, so the "
        "deployment serves real AUM without ever calling AMFI."
    )

    hub_ok, hub_why = aum_cache_sync.available()
    if not hub_ok:
        st.info(f"Sync unavailable: {hub_why}")
    else:
        st.code(f"repo: {aum_cache_sync.repo_id()}", language=None)
        token_present = bool(os.environ.get("HF_TOKEN") or os.environ.get("HUGGING_FACE_HUB_TOKEN"))
        if not token_present:
            st.info(
                "Set HF_TOKEN in your environment before pushing. Keep it in the environment, "
                "not in a file — a token committed once outlives every place you meant it to be used."
            )
        make_public = st.checkbox(
            "Make the dataset public", value=False, key="sync_public",
            help="The contents are AMFI's published AUM figures. Public means the Space needs "
                 "no token to read it; private means adding HF_TOKEN as a Space secret.",
        )
        s1, s2 = st.columns(2)
        with s1:
            if st.button("Push cache", disabled=not token_present, key="sync_push"):
                try:
                    st.success(aum_cache_sync.push(private=not make_public))
                except Exception as exc:
                    st.error(f"Push failed: {exc}")
        with s2:
            if st.button("Pull cache", key="sync_pull"):
                try:
                    st.success(aum_cache_sync.pull())
                    st.rerun()
                except Exception as exc:
                    st.error(f"Pull failed: {exc}")

    st.divider()
    st.caption(
        "To run it unattended instead, schedule `python aum_backfill.py --daily` once a day; "
        "it tops up the newest dates and exits."
    )


def render_aum_health(stats: dict) -> None:
    """Report where each row's AUM actually came from.

    The four sources are genuinely different in quality and must not be
    conflated: a live feed value, a value derived from monthly AUM scaled by
    NAV, one carried from an adjacent day, and none at all. Reporting derived
    figures as live is what let a run with zero answered queries claim 86% live
    coverage.
    """
    if not stats or not stats.get("rows_total"):
        return

    total = stats["rows_total"]
    live = stats.get("rows_live_aum", 0)
    derived = stats.get("rows_derived_aum", 0)
    carried = stats.get("rows_carried", 0)
    blank = stats.get("rows_blank", 0)
    q_total = stats.get("perf_queries", 0)
    q_ok = stats.get("perf_ok", 0)

    parts = []
    if live:
        parts.append(f"{live:,} live from AMFI ({live / total * 100:.0f}%)")
    if derived:
        parts.append(f"{derived:,} derived from monthly AUM")
    if carried:
        parts.append(f"{carried:,} carried from an adjacent day")
    if blank:
        parts.append(f"{blank:,} with no AUM")
    breakdown = "; ".join(parts) if parts else "no rows"

    msg = f"AUM sources across {total:,} rows — {breakdown}. AMFI queries answered: {q_ok}/{q_total}."

    if q_total and q_ok == 0:
        st.warning(
            msg
            + " AMFI answered none of them, so no figure here is live. Derived AUM moves only "
            "with NAV, so the flows it implies are near zero rather than real. AMFI rate-limits "
            "by IP and usually clears within a few hours — re-run then for live figures."
        )
    elif blank or (q_total and q_ok / q_total < 0.5):
        st.warning(
            msg
            + " AMFI under-delivered; rows without AUM report no flow rather than a guess. "
            "Re-running later will fill them, as answered queries are cached to disk."
        )
    else:
        st.caption(msg)


@st.cache_data(ttl=3600, show_spinner=False)
def load_data(force_refresh: bool = False) -> pd.DataFrame:
    return fetch_nav_data(force_refresh=force_refresh)


def format_timestamp(frame: pd.DataFrame) -> str:
    if frame.empty or "NAV Date" not in frame.columns:
        return "Unavailable"
    latest = pd.to_datetime(frame["NAV Date"], errors="coerce").dropna().max()
    if pd.isna(latest):
        return "Unavailable"
    return pd.Timestamp(latest).strftime("%d-%m-%Y")


def render_result_card(selected_summary: pd.Series, selected_rows: pd.DataFrame) -> None:
    render_section_header("📈", "Fund Overview", "Key metrics for the selected scheme family")
    latest_nav = selected_summary.get("Latest NAV")
    nav_display = f"{latest_nav:.4f}" if pd.notna(latest_nav) else "N/A"
    latest_date = pd.to_datetime(selected_summary.get("Latest NAV Date"), errors="coerce")
    date_display = latest_date.strftime("%d-%m-%Y") if pd.notna(latest_date) else "N/A"
    st.markdown(
        f"""
        <div class="metrics-grid">
            <div class="metric-card">
                <span class="metric-label">Fund</span>
                <span class="metric-value">{selected_summary.get("Family Name", "-")}</span>
            </div>
            <div class="metric-card">
                <span class="metric-label">AMC</span>
                <span class="metric-value">{selected_summary.get("AMC Name", "-")}</span>
            </div>
            <div class="metric-card">
                <span class="metric-label">Latest NAV</span>
                <span class="metric-value accent">{nav_display}</span>
            </div>
            <div class="metric-card">
                <span class="metric-label">NAV Date</span>
                <span class="metric-value">{date_display}</span>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    render_section_header("⚖️", "Regular vs Direct Comparison")
    comparison = comparison_table(selected_rows, selected_summary["Family Key"])
    if comparison.empty:
        st.info("No comparison rows available for the selected fund.")
    else:
        st.dataframe(comparison, use_container_width=True, hide_index=True)


def render_history(selected_summary: pd.Series) -> None:
    render_section_header("📉", "Historical NAV Cache", "Trends from locally cached snapshots")
    history = load_historical_snapshots(selected_summary["Family Key"])
    if history.empty:
        st.info("No cached historical snapshots are available yet. The app will build them after a successful refresh.")
        return

    timeline = history[["NAV Date", "Scheme Name", "Plan Type", "Option Type", "NAV"]].copy()
    timeline["NAV Date"] = pd.to_datetime(timeline["NAV Date"], errors="coerce")
    timeline = timeline.dropna(subset=["NAV Date"])
    st.line_chart(timeline.pivot_table(index="NAV Date", values="NAV", aggfunc="mean"))
    st.dataframe(timeline.sort_values("NAV Date", ascending=False), use_container_width=True, hide_index=True)

    render_section_header("🎯", "Benchmark Comparison")
    benchmark_return = st.number_input("Assumed benchmark annual return %", min_value=0.0, value=10.0, step=0.5)
    daily_series = timeline.groupby("NAV Date", as_index=False)["NAV"].mean().sort_values("NAV Date")
    delta = benchmark_delta(daily_series["NAV"], benchmark_return)
    metric_cols = st.columns(3)
    metric_cols[0].metric("Observed change %", "N/A" if delta["observed_change_pct"] is None else f"{delta['observed_change_pct']:.2f}")
    metric_cols[1].metric("Benchmark change %", "N/A" if delta["benchmark_change_pct"] is None else f"{delta['benchmark_change_pct']:.2f}")
    metric_cols[2].metric("Delta %", "N/A" if delta["delta_pct"] is None else f"{delta['delta_pct']:.2f}")


def render_sip_calculator(default_nav: float | None) -> None:
    render_section_header("🧮", "SIP Calculator", "Project future corpus from monthly investments")
    col1, col2, col3 = st.columns(3)
    with col1:
        monthly = st.number_input("Monthly SIP amount", min_value=0.0, value=5000.0, step=500.0)
    with col2:
        annual_return = st.number_input("Expected annual return %", min_value=0.0, value=12.0, step=0.5)
    with col3:
        years = st.number_input("Investment horizon (years)", min_value=1, value=5, step=1)

    future_value = sip_future_value(monthly, annual_return, int(years))
    invested = monthly * 12 * years
    st.metric("Projected corpus", f"₹{future_value:,.2f}")
    st.caption(f"Total invested: ₹{invested:,.2f}")
    if default_nav and default_nav > 0:
        st.caption(f"Reference NAV used for display only: {default_nav:.4f}")


@st.cache_resource(show_spinner=False)
def _sync_aum_cache_once() -> str:
    """Pull the shared AUM cache at start-up, once per process.

    Only on a deployment: locally the cache is the source of truth and pulling
    over it would replace fresh collection with whatever was last published.
    Failure is reported, never raised -- without the dataset the app is exactly
    where it would have been anyway.
    """
    if not os.environ.get("SPACE_ID") and not os.environ.get("AUM_CACHE_PULL"):
        return "local — using the on-disk cache"
    try:
        import aum_cache_sync
        return aum_cache_sync.pull_quietly()
    except Exception as exc:  # noqa: BLE001
        return f"AUM cache not synced: {type(exc).__name__}: {exc}"


def main() -> None:
    inject_custom_css()
    render_top_bar()
    _sync_aum_cache_once()

    with st.sidebar:
        st.markdown("### 📌 Navigation")
        app_module = st.radio(
            "Select Module",
            ["NAV Terminal", "MIS Generator"],
            index=0,
            key="main_app_module_selector",
        )

    if app_module == "MIS Generator":
        from mis_generator import render_mis_generator_page
        render_mis_generator_page()
        render_app_footer()
        return

    try:
        nav_data = load_data(force_refresh=False)
    except Exception as exc:
        st.error(f"Fatal error loading data: {exc}")
        st.stop()

    latest_update = format_timestamp(nav_data)
    scheme_count = len(nav_data) if not nav_data.empty else 0
    render_ticker_strip()
    render_hero(
        title="Mutual Fund NAV Terminal",
        subtitle=(
            "Institutional-grade search, comparison, and export for Indian mutual funds — "
            "live AMFI feeds, fuzzy matching, Regular vs Direct analysis, and styled Excel reports."
        ),
        chips=[
            {"label": "Latest NAV Date", "value": latest_update, "tone": "positive"},
            {"label": "Schemes Indexed", "value": f"{scheme_count:,}", "tone": "gold"},
            {"label": "Data Source", "value": "AMFI India", "tone": ""},
            {"label": "Market", "value": "Mutual Funds", "tone": ""},
        ],
    )

    with finance_panel("Search & Export"):
        render_section_header("🔍", "Fund Discovery", "Search single funds, run batch lookups, or generate historical ISIN reports")
        search_mode = st.radio(
            "Search mode",
            ["Historical ISIN Export", "Fund Performance", "Portfolio Bucket Tracker",
             "MIS Generator", "AUM Cache"],
            horizontal=True,
            index=0,
            label_visibility="collapsed",
        )

        if search_mode == "AUM Cache":
            render_aum_cache_page()

        elif search_mode == "MIS Generator":
            from mis_generator import render_mis_generator_page
            render_mis_generator_page()
        elif search_mode == "Historical ISIN Export":
            render_section_header("📋", "Historical ISIN Export", "Pivoted NAV/AUM reports with corporate Excel styling")
            render_info_card(
                "<strong>Historical NAV Extractor:</strong> Specify a date range and target ISINs to generate "
                "a pivoted, corporate-styled Excel sheet. Weekends and holidays can be filled via carry-forward."
            )

            col1, col2 = st.columns(2)
            with col1:
                start_date = st.date_input(
                    "Start Date", 
                    value=datetime(2026, 1, 1).date(),
                    min_value=datetime(2000, 1, 1).date(),
                    max_value=datetime.today().date()
                )
            with col2:
                end_date = st.date_input(
                    "End Date", 
                    value=datetime.today().date(),
                    min_value=datetime(2000, 1, 1).date(),
                    max_value=datetime.today().date()
                )

            default_isins_str = "\n".join([
                "INF209K01AJ8", "INF846K01CH7", "INF846K016E3", "INF194K01524", "INF760K01019", 
                "INF760K01KR2", "INF740K01128", "INF179K01608", "INF179KA1RT1", "INF179K01CR2", 
                "INF179KA1RZ8", "INF109KA1TX4", "INF109K01BZ4", "INF205KA1189", "INF205K011T7", 
                "INF174KA1EK3", "INF174K01DS9", "INF174KA1HS9", "INF247L01478", "INF204K01GE7", 
                "INF204K01562", "INF204K01489", "INF879O01019", "INF966L01457", "INF966L01AW4", 
                "INF966L01234", "INF200K01370", "INF200K01CT2", "INF200K01297"
            ])
        
            isin_input = st.text_area("Target ISINs (one per line or comma-separated)", value=default_isins_str, height=200)
        
            c1, c2, c3 = st.columns(3)
            with c1:
                carry_forward = st.checkbox("Carry forward NAV on holidays/weekends", value=True)
            with c2:
                skip_saturdays = st.checkbox("Skip Saturdays", value=True, key="hist_skip_sat")
            with c3:
                skip_sundays = st.checkbox("Skip Sundays", value=True, key="hist_skip_sun")

            col_c1, col_c2, col_c3 = st.columns(3)
            with col_c1:
                want_nav = st.checkbox("Want NAV", value=True)
            with col_c2:
                want_aum = st.checkbox("Want AUM", value=True)
            with col_c3:
                want_flows = st.checkbox("Want Flows", value=False)

            fetch_live_aum = st.checkbox("Fetch actual daily AUM from AMFI (slower)", value=True, key="hist_live_aum")

            if st.button("Fetch & generate Excel", type="primary", use_container_width=True):
                st.session_state["hist_result"] = None
                parsed_isins = [x.strip() for x in re.split(r"[,\n\s]+", isin_input) if x.strip()]
                filtered_isins = []
                for isin in parsed_isins:
                    # Look up in nav_data cache
                    match_row = pd.DataFrame()
                    if not nav_data.empty:
                        col_g = "ISIN Div Payout/ ISIN Growth" if "ISIN Div Payout/ ISIN Growth" in nav_data.columns else "ISIN Div Payout / ISIN Growth"
                        match_row = nav_data[
                            (nav_data[col_g].astype(str).str.strip().str.upper() == isin.upper()) |
                            (nav_data["ISIN Div Reinvestment"].astype(str).str.strip().str.upper() == isin.upper())
                        ]
                    if not match_row.empty:
                        row = match_row.iloc[0]
                        scheme_name = str(row.get("Scheme Name", "")).strip().lower()
                        plan_type = str(row.get("Plan Type", "")).strip().lower()
                        option_type = str(row.get("Option Type", "")).strip().lower()
                        
                        is_direct = "direct" in plan_type or "direct" in scheme_name or "dir" in plan_type or re.search(r"\bdir\b", scheme_name)
                        is_idcw = is_idcw_scheme(option_type) or is_idcw_scheme(scheme_name)
                        
                        if is_direct or is_idcw:
                            continue
                    filtered_isins.append(isin)
                parsed_isins = filtered_isins
                
                if not want_nav and not want_aum and not want_flows:
                    st.error("Please select at least one data type (NAV, AUM, or Flows) to export.")
                elif not parsed_isins:
                    st.error("Please enter at least one valid Regular Growth ISIN.")
                elif start_date > end_date:
                    st.error("Start Date cannot be after End Date.")
                else:
                    with st.spinner("Connecting to AMFI India and fetching historical data..."):
                        df_final, is_aum_only, err = run_historical_export(
                            parsed_isins=parsed_isins,
                            start_date=start_date,
                            end_date=end_date,
                            carry_forward=carry_forward,
                            skip_sundays=skip_sundays,
                            want_nav=want_nav,
                            want_aum=want_aum,
                            want_flows=want_flows,
                            fetch_live_aum=fetch_live_aum,
                            skip_saturdays=skip_saturdays
                        )
                    if err:
                        st.warning(err)
                    elif df_final.empty:
                        st.warning("No records found matching the specified criteria.")
                    else:
                        # Persist so the preview + download survive the rerun triggered by a
                        # download-button click (otherwise Streamlit discards the results).
                        st.session_state["hist_result"] = {
                            "df": df_final,
                            "is_aum_only": is_aum_only,
                            "start": str(start_date),
                            "end": str(end_date),
                            "stats": dict(_LAST_AUM_STATS),
                        }

            # Render the persisted result outside the button gate so the download works reliably.
            hist_res = st.session_state.get("hist_result")
            if hist_res:
                df_final = hist_res["df"]
                st.success(f"Successfully processed {len(df_final)} vertical records!")
                render_aum_health(hist_res.get("stats"))

                render_section_header("👁️", "Data Preview")
                st.dataframe(
                    df_final,
                    use_container_width=True,
                    column_config={
                        "Daily return": st.column_config.NumberColumn(
                            "Daily return",
                            format="%.2f%%"
                        )
                    }
                )

                excel_bytes = generate_historical_excel(df_final, [], is_aum_only=hist_res["is_aum_only"])
                st.download_button(
                    label="📥 Download Styled Excel (.xlsx)",
                    data=excel_bytes,
                    file_name=f"amfi_nav_export_{hist_res['start']}_to_{hist_res['end']}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="hist_download",
                )

        elif search_mode == "Fund Performance":
            maturity_type = st.selectbox("Maturity Type", ["Open Ended", "Close Ended", "Interval"], index=0)
            category = st.selectbox("Category", ["Equity", "Debt", "Hybrid", "Solution Oriented", "Other"], index=0)
            maturity_id_map = {"Open Ended": 1, "Close Ended": 2, "Interval": 2}
            cat_id_map = {"Equity": 1, "Debt": 2, "Hybrid": 3, "Solution Oriented": 4, "Other": 5}
            maturity_id = maturity_id_map[maturity_type]
            cat_id = cat_id_map[category]

            with st.spinner("Fetching subcategories..."):
                subcategories, sub_source = load_subcategories(cat_id)
            if sub_source == "cached":
                st.caption("⚠️ AMFI's category list is unavailable; using the last one retrieved.")
            elif sub_source == "builtin":
                st.caption("⚠️ AMFI's category list is unavailable; using the built-in SEBI categories.")

            sub_names = [name for name, _ in subcategories]
            # "ALL" pulls every subcategory in the chosen category together.
            sub_options = ["ALL"] + sub_names if sub_names else ["ALL"]
            subcategory_name = st.selectbox("Subcategory", sub_options, index=0)
            if subcategory_name == "ALL":
                sub_id = "ALL"
            else:
                sub_id = dict(subcategories).get(subcategory_name, 0)

            col_d1, col_d2 = st.columns(2)
            with col_d1:
                start_date = st.date_input(
                    "Start Date", 
                    value=datetime(2026, 1, 1).date(),
                    min_value=datetime(2000, 1, 1).date(),
                    max_value=datetime.today().date()
                )
            with col_d2:
                end_date = st.date_input(
                    "End Date", 
                    value=datetime.today().date(),
                    min_value=datetime(2000, 1, 1).date(),
                    max_value=datetime.today().date()
                )

            c1, c2, c3 = st.columns(3)
            with c1:
                carry_forward = st.checkbox("Carry forward NAV on holidays/weekends", value=True, key="perf_carry")
            with c2:
                skip_saturdays = st.checkbox("Skip Saturdays", value=True, key="perf_skip_sat")
            with c3:
                skip_sundays = st.checkbox("Skip Sundays", value=True, key="perf_skip")

            col_c1, col_c2, col_c3 = st.columns(3)
            with col_c1:
                want_nav = st.checkbox("Want NAV", value=True, key="perf_nav")
            with col_c2:
                want_aum = st.checkbox("Want AUM", value=True, key="perf_aum")
            with col_c3:
                want_flows = st.checkbox("Want Flows", value=True, key="perf_flows")

            fetch_live_aum = st.checkbox("Fetch actual daily AUM from AMFI (slower)", value=True, key="perf_live_aum")

            if st.button("Go", type="primary", use_container_width=True):
                st.session_state["fp_result"] = None
                if not want_nav and not want_aum and not want_flows:
                    st.error("Please select at least one data type (NAV, AUM, or Flows) to export.")
                elif start_date > end_date:
                    st.error("Start Date cannot be after End Date.")
                else:
                    matched_rows = []
                    for idx, row in nav_data.iterrows():
                        ac = row.get("Asset Class", "")
                        m_id, c_id, s_id = map_section_to_ids(ac)
                        if m_id == maturity_id and c_id == cat_id and (sub_id == "ALL" or s_id == sub_id):
                            matched_rows.append(row)

                    df_matched = pd.DataFrame(matched_rows)
                    if df_matched.empty:
                        st.warning("No schemes found in current AMFI index matching this subcategory.")
                    else:
                        parsed_isins, n_kept, n_direct, n_option = select_regular_growth_isins(df_matched)
                        st.caption(
                            f"{len(df_matched)} scheme rows in this category → **{n_kept} Regular Growth "
                            f"funds** ({n_direct} Direct plans and {n_option} IDCW/Bonus options excluded)."
                        )

                        if not parsed_isins:
                            st.warning("No valid ISINs found for schemes in this subcategory.")
                        else:
                            with st.spinner("Connecting to AMFI India and fetching historical data..."):
                                df_final, is_aum_only, err = run_historical_export(
                                    parsed_isins=parsed_isins,
                                    start_date=start_date,
                                    end_date=end_date,
                                    carry_forward=carry_forward,
                                    skip_sundays=skip_sundays,
                                    want_nav=want_nav,
                                    want_aum=want_aum,
                                    want_flows=want_flows,
                                    fetch_live_aum=fetch_live_aum,
                                    skip_saturdays=skip_saturdays
                                )
                            if err:
                                st.warning(err)
                            elif df_final.empty:
                                st.warning("No records found matching the specified criteria.")
                            else:
                                # Persist so the preview + download survive the rerun that a
                                # download-button click triggers (Streamlit re-executes the script,
                                # which would otherwise discard results computed inside this block).
                                st.session_state["fp_result"] = {
                                    "df": df_final,
                                    "is_aum_only": is_aum_only,
                                    "name": subcategory_name,
                                    "start": str(start_date),
                                    "end": str(end_date),
                                    "stats": dict(_LAST_AUM_STATS),
                                }

            # Render the persisted Fund Performance result (outside the button gate) so the
            # Excel download works reliably across reruns.
            fp_res = st.session_state.get("fp_result")
            if fp_res:
                df_final = fp_res["df"]
                st.success(f"Successfully processed {len(df_final)} vertical records!")
                render_aum_health(fp_res.get("stats"))

                render_section_header("👁️", "Data Preview")
                st.dataframe(
                    df_final,
                    use_container_width=True,
                    column_config={
                        "Daily return": st.column_config.NumberColumn(
                            "Daily return",
                            format="%.2f%%"
                        )
                    }
                )

                excel_bytes = generate_historical_excel(df_final, [], is_aum_only=fp_res["is_aum_only"])
                st.download_button(
                    label="📥 Download Styled Excel (.xlsx)",
                    data=excel_bytes,
                    file_name=f"fund_performance_{fp_res['name']}_{fp_res['start']}_to_{fp_res['end']}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="fp_download",
                )
        elif search_mode == "Portfolio Bucket Tracker":
            st.markdown('<hr class="panel-divider">', unsafe_allow_html=True)
            
            # 1. Initialize session state buckets if they don't exist
            if "portfolio_buckets" not in st.session_state:
                default_df = pd.DataFrame([
                    {"Scheme Name": "Quant Large Cap Fund-Reg(G)", "ISIN": "INF966L01AW4", "Weight (%)": 10.0},
                    {"Scheme Name": "DSP Equity Opportunities Fund-Reg(G)", "ISIN": "INF740K01094", "Weight (%)": 9.0},
                    {"Scheme Name": "SBI Large & Midcap Fund-Reg(G)", "ISIN": "INF200K01305", "Weight (%)": 6.0},
                    {"Scheme Name": "HDFC Flexi Cap Fund(G)", "ISIN": "INF179K01608", "Weight (%)": 9.0},
                    {"Scheme Name": "ICICI Pru Focused Equity Fund(G)", "ISIN": "INF109K01BZ4", "Weight (%)": 7.0},
                    {"Scheme Name": "ICICI Pru Dividend Yield Equity Fund(G)", "ISIN": "INF109KA1TX4", "Weight (%)": 7.0},
                    {"Scheme Name": "Canara Rob Multi Cap Fund-Reg(G)", "ISIN": "INF760K01KR2", "Weight (%)": 7.0},
                    {"Scheme Name": "Kotak Multicap Fund-Reg(G)", "ISIN": "INF174KA1HS9", "Weight (%)": 7.0},
                    {"Scheme Name": "Bandhan Large & Mid Cap Fund-Reg(G)", "ISIN": "INF194K01524", "Weight (%)": 7.0},
                    {"Scheme Name": "Invesco India Focused Fund-Reg(G)", "ISIN": "INF205KA1189", "Weight (%)": 7.0},
                    {"Scheme Name": "Kotak Emerging Equity Fund(G)", "ISIN": "INF174K01DS9", "Weight (%)": 6.0},
                    {"Scheme Name": "SBI Infrastructure Fund-Reg(G)", "ISIN": "INF200K01CT2", "Weight (%)": 6.0},
                    {"Scheme Name": "Invesco India Smallcap Fund-Reg(G)", "ISIN": "INF205K011T7", "Weight (%)": 7.0},
                    {"Scheme Name": "HDFC Small Cap Fund-Reg(G)", "ISIN": "INF179KA1RZ8", "Weight (%)": 5.0},
                ])
                st.session_state["portfolio_buckets"] = {"Default Balanced Portfolio": default_df}
                st.session_state["active_bucket_name"] = "Default Balanced Portfolio"
                
            if "portfolio_results" not in st.session_state:
                st.session_state["portfolio_results"] = None
            if "last_active_bucket" not in st.session_state:
                st.session_state["last_active_bucket"] = st.session_state["active_bucket_name"]
            if "last_processed_file" not in st.session_state:
                st.session_state["last_processed_file"] = None
                
            # If active bucket changed, reset portfolio results to avoid stale data display
            if st.session_state["last_active_bucket"] != st.session_state["active_bucket_name"]:
                st.session_state["portfolio_results"] = None
                st.session_state["last_active_bucket"] = st.session_state["active_bucket_name"]

            # Bucket selection & management controls
            render_section_header("📁", "Bucket Management", "Create, switch, or delete fund buckets")
            col_m1, col_m2 = st.columns(2)
            with col_m1:
                bucket_options = list(st.session_state["portfolio_buckets"].keys())
                active_bucket = st.selectbox(
                    "Select Active Bucket", 
                    bucket_options, 
                    index=bucket_options.index(st.session_state["active_bucket_name"])
                )
                st.session_state["active_bucket_name"] = active_bucket
                
            with col_m2:
                new_bucket_name_raw = st.text_input("New Bucket Name", placeholder="e.g. My Conservative Portfolio")
                new_bucket_name = new_bucket_name_raw.strip()
                c_add, c_del = st.columns(2)
                with c_add:
                    if st.button("Create New Bucket", use_container_width=True):
                        if new_bucket_name and new_bucket_name not in st.session_state["portfolio_buckets"]:
                            curr_df = st.session_state["portfolio_buckets"][st.session_state["active_bucket_name"]].copy()
                            st.session_state["portfolio_buckets"][new_bucket_name] = curr_df
                            st.session_state["active_bucket_name"] = new_bucket_name
                            st.session_state["portfolio_results"] = None
                            st.rerun()
                with c_del:
                    if st.button("Delete Active Bucket", use_container_width=True, type="secondary"):
                        if len(st.session_state["portfolio_buckets"]) > 1:
                            del st.session_state["portfolio_buckets"][st.session_state["active_bucket_name"]]
                            st.session_state["active_bucket_name"] = list(st.session_state["portfolio_buckets"].keys())[0]
                            st.session_state["portfolio_results"] = None
                            st.rerun()
                        else:
                            st.warning("Cannot delete the last remaining bucket.")

            st.markdown('<hr class="panel-divider">', unsafe_allow_html=True)
            
            # 2. Investment Setup
            render_section_header("📅", "Investment Setup", "Pick when you started investing — the tracker runs to today automatically")
            col_a, col_b, col_c = st.columns(3)
            with col_a:
                start_date = st.date_input(
                    "Investment Start Date", 
                    value=(datetime.today() - timedelta(days=90)).date(),
                    min_value=datetime(2000, 1, 1).date(),
                    max_value=datetime.today().date()
                )
            with col_b:
                initial_amount_str = st.text_input("Investment Amount (₹)", value="100000")
                try:
                    initial_amount = float(initial_amount_str.replace(",", "").strip())
                    if initial_amount <= 0:
                        st.error("Investment Amount must be greater than 0.")
                        initial_amount = 100000.0
                except ValueError:
                    st.error("Please enter a valid numeric Investment Amount.")
                    initial_amount = 100000.0
            with col_c:
                skip_sunday = st.checkbox("Skip Sundays", value=True)

            st.markdown('<hr class="panel-divider">', unsafe_allow_html=True)
            
            # 3. Bucket Composition Editor
            render_section_header("📋", "Bucket Composition", f"Add/remove funds and set weights for '{st.session_state['active_bucket_name']}'")
            
            uploaded_file = st.file_uploader("Upload Bucket (Excel/CSV with ISIN and Weights)", type=["xlsx", "xls", "csv"])
            
            current_bucket_df = st.session_state["portfolio_buckets"][st.session_state["active_bucket_name"]]
            if uploaded_file is not None:
                file_key = f"processed_{uploaded_file.name}_{uploaded_file.size}"
                if st.session_state.get("last_processed_file") != file_key:
                    parsed_df = parse_bucket_input(uploaded_file=uploaded_file)
                    if not parsed_df.empty:
                        st.session_state["portfolio_buckets"][st.session_state["active_bucket_name"]] = parsed_df
                        current_bucket_df = parsed_df
                        st.session_state["portfolio_results"] = None
                        st.session_state["last_processed_file"] = file_key
                        st.success("Successfully loaded bucket from file!")
            else:
                st.session_state["last_processed_file"] = None
                    
            current_bucket_df_editor = current_bucket_df.copy()
            if "Weight (%)" in current_bucket_df_editor.columns:
                current_bucket_df_editor["Weight (%)"] = current_bucket_df_editor["Weight (%)"].astype(str)
                
            edited_df = st.data_editor(
                current_bucket_df_editor,
                column_config={
                    "Scheme Name": st.column_config.TextColumn("Scheme Name", width="large", help="Optional name for display"),
                    "ISIN": st.column_config.TextColumn("ISIN", required=True, help="Mutual Fund ISIN (Growth or Reinvestment)"),
                    "Weight (%)": st.column_config.TextColumn("Weight (%)", required=True, help="Percentage weight in portfolio (e.g. 8.9123456789)")
                },
                num_rows="dynamic",
                use_container_width=True,
                key=f"editor_{st.session_state['active_bucket_name']}"
            )
            
            if edited_df is not None:
                # Convert manual string weight input back to float, replacing invalid text with 0.0
                if "Weight (%)" in edited_df.columns:
                    edited_df["Weight (%)"] = pd.to_numeric(
                        edited_df["Weight (%)"].astype(str).str.replace("%", "").str.strip(), 
                        errors="coerce"
                    ).fillna(0.0)
                st.session_state["portfolio_buckets"][st.session_state["active_bucket_name"]] = edited_df
                
            weight_sum = edited_df["Weight (%)"].sum() if not edited_df.empty else 0
            weight_sum_str = f"{weight_sum:.10f}".rstrip('0').rstrip('.')
            if weight_sum != 100.0:
                st.info(f"⚖️ Current weights sum to **{weight_sum_str}%**. Weights will be automatically normalized to 100%.")
            else:
                st.success("⚖️ Weights sum to exactly **100%**!")

            st.markdown('<hr class="panel-divider">', unsafe_allow_html=True)
            
            # 4. Track / Refresh buttons
            col_btn1, col_btn2 = st.columns([3, 1])
            with col_btn1:
                track_btn = st.button("📊 Track Portfolio", type="primary", use_container_width=True)
            with col_btn2:
                refresh_btn = st.button("🔄 Refresh", use_container_width=True)
            
            if track_btn or refresh_btn:
                if start_date > datetime.today().date():
                    st.error("Start Date cannot be in the future.")
                else:
                    with st.spinner("Fetching latest NAVs from AMFI..."):
                        res_temp, err_temp = run_live_portfolio(
                            edited_df,
                            start_date,
                            initial_amount,
                            skip_sunday
                        )
                    if err_temp:
                        st.error(err_temp)
                        st.session_state["portfolio_results"] = None
                    else:
                        st.session_state["portfolio_results"] = res_temp
                        
            if st.session_state["portfolio_results"] is not None:
                res = st.session_state["portfolio_results"]
                met = res["metrics"]
                nav_date_display = met.get("NAV Date", "N/A")
                
                # ── Prominent LIVE current value ──────────────────────────
                gain_color = "#10b981" if met["Total Gain/Loss"] >= 0 else "#ef4444"
                today_color = "#10b981" if met["Today's Change"] >= 0 else "#ef4444"
                gain_sign = "+" if met["Total Gain/Loss"] >= 0 else ""
                today_sign = "+" if met["Today's Change"] >= 0 else ""
                
                st.markdown(
                    f"""
                    <div style="background: linear-gradient(135deg, #1a1a2e 0%, #16213e 100%); border-radius: 16px; padding: 28px 32px; margin-bottom: 20px; border: 1px solid rgba(212,175,55,0.3);">
                        <div style="display: flex; justify-content: space-between; align-items: flex-start; flex-wrap: wrap; gap: 16px;">
                            <div>
                                <div style="color: #94a3b8; font-size: 13px; text-transform: uppercase; letter-spacing: 1px;">Current Portfolio Value</div>
                                <div style="color: #f8fafc; font-size: 36px; font-weight: 700; margin: 4px 0;">₹{met['Current Value']:,.2f}</div>
                                <div style="color: {gain_color}; font-size: 16px; font-weight: 600;">{gain_sign}₹{met['Total Gain/Loss']:,.2f} ({gain_sign}{met['Total Return (%)']:.2f}%) total</div>
                            </div>
                            <div style="text-align: right;">
                                <div style="color: #94a3b8; font-size: 13px; text-transform: uppercase; letter-spacing: 1px;">Today's Change</div>
                                <div style="color: {today_color}; font-size: 24px; font-weight: 700; margin: 4px 0;">{today_sign}₹{met["Today's Change"]:,.2f}</div>
                                <div style="color: {today_color}; font-size: 14px;">{today_sign}{met["Today's Change (%)"]:.2f}%</div>
                                <div style="color: #64748b; font-size: 12px; margin-top: 6px;">NAV as of {nav_date_display}</div>
                            </div>
                        </div>
                        <div style="display: flex; gap: 24px; margin-top: 16px; padding-top: 16px; border-top: 1px solid rgba(255,255,255,0.1);">
                            <div><span style="color: #64748b; font-size: 12px;">Invested</span><br><span style="color: #e2e8f0; font-weight: 600;">₹{met['Initial Value']:,.2f}</span></div>
                            <div><span style="color: #64748b; font-size: 12px;">Max Drawdown</span><br><span style="color: #ef4444; font-weight: 600;">{met['Max Drawdown (%)']:.2f}%</span></div>
                            <div><span style="color: #64748b; font-size: 12px;">Best Day</span><br><span style="color: #10b981; font-weight: 600;">{met['Best Day Return (%)']:.2f}%</span></div>
                            <div><span style="color: #64748b; font-size: 12px;">Worst Day</span><br><span style="color: #ef4444; font-weight: 600;">{met['Worst Day Return (%)']:.2f}%</span></div>
                        </div>
                    </div>
                    """,
                    unsafe_allow_html=True
                )
                
                # ── Fund-wise holdings ───────────────────────────────────
                render_section_header("💼", "Fund-wise Holdings")
                st.dataframe(
                    res["composition"],
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        "Weight (%)": st.column_config.NumberColumn("Weight (%)", format="%.10f%%"),
                        "Units": st.column_config.NumberColumn("Units", format="%.10f"),
                        "Return (%)": st.column_config.NumberColumn("Return (%)", format="%.2f%%"),
                    }
                )
                
                # ── Portfolio Value Chart ─────────────────────────────────
                render_section_header("📈", "Portfolio Value Over Time")
                chart_df = res["tracker"].copy()
                chart_df["Date_dt"] = pd.to_datetime(chart_df["Date"], format="%d-%m-%Y")
                chart_df = chart_df.set_index("Date_dt")
                chart_cols = ["Total Portfolio Value (₹)"]
                if "Nifty 50 Valuation (₹)" in chart_df.columns:
                    chart_cols.append("Nifty 50 Valuation (₹)")
                if "Nifty 500 Valuation (₹)" in chart_df.columns:
                    chart_cols.append("Nifty 500 Valuation (₹)")
                st.line_chart(chart_df[chart_cols])
                
                # ── Daily History ─────────────────────────────────────────
                with st.expander("📋 Daily Valuation History", expanded=False):
                    st.dataframe(
                        res["tracker"],
                        use_container_width=True,
                        hide_index=True,
                        column_config={
                            "Daily Return (%)": st.column_config.NumberColumn("Daily Return (%)", format="%.2f%%"),
                            "Cumulative Return (%)": st.column_config.NumberColumn("Cumulative Return (%)", format="%.2f%%")
                        }
                    )
                
                # ── Export ─────────────────────────────────────────────────
                render_section_header("📥", "Export Report")
                with st.spinner("Generating Excel report..."):
                    excel_bytes = style_portfolio_excel(res)
                    
                st.download_button(
                    label="Download Portfolio Report (.xlsx)",
                    data=excel_bytes,
                    file_name=f"portfolio_{st.session_state['active_bucket_name']}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )





    with finance_panel("Market Snapshot"):
        with st.expander("Browse latest market snapshot", expanded=False):
            summary = summarize_families(nav_data)
            if summary.empty:
                st.info("No fund summaries are available.")
            else:
                browse = summary[["Family Name", "AMC Name", "Latest Scheme Name", "Latest Scheme Code", "Latest NAV", "Latest NAV Date", "Plan Types", "Option Types"]].copy()
                browse["Latest NAV Date"] = pd.to_datetime(browse["Latest NAV Date"], errors="coerce").dt.strftime("%d-%m-%Y")
                st.dataframe(browse.head(200), use_container_width=True, hide_index=True)

        render_section_header("⚡", "Quick Export", "Jump straight to Excel by fund name or Scheme Code")
        quick_query = st.text_input(
            "Enter full/partial fund name or exact Scheme Code",
            placeholder="e.g. HDFC Equity or 120439",
            label_visibility="collapsed",
        )
        if quick_query.strip():
            if st.button("Export NAV Excel", key="quick_export"):
                # Try direct Scheme Code match first
                q = quick_query.strip()
                found_family_key = None
                # exact numeric or string match against Scheme Code
                exact_code = nav_data[nav_data["Scheme Code"].astype(str).str.strip().eq(q)]
                if not exact_code.empty:
                    found_family_key = exact_code.iloc[0]["Family Key"]
                else:
                    # fuzzy search across families
                    matches = search_fund(nav_data, q, plan_filter=None, option_filter=None, limit=1)
                    if not matches.empty:
                        found_family_key = matches.iloc[0]["Family Key"]

                if not found_family_key:
                    st.warning("No matching fund found for export. Try a different name or the exact Scheme Code.")
                else:
                    latest_rows = filter_family_rows(nav_data, found_family_key)
                    history = load_historical_snapshots(found_family_key)

                    if latest_rows.empty and (history is None or history.empty):
                        st.info("No NAV rows available for the selected fund.")
                    else:
                        # Build Excel with two sheets when history exists
                        out = BytesIO()
                        with pd.ExcelWriter(out, engine="openpyxl") as writer:
                            if not latest_rows.empty:
                                export_latest = latest_rows.copy()
                                if "NAV Date" in export_latest.columns:
                                    export_latest["NAV Date"] = pd.to_datetime(export_latest["NAV Date"], errors="coerce").dt.strftime("%d-%m-%Y")
                                export_latest.to_excel(writer, index=False, sheet_name="Latest NAVs")
                            if history is not None and not history.empty:
                                export_hist = history.copy()
                                if "NAV Date" in export_hist.columns:
                                    export_hist["NAV Date"] = pd.to_datetime(export_hist["NAV Date"], errors="coerce").dt.strftime("%d-%m-%Y")
                                export_hist.to_excel(writer, index=False, sheet_name="History")
                        out.seek(0)
                        file_name = f"nav_export_{normalize_filename(quick_query)}.xlsx"
                        st.download_button(
                            "Download NAV Excel",
                            data=out.getvalue(),
                            file_name=file_name,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        )

    render_app_footer()


if __name__ == "__main__":
    main()
